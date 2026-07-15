import matplotlib
matplotlib.use("Agg")

import os, random
import numpy as np
import pandas as pd
from datetime import date, timedelta
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

rng = np.random.default_rng(42)
random.seed(42)

# ─── directories ───────────────────────────────────────────────────────────────
BASE   = r"D:\aml-transaction-monitoring-mis"
D_DIR  = os.path.join(BASE, "data")
E_DIR  = os.path.join(BASE, "excel")
I_DIR  = os.path.join(BASE, "images")

# ─── constants ─────────────────────────────────────────────────────────────────
START_DATE   = date(2026, 1, 1)
END_DATE     = date(2026, 4, 30)
N_DAYS       = (END_DATE - START_DATE).days + 1        # 120
N_CUSTOMERS  = 300
N_STRUCT     = 20           # customers in structuring cluster
THRESHOLD    = 10_000       # reporting threshold (USD)
AVG_DAILY    = 14           # avg transactions per day (normal)

# structuring window: week of Feb 9-15
STRUCT_START = date(2026, 2, 9)
STRUCT_END   = date(2026, 2, 15)

CHANNELS  = ["Wire Transfer", "ACH", "Cash Deposit", "Card", "Mobile App"]
COUNTRIES = ["US","US","US","US","US","US","MX","CA","NG","CN","RU","AE","US","US","US"]
MCCs      = ["Retail","Food & Beverage","Travel","Real Estate","Jewelry","Electronics",
             "Crypto Exchange","Money Services","Auto","Healthcare"]

ALERT_TYPES  = ["Structuring","Velocity","High-Risk Country","Round-Amount","Dormant-Reactivation"]
STATUSES     = ["Closed","Open","Escalated"]
KYC_STATUSES = ["Complete","Complete","Complete","Complete","Pending","Expired"]
RISK_RATINGS = ["Low","Low","Low","Medium","Medium","High"]
SEGMENTS     = ["Retail","Retail","SMB","SMB","Corporate","HNI"]

print("[1/6] Generating customer master ...")

# ─── customer risk profiles ────────────────────────────────────────────────────
cust_ids  = [f"C_{i:05d}" for i in range(1, N_CUSTOMERS + 1)]
struct_ids = cust_ids[:N_STRUCT]                   # first 20 are structuring customers

risk_rows = []
for cid in cust_ids:
    if cid in struct_ids:
        risk  = "High"
        kyc   = "Complete"
        seg   = "HNI"
    else:
        risk  = rng.choice(RISK_RATINGS)
        kyc   = rng.choice(KYC_STATUSES)
        seg   = rng.choice(SEGMENTS)
    age = int(rng.integers(6, 120))
    risk_rows.append({"Customer_ID": cid, "Risk_Rating": risk,
                      "KYC_Status": kyc, "Account_Age_Months": age,
                      "Customer_Segment": seg})

# seed 20 customers without profiles (orphaned FKs in transactions)
orphan_ids = [f"C_ORPHAN_{i:03d}" for i in range(1, 21)]

risk_df = pd.DataFrame(risk_rows)
risk_df.to_csv(os.path.join(D_DIR, "customer_risk_profile.csv"), index=False)
print(f"   -> customer_risk_profile.csv  ({len(risk_df)} rows)")

# ─── transactions ──────────────────────────────────────────────────────────────
print("[2/6] Generating transactions ...")

trans_rows = []
tx_id = 1

# normal transactions
all_dates = [START_DATE + timedelta(days=d) for d in range(N_DAYS)]
for d in all_dates:
    n = int(rng.poisson(AVG_DAILY))
    for _ in range(n):
        # small chance of orphaned customer
        if rng.random() < 0.03:
            cid = rng.choice(orphan_ids)
        else:
            cid = rng.choice(cust_ids[N_STRUCT:])   # non-structuring normal customers
        amt  = round(float(rng.lognormal(7.5, 1.2)), 2)
        amt  = min(amt, 200_000)
        chan = rng.choice(CHANNELS)
        ctry = rng.choice(COUNTRIES)
        mcc  = rng.choice(MCCs)
        trans_rows.append({"Transaction_ID": f"TXN_{tx_id:06d}", "Customer_ID": cid,
                           "Date": str(d), "Amount": amt,
                           "Channel": chan, "Country": ctry,
                           "Merchant_Category": mcc})
        tx_id += 1

# structuring cluster: Feb 9-15 — 20 customers × 5 transactions each
struct_dates = [STRUCT_START + timedelta(days=d) for d in range(7)]
for cid in struct_ids:
    n_txn = int(rng.integers(4, 7))
    chosen_dates = rng.choice(struct_dates, size=n_txn, replace=True)
    for d in chosen_dates:
        amt  = round(float(rng.uniform(9_000, 9_999)), 2)
        chan = rng.choice(["Cash Deposit", "Wire Transfer"])
        ctry = "US"
        mcc  = "Money Services"
        trans_rows.append({"Transaction_ID": f"TXN_{tx_id:06d}", "Customer_ID": cid,
                           "Date": str(d), "Amount": amt,
                           "Channel": chan, "Country": ctry,
                           "Merchant_Category": mcc})
        tx_id += 1

# 5 near-duplicate transactions (same customer, date, amount — DQ issue)
for k in range(5):
    base = trans_rows[k * 100]
    dup  = base.copy()
    dup["Transaction_ID"] = f"TXN_{tx_id:06d}"
    trans_rows.append(dup)
    tx_id += 1

trans_df = pd.DataFrame(trans_rows).sort_values("Date").reset_index(drop=True)
trans_df.to_csv(os.path.join(D_DIR, "transactions.csv"), index=False)
print(f"   -> transactions.csv  ({len(trans_df)} rows)")

N_TRANS = len(trans_df)

# ─── alerts ────────────────────────────────────────────────────────────────────
print("[3/6] Generating alerts log ...")

all_txn_ids   = set(trans_df["Transaction_ID"].tolist())
struct_txn_ids = trans_df[trans_df["Customer_ID"].isin(struct_ids)]["Transaction_ID"].tolist()

alert_rows = []
alrt_id = 1

# structuring alerts for every structuring transaction
for tid in struct_txn_ids:
    d    = trans_df.loc[trans_df["Transaction_ID"] == tid, "Date"].values[0]
    d    = pd.to_datetime(d).date()
    res  = int(rng.integers(1, 5))
    stat = "Closed" if res <= 3 else "Open"
    alert_rows.append({"Alert_ID": f"ALT_{alrt_id:05d}", "Transaction_ID": tid,
                        "Alert_Type": "Structuring",
                        "Alert_Date": str(d + timedelta(days=1)),
                        "Status": stat, "Resolution_Time_Days": res if stat == "Closed" else ""})
    alrt_id += 1

# background alerts (~12% of non-structuring transactions)
non_struct_txns = trans_df[~trans_df["Customer_ID"].isin(struct_ids)]["Transaction_ID"].tolist()
rng.shuffle(non_struct_txns)
n_bg_alerts = int(len(non_struct_txns) * 0.12)
for tid in non_struct_txns[:n_bg_alerts]:
    d    = trans_df.loc[trans_df["Transaction_ID"] == tid, "Date"].values[0]
    d    = pd.to_datetime(d).date()
    atype = rng.choice(ALERT_TYPES[1:])   # no structuring for background
    res  = int(rng.integers(1, 15))
    stat = rng.choice(["Closed", "Closed", "Open", "Escalated"])
    alert_rows.append({"Alert_ID": f"ALT_{alrt_id:05d}", "Transaction_ID": tid,
                        "Alert_Type": atype,
                        "Alert_Date": str(d + timedelta(days=int(rng.integers(0, 3)))),
                        "Status": stat, "Resolution_Time_Days": res if stat == "Closed" else ""})
    alrt_id += 1

# 10 orphaned alerts (Transaction_IDs not in transactions — DQ issue)
for k in range(10):
    alert_rows.append({"Alert_ID": f"ALT_{alrt_id:05d}",
                        "Transaction_ID": f"TXN_ORPHAN_{k:03d}",
                        "Alert_Type": rng.choice(ALERT_TYPES),
                        "Alert_Date": str(START_DATE + timedelta(days=int(rng.integers(0, 90)))),
                        "Status": "Open", "Resolution_Time_Days": ""})
    alrt_id += 1

alerts_df = pd.DataFrame(alert_rows)
alerts_df.to_csv(os.path.join(D_DIR, "alerts_log.csv"), index=False)
N_ALERTS = len(alerts_df)
print(f"   -> alerts_log.csv  ({N_ALERTS} rows)")

# ─── Excel workbook ────────────────────────────────────────────────────────────
print("[4/6] Building Excel workbook ...")

wb = Workbook()
wb.remove(wb.active)

# colour palette
HDR_FILL   = PatternFill("solid", fgColor="1B3A6B")  # dark navy
HDR2_FILL  = PatternFill("solid", fgColor="2E5E9E")  # mid navy
KPI_FILL   = PatternFill("solid", fgColor="F0F4FB")  # light blue
RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
YEL_FILL   = PatternFill("solid", fgColor="FFEB9C")
GRN_FILL   = PatternFill("solid", fgColor="C6EFCE")
WHT_FILL   = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL   = PatternFill("solid", fgColor="EEF2FA")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
HDR2_FONT  = Font(bold=True, color="FFFFFF", size=9)
TTL_FONT   = Font(bold=True, color="1B3A6B", size=11)
SMALL_FONT = Font(size=9)
BOLD_FONT  = Font(bold=True, size=9)

thin  = Side(style="thin",   color="B0BAC9")
thick = Side(style="medium", color="1B3A6B")
def borders(top=None, left=None, right=None, bottom=None):
    return Border(top=top or thin, left=left or thin,
                  right=right or thin, bottom=bottom or thin)

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT   = Alignment(horizontal="right",  vertical="center")

def col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ── helper: write header row ───────────────────────────────────────────────────
def write_header(ws, row, cols, fill=HDR_FILL, font=HDR_FONT):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = CENTER
        cell.border = borders()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Raw_Transactions
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Raw_Transactions")
ws1.sheet_view.showGridLines = False

T_COLS = list(trans_df.columns)
write_header(ws1, 1, T_COLS)
for r, row_data in enumerate(trans_df.itertuples(index=False), 2):
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.fill  = ALT_FILL if r % 2 == 0 else WHT_FILL
        cell.font  = SMALL_FONT
        cell.border = borders()
        if T_COLS[c-1] == "Amount":
            cell.number_format = '#,##0.00'

for w, col in zip([14, 10, 12, 14, 14, 10, 18], "ABCDEFG"):
    col_width(ws1, col, w)
ws1.row_dimensions[1].height = 22
freeze(ws1, "A2")
print("   -> Sheet: Raw_Transactions")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Raw_Risk_Profiles
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Raw_Risk_Profiles")
ws2.sheet_view.showGridLines = False

R_COLS = list(risk_df.columns)
write_header(ws2, 1, R_COLS)
for r, row_data in enumerate(risk_df.itertuples(index=False), 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill  = ALT_FILL if r % 2 == 0 else WHT_FILL
        cell.font  = SMALL_FONT
        cell.border = borders()

for w, col in zip([12, 13, 13, 18, 16], "ABCDE"):
    col_width(ws2, col, w)
ws2.row_dimensions[1].height = 22
freeze(ws2, "A2")

# conditional formatting: risk rating colour-coding
ws2.conditional_formatting.add(
    f"B2:B{len(risk_df)+1}",
    CellIsRule(operator="equal", formula=['"High"'],   fill=RED_FILL, font=Font(bold=True, color="9C0006")))
ws2.conditional_formatting.add(
    f"B2:B{len(risk_df)+1}",
    CellIsRule(operator="equal", formula=['"Medium"'], fill=YEL_FILL, font=Font(bold=True, color="7D6608")))
ws2.conditional_formatting.add(
    f"B2:B{len(risk_df)+1}",
    CellIsRule(operator="equal", formula=['"Low"'],    fill=GRN_FILL, font=Font(bold=True, color="276221")))
print("   -> Sheet: Raw_Risk_Profiles")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Raw_Alerts
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Raw_Alerts")
ws3.sheet_view.showGridLines = False

A_COLS = list(alerts_df.columns)
write_header(ws3, 1, A_COLS)
for r, row_data in enumerate(alerts_df.itertuples(index=False), 2):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=str(val) if val == val else "")
        cell.fill  = ALT_FILL if r % 2 == 0 else WHT_FILL
        cell.font  = SMALL_FONT
        cell.border = borders()

for w, col in zip([12, 14, 22, 13, 11, 20], "ABCDEF"):
    col_width(ws3, col, w)
ws3.row_dimensions[1].height = 22
freeze(ws3, "A2")
print("   -> Sheet: Raw_Alerts")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Consolidated_View
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Consolidated_View")
ws4.sheet_view.showGridLines = False

C_COLS = ["Transaction_ID","Customer_ID","Date","Amount","Channel","Country",
          "Merchant_Category","Risk_Rating","KYC_Status","Account_Age_Months",
          "Customer_Segment","Alert_ID","Alert_Type","Alert_Date",
          "Resolution_Days","Data_Quality_Flag"]
write_header(ws4, 1, C_COLS)

N = N_TRANS    # number of transaction rows
for i in range(2, N + 2):
    r = i
    # columns A-G: pull from Raw_Transactions (same row index)
    ws4.cell(r, 1, f"=Raw_Transactions!A{r}")
    ws4.cell(r, 2, f"=Raw_Transactions!B{r}")
    ws4.cell(r, 3, f"=Raw_Transactions!C{r}")
    ws4.cell(r, 4, f"=Raw_Transactions!D{r}")
    ws4.cell(r, 5, f"=Raw_Transactions!E{r}")
    ws4.cell(r, 6, f"=Raw_Transactions!F{r}")
    ws4.cell(r, 7, f"=Raw_Transactions!G{r}")

    # XLOOKUP: Risk_Rating from Raw_Risk_Profiles by Customer_ID
    ws4.cell(r, 8,  f'=XLOOKUP(B{r},Raw_Risk_Profiles!$A:$A,Raw_Risk_Profiles!$B:$B,"Not Found")')
    # XLOOKUP: KYC_Status
    ws4.cell(r, 9,  f'=XLOOKUP(B{r},Raw_Risk_Profiles!$A:$A,Raw_Risk_Profiles!$C:$C,"Missing")')
    # XLOOKUP: Account_Age_Months
    ws4.cell(r, 10, f'=XLOOKUP(B{r},Raw_Risk_Profiles!$A:$A,Raw_Risk_Profiles!$D:$D,"")')
    # INDEX-MATCH: Customer_Segment (alternate technique)
    ws4.cell(r, 11, f'=IFERROR(INDEX(Raw_Risk_Profiles!$E:$E,MATCH(B{r},Raw_Risk_Profiles!$A:$A,0)),"Unknown")')
    # XLOOKUP: Alert_ID (matched on Transaction_ID)
    ws4.cell(r, 12, f'=IFERROR(XLOOKUP(A{r},Raw_Alerts!$B:$B,Raw_Alerts!$A:$A),"No Alert")')
    # XLOOKUP: Alert_Type
    ws4.cell(r, 13, f'=IFERROR(XLOOKUP(A{r},Raw_Alerts!$B:$B,Raw_Alerts!$C:$C),"")')
    # XLOOKUP: Alert_Date
    ws4.cell(r, 14, f'=IFERROR(XLOOKUP(A{r},Raw_Alerts!$B:$B,Raw_Alerts!$D:$D),"")')
    # XLOOKUP: Resolution_Days
    ws4.cell(r, 15, f'=IFERROR(XLOOKUP(A{r},Raw_Alerts!$B:$B,Raw_Alerts!$F:$F),"")')
    # 4-level nested IF: Data_Quality_Flag
    ws4.cell(r, 16,
        f'=IF(H{r}="Not Found","CRITICAL: No Risk Profile",'
        f'IF(I{r}="Missing","WARN: Missing KYC",'
        f'IF(I{r}="Expired","WARN: KYC Expired",'
        f'IF(I{r}="Pending","INFO: KYC Pending","OK"))))')

    # styling
    for c in range(1, 17):
        cell = ws4.cell(r, c)
        cell.fill   = ALT_FILL if r % 2 == 0 else WHT_FILL
        cell.font   = SMALL_FONT
        cell.border = borders()
        if c == 4:
            cell.number_format = '#,##0.00'

col_widths = [14, 10, 12, 12, 13, 8, 17, 13, 12, 15, 14, 10, 22, 12, 14, 25]
for w, col in zip(col_widths, [get_column_letter(c) for c in range(1, 17)]):
    col_width(ws4, col, w)
ws4.row_dimensions[1].height = 22
freeze(ws4, "A2")

# CF on Data_Quality_Flag column (P)
DQ_RANGE = f"P2:P{N+1}"
ws4.conditional_formatting.add(DQ_RANGE,
    FormulaRule(formula=[f'LEFT(P2,8)="CRITICAL"'], fill=RED_FILL, font=Font(bold=True, color="9C0006")))
ws4.conditional_formatting.add(DQ_RANGE,
    FormulaRule(formula=[f'LEFT(P2,4)="WARN"'],     fill=YEL_FILL, font=Font(bold=True, color="7D6608")))
ws4.conditional_formatting.add(DQ_RANGE,
    FormulaRule(formula=[f'LEFT(P2,4)="INFO"'],     fill=PatternFill("solid", fgColor="DEEBF7"),
                font=Font(bold=False, color="1F4E79")))
ws4.conditional_formatting.add(DQ_RANGE,
    FormulaRule(formula=[f'P2="OK"'],               fill=GRN_FILL))
print("   -> Sheet: Consolidated_View")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Weekly_MIS_Report
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Weekly_MIS_Report")
ws5.sheet_view.showGridLines = False

# compute week bins
weeks = []
d = START_DATE
while d <= END_DATE:
    wend = min(d + timedelta(days=6), END_DATE)
    weeks.append((d, wend))
    d = wend + timedelta(days=1)

W_COLS = ["Week #","Week Start","Week End","Total Transactions","Total Alerts",
          "Structuring Alerts","Velocity Alerts","High-Risk Country","Round-Amount",
          "Dormant-Reactivation","Open Alerts","Escalated Alerts","Closed Alerts",
          "Avg Resolution Days","SLA Breach (>7d)","Alert Rate %"]
write_header(ws5, 1, W_COLS)

for idx, (wstart, wend) in enumerate(weeks, 2):
    ws5.cell(idx, 1, idx - 1)
    ws5.cell(idx, 2, str(wstart))
    ws5.cell(idx, 3, str(wend))
    ws5.cell(idx, 4,
        f'=COUNTIFS(Raw_Transactions!$C:$C,">="&B{idx},'
        f'Raw_Transactions!$C:$C,"<="&C{idx})')
    ws5.cell(idx, 5,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx})')
    for col_off, atype in enumerate(["Structuring","Velocity","High-Risk Country",
                                      "Round-Amount","Dormant-Reactivation"], 6):
        ws5.cell(idx, col_off,
            f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
            f'Raw_Alerts!$D:$D,"<="&C{idx},'
            f'Raw_Alerts!$C:$C,"{atype}")')
    ws5.cell(idx, 11,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Open")')
    ws5.cell(idx, 12,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Escalated")')
    ws5.cell(idx, 13,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Closed")')
    ws5.cell(idx, 14,
        f'=IFERROR(AVERAGEIFS(Raw_Alerts!$F:$F,'
        f'Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Closed"),0)')
    ws5.cell(idx, 15,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$F:$F,">"&7)')
    ws5.cell(idx, 16,
        f'=IFERROR(E{idx}/D{idx},0)')

    for c in range(1, 17):
        cell = ws5.cell(idx, c)
        cell.fill   = ALT_FILL if idx % 2 == 0 else WHT_FILL
        cell.font   = SMALL_FONT
        cell.border = borders()
        if c == 14:
            cell.number_format = '0.0'
        if c == 16:
            cell.number_format = '0.0%'

for w, col in zip([8,13,11,18,13,17,13,18,13,20,11,14,12,16,15,12],
                  [get_column_letter(c) for c in range(1, 17)]):
    col_width(ws5, col, w)
ws5.row_dimensions[1].height = 30
freeze(ws5, "D2")

# highlight week 6 (structuring spike)
STRUCT_ROW = 7  # week 6
for c in range(1, 17):
    ws5.cell(STRUCT_ROW, c).fill = PatternFill("solid", fgColor="FFF2CC")
    ws5.cell(STRUCT_ROW, c).font = Font(bold=True, size=9)

# CF: SLA breach column
N_WEEKS = len(weeks)
ws5.conditional_formatting.add(
    f"O2:O{N_WEEKS+1}",
    CellIsRule(operator="greaterThan", formula=["0"], fill=RED_FILL))
print("   -> Sheet: Weekly_MIS_Report")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: Monthly_MIS_Report
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Monthly_MIS_Report")
ws6.sheet_view.showGridLines = False

months_info = [
    ("January 2026",  "2026-01-01", "2026-01-31"),
    ("February 2026", "2026-02-01", "2026-02-28"),
    ("March 2026",    "2026-03-01", "2026-03-31"),
    ("April 2026",    "2026-04-01", "2026-04-30"),
]

M_COLS = ["Month","Month Start","Month End","Total Transactions","Total Alerts",
          "Structuring Alerts","Velocity Alerts","Open Alerts","Escalated Alerts",
          "Closed Alerts","Avg Resolution Days","SLA Breach Count",
          "Alert Rate %","MoM Alert Change","MoM SLA Breach Change"]
write_header(ws6, 1, M_COLS)

for idx, (mname, ms, me) in enumerate(months_info, 2):
    ws6.cell(idx, 1, mname)
    ws6.cell(idx, 2, ms)
    ws6.cell(idx, 3, me)
    ws6.cell(idx, 4,
        f'=COUNTIFS(Raw_Transactions!$C:$C,">="&B{idx},'
        f'Raw_Transactions!$C:$C,"<="&C{idx})')
    ws6.cell(idx, 5,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx})')
    ws6.cell(idx, 6,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$C:$C,"Structuring")')
    ws6.cell(idx, 7,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$C:$C,"Velocity")')
    ws6.cell(idx, 8,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Open")')
    ws6.cell(idx, 9,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Escalated")')
    ws6.cell(idx, 10,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Closed")')
    ws6.cell(idx, 11,
        f'=IFERROR(AVERAGEIFS(Raw_Alerts!$F:$F,'
        f'Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$E:$E,"Closed"),0)')
    ws6.cell(idx, 12,
        f'=COUNTIFS(Raw_Alerts!$D:$D,">="&B{idx},'
        f'Raw_Alerts!$D:$D,"<="&C{idx},'
        f'Raw_Alerts!$F:$F,">"&7)')
    ws6.cell(idx, 13,
        f'=IFERROR(E{idx}/D{idx},0)')
    if idx == 2:
        ws6.cell(idx, 14, "N/A")
        ws6.cell(idx, 15, "N/A")
    else:
        ws6.cell(idx, 14,
            f'=IFERROR((E{idx}-E{idx-1})/E{idx-1},0)')
        ws6.cell(idx, 15,
            f'=IFERROR((L{idx}-L{idx-1})/L{idx-1},0)')

    for c in range(1, 16):
        cell = ws6.cell(idx, c)
        cell.fill   = ALT_FILL if idx % 2 == 0 else WHT_FILL
        cell.font   = SMALL_FONT
        cell.border = borders()
        if c == 11:
            cell.number_format = '0.0'
        if c in (13, 14, 15):
            if idx > 2 or c == 13:
                cell.number_format = '0.0%'

for w, col in zip([16,13,11,18,13,17,13,11,14,12,16,14,12,16,18],
                  [get_column_letter(c) for c in range(1, 16)]):
    col_width(ws6, col, w)
ws6.row_dimensions[1].height = 30
freeze(ws6, "D2")
print("   -> Sheet: Monthly_MIS_Report")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7: Ad_Hoc_Scorecard
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Ad_Hoc_Scorecard")
ws7.sheet_view.showGridLines = False

ws7.merge_cells("A1:F1")
title_cell = ws7.cell(1, 1, "AML Alert Scorecard — Filter by Alert Type")
title_cell.fill = HDR_FILL
title_cell.font = Font(bold=True, color="FFFFFF", size=12)
title_cell.alignment = CENTER

ws7.cell(3, 1, "Select Alert Type:").font = Font(bold=True, size=10)
ws7.cell(3, 2, "All").font = Font(bold=True, size=10, color="1B3A6B")

# data validation dropdown
dv = DataValidation(type="list", formula1='"All,Structuring,Velocity,High-Risk Country,Round-Amount,Dormant-Reactivation"', allow_blank=False)
dv.sqref = "B3"
ws7.add_data_validation(dv)

score_cols = ["Metric", "Q1 (Jan)", "Q2 Feb", "Q3 Mar", "Q4 Apr", "Total"]
write_header(ws7, 5, score_cols, fill=HDR2_FILL)

score_metrics = [
    ("Alert Count",
     lambda m, ms, me: f'=IF($B$3="All",COUNTIFS(Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"),COUNTIFS(Raw_Alerts!$C:$C,$B$3,Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"))'),
    ("Open Alerts",
     lambda m, ms, me: f'=IF($B$3="All",COUNTIFS(Raw_Alerts!$E:$E,"Open",Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"),COUNTIFS(Raw_Alerts!$C:$C,$B$3,Raw_Alerts!$E:$E,"Open",Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"))'),
    ("Escalated Alerts",
     lambda m, ms, me: f'=IF($B$3="All",COUNTIFS(Raw_Alerts!$E:$E,"Escalated",Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"),COUNTIFS(Raw_Alerts!$C:$C,$B$3,Raw_Alerts!$E:$E,"Escalated",Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"))'),
    ("Closed Alerts",
     lambda m, ms, me: f'=IF($B$3="All",COUNTIFS(Raw_Alerts!$E:$E,"Closed",Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"),COUNTIFS(Raw_Alerts!$C:$C,$B$3,Raw_Alerts!$E:$E,"Closed",Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"))'),
    ("SLA Breaches (>7d)",
     lambda m, ms, me: f'=IF($B$3="All",COUNTIFS(Raw_Alerts!$F:$F,">"&7,Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"),COUNTIFS(Raw_Alerts!$C:$C,$B$3,Raw_Alerts!$F:$F,">"&7,Raw_Alerts!$D:$D,">="&"{ms}",Raw_Alerts!$D:$D,"<="&"{me}"))'),
]

month_ranges = [("Jan","2026-01-01","2026-01-31"),
                ("Feb","2026-02-01","2026-02-28"),
                ("Mar","2026-03-01","2026-03-31"),
                ("Apr","2026-04-01","2026-04-30")]

for row_off, (metric_name, formula_fn) in enumerate(score_metrics, 6):
    ws7.cell(row_off, 1, metric_name).font = BOLD_FONT
    ws7.cell(row_off, 1).border = borders()
    for col_off, (mname, ms, me) in enumerate(month_ranges, 2):
        cell = ws7.cell(row_off, col_off)
        cell.value  = formula_fn(mname, ms, me)
        cell.font   = SMALL_FONT
        cell.border = borders()
        cell.fill   = ALT_FILL if row_off % 2 == 0 else WHT_FILL
    # total
    ws7.cell(row_off, 6, f"=SUM(B{row_off}:E{row_off})").font = BOLD_FONT
    ws7.cell(row_off, 6).border = borders()

for w, col in zip([25, 14, 14, 14, 14, 14], "ABCDEF"):
    col_width(ws7, col, w)
ws7.row_dimensions[1].height = 28
ws7.row_dimensions[5].height = 22
print("   -> Sheet: Ad_Hoc_Scorecard")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8: Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Dashboard")
ws8.sheet_view.showGridLines = False

# title banner
ws8.merge_cells("A1:N1")
t = ws8.cell(1, 1, "AML TRANSACTION MONITORING MIS DASHBOARD  |  Jan-Apr 2026")
t.fill = HDR_FILL
t.font = Font(bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws8.row_dimensions[1].height = 32

ws8.merge_cells("A2:N2")
s = ws8.cell(2, 1, "American Express — AML Compliance  |  Reporting Period: Q1 2026 (120 days)")
s.fill = PatternFill("solid", fgColor="2E5E9E")
s.font = Font(italic=True, color="FFFFFF", size=10)
s.alignment = CENTER

# KPI cards — two separate merges per card: row 4 = label, row 5 = value
from openpyxl.utils import column_index_from_string
kpis = [
    ("A4:C4", "A5:C5", "TOTAL TRANSACTIONS",
     f"=COUNTA(Raw_Transactions!$A:$A)-1", "#,##0"),
    ("D4:F4", "D5:F5", "TOTAL ALERTS",
     f"=COUNTA(Raw_Alerts!$A:$A)-1", "#,##0"),
    ("G4:I4", "G5:I5", "OPEN / ESCALATED",
     f'=COUNTIF(Raw_Alerts!$E:$E,"Open")+COUNTIF(Raw_Alerts!$E:$E,"Escalated")', "#,##0"),
    ("J4:L4", "J5:L5", "STRUCTURING ALERTS",
     f'=COUNTIF(Raw_Alerts!$C:$C,"Structuring")', "#,##0"),
    ("M4:N4", "M5:N5", "HIGH-RISK CUSTOMERS",
     f'=COUNTIF(Raw_Risk_Profiles!$B:$B,"High")', "#,##0"),
]
for lbl_range, val_range, label, formula, fmt in kpis:
    lbl_start = lbl_range.split(":")[0]
    lbl_col = column_index_from_string(''.join(c for c in lbl_start if c.isalpha()))
    ws8.merge_cells(lbl_range)
    sc = ws8.cell(4, lbl_col)
    sc.value = label
    sc.fill  = HDR2_FILL
    sc.font  = Font(bold=True, color="FFFFFF", size=9)
    sc.alignment = CENTER

    ws8.merge_cells(val_range)
    vc = ws8.cell(5, lbl_col)
    vc.value = formula
    vc.fill  = KPI_FILL
    vc.font  = Font(bold=True, size=18, color="1B3A6B")
    vc.alignment = CENTER
    vc.number_format = fmt

ws8.row_dimensions[4].height = 18
ws8.row_dimensions[5].height = 36

# weekly trend table (rows 8-28)
ws8.cell(8, 1, "Weekly Alert Trend (Structuring Highlighted)").font = TTL_FONT
ws8.merge_cells("A8:F8")

trend_cols = ["Week","Wk Start","Total Alerts","Structuring","Velocity","Open/Esc"]
write_header(ws8, 9, trend_cols, fill=HDR2_FILL)
for idx in range(len(weeks)):
    src_row = idx + 2
    r = 10 + idx
    ws8.cell(r, 1, f"=Weekly_MIS_Report!A{src_row}")
    ws8.cell(r, 2, f"=Weekly_MIS_Report!B{src_row}")
    ws8.cell(r, 3, f"=Weekly_MIS_Report!E{src_row}")
    ws8.cell(r, 4, f"=Weekly_MIS_Report!F{src_row}")
    ws8.cell(r, 5, f"=Weekly_MIS_Report!G{src_row}")
    ws8.cell(r, 6, f"=Weekly_MIS_Report!K{src_row}+Weekly_MIS_Report!L{src_row}")
    for c in range(1, 7):
        ws8.cell(r, c).font   = SMALL_FONT
        ws8.cell(r, c).border = borders()
        ws8.cell(r, c).fill   = ALT_FILL if r % 2 == 0 else WHT_FILL

# structuring week highlight in dashboard table
ws8.conditional_formatting.add(
    f"D10:D{10+len(weeks)-1}",
    CellIsRule(operator="greaterThan", formula=["20"], fill=RED_FILL,
               font=Font(bold=True, color="9C0006")))

# monthly summary (rows 8-13, columns H onwards)
ws8.cell(8, 8, "Monthly MIS Summary").font = TTL_FONT
ws8.merge_cells("H8:N8")
mon_cols = ["Month","Txns","Alerts","Struct","SLA Breach","Alert%","MoM %"]
write_header(ws8, 9, mon_cols, fill=HDR2_FILL)
for c, v in enumerate(mon_cols, 8):
    ws8.cell(9, c).value = v
    ws8.cell(9, c).fill = HDR2_FILL
    ws8.cell(9, c).font = HDR2_FONT
    ws8.cell(9, c).alignment = CENTER

for midx, (mname, _, _) in enumerate(months_info, 10):
    src = midx - 8
    ws8.cell(midx, 8,  f"=Monthly_MIS_Report!A{src}")
    ws8.cell(midx, 9,  f"=Monthly_MIS_Report!D{src}")
    ws8.cell(midx, 10, f"=Monthly_MIS_Report!E{src}")
    ws8.cell(midx, 11, f"=Monthly_MIS_Report!F{src}")
    ws8.cell(midx, 12, f"=Monthly_MIS_Report!L{src}")
    ws8.cell(midx, 13, f"=Monthly_MIS_Report!M{src}")
    ws8.cell(midx, 14, f"=Monthly_MIS_Report!N{src}" if src > 2 else "N/A")
    for c in range(8, 15):
        ws8.cell(midx, c).font   = SMALL_FONT
        ws8.cell(midx, c).border = borders()
        ws8.cell(midx, c).fill   = ALT_FILL if midx % 2 == 0 else WHT_FILL
    ws8.cell(midx, 13).number_format = "0.0%"
    if src > 2:
        ws8.cell(midx, 14).number_format = "0.0%"

# DQ flag summary (rows 16-22)
dq_start = 10 + len(weeks) + 2
ws8.cell(dq_start, 1, "Data Quality Flag Summary").font = TTL_FONT
ws8.merge_cells(f"A{dq_start}:F{dq_start}")
dq_items = [
    ("CRITICAL: No Risk Profile", f'=COUNTIF(Consolidated_View!$P:$P,"CRITICAL*")'),
    ("WARN: Missing KYC",         f'=COUNTIF(Consolidated_View!$P:$P,"WARN: Missing KYC")'),
    ("WARN: KYC Expired",         f'=COUNTIF(Consolidated_View!$P:$P,"WARN: KYC Expired")'),
    ("INFO: KYC Pending",         f'=COUNTIF(Consolidated_View!$P:$P,"INFO*")'),
    ("OK — No Issues",            f'=COUNTIF(Consolidated_View!$P:$P,"OK")'),
]
write_header(ws8, dq_start + 1, ["Flag Type", "Count"], fill=HDR2_FILL)
ws8.merge_cells(f"B{dq_start+1}:F{dq_start+1}")
for off, (flag, formula) in enumerate(dq_items, dq_start + 2):
    ws8.cell(off, 1, flag).font = SMALL_FONT
    ws8.cell(off, 1).border = borders()
    c = ws8.cell(off, 2, formula)
    ws8.merge_cells(f"B{off}:F{off}")
    c.font  = BOLD_FONT
    c.border = borders()
    c.fill  = ALT_FILL if off % 2 == 0 else WHT_FILL

for col, w in zip("ABCDEFGHIJKLMN", [8,12,12,13,12,12, 16,12,12,14,13,10,10,10]):
    col_width(ws8, col, w)

print("   -> Sheet: Dashboard")

wb.save(os.path.join(E_DIR, "aml_transaction_monitoring_mis.xlsx"))
print("   -> Workbook saved.")

# ─── matplotlib images ─────────────────────────────────────────────────────────
print("[5/6] Generating images ...")

# compute actuals for charts
trans_df["Date_dt"] = pd.to_datetime(trans_df["Date"])
alerts_df["Alert_Date_dt"] = pd.to_datetime(alerts_df["Alert_Date"], errors="coerce")

# weekly alert counts
week_labels, week_alerts, week_struct = [], [], []
for wstart, wend in weeks:
    ws_dt = pd.Timestamp(wstart)
    we_dt = pd.Timestamp(wend)
    total = len(alerts_df[(alerts_df["Alert_Date_dt"] >= ws_dt) &
                           (alerts_df["Alert_Date_dt"] <= we_dt)])
    struct = len(alerts_df[(alerts_df["Alert_Date_dt"] >= ws_dt) &
                            (alerts_df["Alert_Date_dt"] <= we_dt) &
                            (alerts_df["Alert_Type"] == "Structuring")])
    week_labels.append(f"W{wstart.isocalendar().week}")
    week_alerts.append(total)
    week_struct.append(struct)

# ── image 1: dashboard overview ────────────────────────────────────────────────
fig, axes = plt.subplots(2, 2, figsize=(14, 9))
fig.patch.set_facecolor("#F8FAFF")
fig.suptitle("AML Transaction Monitoring — MIS Dashboard  |  Jan–Apr 2026",
             fontsize=14, fontweight="bold", color="#1B3A6B", y=0.98)

# KPI cards panel
ax = axes[0, 0]
ax.set_facecolor("#1B3A6B")
ax.axis("off")
n_total = len(trans_df)
n_alerts = len(alerts_df)
n_open   = len(alerts_df[alerts_df["Status"].isin(["Open","Escalated"])])
n_struct = len(alerts_df[alerts_df["Alert_Type"] == "Structuring"])
kpi_vals = [
    ("TOTAL TRANSACTIONS", f"{n_total:,}"),
    ("TOTAL ALERTS", f"{n_alerts:,}"),
    ("OPEN / ESCALATED", f"{n_open:,}"),
    ("STRUCTURING ALERTS", f"{n_struct:,}"),
]
for i, (label, val) in enumerate(kpi_vals):
    x = 0.25 * i + 0.03
    ax.add_patch(mpatches.FancyBboxPatch((x, 0.1), 0.21, 0.80,
        boxstyle="round,pad=0.02", facecolor="#2E5E9E", edgecolor="white", linewidth=1))
    ax.text(x+0.105, 0.65, label, ha="center", va="center", fontsize=6.5,
            color="#AECBF5", fontweight="bold")
    ax.text(x+0.105, 0.35, val, ha="center", va="center", fontsize=16,
            color="white", fontweight="bold")
ax.set_xlim(0, 1); ax.set_ylim(0, 1)
ax.set_title("Key Performance Indicators", color="white", fontsize=10, pad=6)

# weekly alert trend
ax2 = axes[0, 1]
colors = ["#D9534F" if w == week_struct.index(max(week_struct)) else "#2E5E9E"
          for w in range(len(week_alerts))]
bars = ax2.bar(range(len(week_labels)), week_alerts, color=colors, width=0.7, zorder=3)
ax2.plot(range(len(week_labels)), week_struct, "o--", color="#E8A838", lw=2,
         label="Structuring only", zorder=4)
ax2.axvspan(week_struct.index(max(week_struct))-0.4,
            week_struct.index(max(week_struct))+0.4,
            alpha=0.15, color="red", label="Structuring spike")
ax2.set_xticks(range(len(week_labels)))
ax2.set_xticklabels(week_labels, rotation=45, fontsize=7)
ax2.set_title("Weekly Alert Volume (structuring spike — Week 6)", fontsize=10,
              color="#1B3A6B", fontweight="bold")
ax2.set_ylabel("Alert Count", fontsize=9)
ax2.legend(fontsize=8)
ax2.yaxis.grid(True, alpha=0.3)
ax2.set_facecolor("#F8FAFF")

# alert type breakdown
ax3 = axes[1, 0]
type_counts = alerts_df["Alert_Type"].value_counts()
wedge_colors = ["#D9534F","#2E5E9E","#E8A838","#5CB85C","#9B59B6"]
wedges, texts, autotexts = ax3.pie(
    type_counts.values, labels=type_counts.index,
    autopct="%1.1f%%", colors=wedge_colors[:len(type_counts)],
    startangle=90, wedgeprops={"edgecolor":"white","linewidth":1.5},
    textprops={"fontsize": 8})
for at in autotexts:
    at.set_fontsize(8)
    at.set_color("white")
    at.set_fontweight("bold")
ax3.set_title("Alert Type Distribution", fontsize=10, color="#1B3A6B", fontweight="bold")

# DQ flag summary
ax4 = axes[1, 1]
flag_labels = ["CRITICAL:\nNo Risk Profile","WARN:\nMissing KYC",
               "WARN:\nKYC Expired","INFO:\nKYC Pending","OK"]
flag_colors = ["#D9534F","#E8A838","#F0AD4E","#5BC0DE","#5CB85C"]
# compute actual DQ flags from data
merged = trans_df[["Transaction_ID","Customer_ID"]].merge(
    risk_df[["Customer_ID","Risk_Rating","KYC_Status"]], on="Customer_ID", how="left")
critical_count = merged["Risk_Rating"].isna().sum()
missing_kyc    = ((merged["KYC_Status"].isna()) & (merged["Risk_Rating"].notna())).sum()
expired_kyc    = (merged["KYC_Status"] == "Expired").sum()
pending_kyc    = (merged["KYC_Status"] == "Pending").sum()
ok_count       = len(merged) - critical_count - missing_kyc - expired_kyc - pending_kyc
flag_vals = [critical_count, missing_kyc, expired_kyc, pending_kyc, ok_count]
hbars = ax4.barh(flag_labels, flag_vals, color=flag_colors, height=0.6, zorder=3)
ax4.xaxis.grid(True, alpha=0.3)
ax4.set_title("Data Quality Flag Distribution", fontsize=10, color="#1B3A6B", fontweight="bold")
ax4.set_xlabel("Transaction Count", fontsize=9)
ax4.set_facecolor("#F8FAFF")
for bar, val in zip(hbars, flag_vals):
    ax4.text(bar.get_width() + 2, bar.get_y() + bar.get_height()/2,
             f"{val:,}", va="center", fontsize=8, fontweight="bold")

plt.tight_layout(rect=[0, 0, 1, 0.97])
plt.savefig(os.path.join(I_DIR, "01-dashboard-overview.png"), dpi=150, bbox_inches="tight")
plt.close()
print("   -> 01-dashboard-overview.png")

# ── image 2: weekly trend (standalone detail) ──────────────────────────────────
fig, ax = plt.subplots(figsize=(13, 5))
fig.patch.set_facecolor("#F8FAFF")
ax.set_facecolor("#F8FAFF")
x = range(len(week_labels))
non_struct = [t - s for t, s in zip(week_alerts, week_struct)]
b1 = ax.bar(x, non_struct, label="Non-Structuring", color="#2E5E9E", width=0.65, zorder=3)
b2 = ax.bar(x, week_struct, bottom=non_struct, label="Structuring",
            color="#D9534F", width=0.65, zorder=3)

struct_wk = week_struct.index(max(week_struct))
ax.axvspan(struct_wk - 0.45, struct_wk + 0.45, alpha=0.12, color="#D9534F")
ax.annotate("Structuring cluster:\n20 customers, $9K-$9.9K txns\n(Feb 9-15)",
             xy=(struct_wk, week_alerts[struct_wk]),
             xytext=(struct_wk + 1.8, week_alerts[struct_wk] + 3),
             fontsize=8.5, color="#9C0006", fontweight="bold",
             arrowprops=dict(arrowstyle="->", color="#9C0006", lw=1.5))
ax.set_xticks(list(x))
ax.set_xticklabels(week_labels, rotation=40, ha="right", fontsize=8)
ax.set_title("Weekly Alert Volume — Jan–Apr 2026  (Structuring spike Week 6)",
             fontsize=12, fontweight="bold", color="#1B3A6B")
ax.set_ylabel("Number of Alerts", fontsize=10)
ax.legend(fontsize=9, loc="upper right")
ax.yaxis.grid(True, alpha=0.3, linestyle="--")
ax.set_xlabel("Week", fontsize=10)
plt.tight_layout()
plt.savefig(os.path.join(I_DIR, "02-weekly-trend-chart.png"), dpi=150, bbox_inches="tight")
plt.close()
print("   -> 02-weekly-trend-chart.png")

# ── image 3: alert type + status breakdown ─────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5))
fig.patch.set_facecolor("#F8FAFF")
fig.suptitle("Alert Classification & Status Breakdown", fontsize=12,
             fontweight="bold", color="#1B3A6B")

type_counts = alerts_df["Alert_Type"].value_counts()
ax1 = axes[0]
ax1.set_facecolor("#F8FAFF")
y_pos = range(len(type_counts))
bar_colors = ["#D9534F" if k == "Structuring" else "#2E5E9E" for k in type_counts.index]
hb = ax1.barh(list(type_counts.index), type_counts.values, color=bar_colors, height=0.6, zorder=3)
ax1.xaxis.grid(True, alpha=0.3)
ax1.set_title("Alerts by Type", fontsize=10, color="#1B3A6B", fontweight="bold")
ax1.set_xlabel("Count", fontsize=9)
for bar, val in zip(hb, type_counts.values):
    ax1.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
             str(val), va="center", fontsize=9, fontweight="bold")

status_counts = alerts_df["Status"].value_counts()
ax2 = axes[1]
status_colors = {"Closed":"#5CB85C","Open":"#E8A838","Escalated":"#D9534F"}
pie_colors = [status_colors.get(s, "#999") for s in status_counts.index]
wedges, texts, autotexts = ax2.pie(
    status_counts.values, labels=status_counts.index,
    autopct="%1.1f%%", colors=pie_colors,
    startangle=90, wedgeprops={"edgecolor":"white","linewidth":2},
    textprops={"fontsize": 9})
for at in autotexts:
    at.set_fontsize(9)
    at.set_fontweight("bold")
    at.set_color("white")
ax2.set_title("Alert Status Distribution", fontsize=10, color="#1B3A6B", fontweight="bold")
plt.tight_layout()
plt.savefig(os.path.join(I_DIR, "03-alert-type-breakdown.png"), dpi=150, bbox_inches="tight")
plt.close()
print("   -> 03-alert-type-breakdown.png")

# ── image 4: formula bar simulation ───────────────────────────────────────────
fig, axes = plt.subplots(2, 1, figsize=(14, 6))
fig.patch.set_facecolor("#252526")

formulas = [
    ("Consolidated_View!P2  —  4-Level Nested IF (Data Quality Flag)",
     "=IF(H2=\"Not Found\",\"CRITICAL: No Risk Profile\",\n"
     " IF(I2=\"Missing\",\"WARN: Missing KYC\",\n"
     "  IF(I2=\"Expired\",\"WARN: KYC Expired\",\n"
     "   IF(I2=\"Pending\",\"INFO: KYC Pending\",\"OK\"))))"),
    ("Consolidated_View!H2  —  XLOOKUP + INDEX/MATCH (Multi-Source Join)",
     "=XLOOKUP(B2, Raw_Risk_Profiles!$A:$A, Raw_Risk_Profiles!$B:$B, \"Not Found\")\n"
     "-- Risk_Rating column --\n\n"
     "=IFERROR(INDEX(Raw_Risk_Profiles!$E:$E, MATCH(B2, Raw_Risk_Profiles!$A:$A, 0)), \"Unknown\")\n"
     "-- Customer_Segment column (INDEX-MATCH technique) --"),
]

for ax, (title, formula) in zip(axes, formulas):
    ax.set_facecolor("#1E1E1E")
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.axis("off")
    # title bar
    ax.add_patch(mpatches.FancyBboxPatch((0, 0.82), 1.0, 0.17,
        boxstyle="square,pad=0", facecolor="#2D2D30", edgecolor="none"))
    ax.text(0.01, 0.91, title, color="#DCDCAA", fontsize=8.5,
            fontfamily="monospace", fontweight="bold", va="center")
    # formula body
    ax.add_patch(mpatches.FancyBboxPatch((0, 0.0), 1.0, 0.80,
        boxstyle="square,pad=0", facecolor="#1E1E1E", edgecolor="#3E3E42"))
    ax.text(0.02, 0.72, formula, color="#9CDCFE", fontsize=9,
            fontfamily="monospace", va="top", multialignment="left")
    for kw in ["XLOOKUP","INDEX","MATCH","IF","IFERROR"]:
        pass  # colour-coding handled by single text block above

fig.suptitle("Excel Formula Bar — Live Multi-Source Formulas", fontsize=11,
             color="#CCCCCC", y=1.01)
plt.tight_layout(pad=0.3)
plt.savefig(os.path.join(I_DIR, "04-formula-view.png"), dpi=150, bbox_inches="tight",
            facecolor="#252526")
plt.close()
print("   -> 04-formula-view.png")

# ─── data quality write-up ─────────────────────────────────────────────────────
print("[6/6] Writing data quality report ...")

n_orphan_alerts = 10
n_dup_txn       = 5
n_missing_risk  = int(merged["Risk_Rating"].isna().sum())
n_kyc_issues    = int((merged["KYC_Status"].isin(["Expired","Pending"])).sum())

dq_md = f"""# AML MIS Data Quality & Validation Notes
**Dataset:** 120-day AML monitoring period (Jan–Apr 2026)
**Prepared by:** Reporting Analytics | Date: {date.today()}

---

## Data Quality Checks Performed

### Check 1: Orphaned Alert Records
**Issue:** Alerts referencing Transaction_IDs not present in the transactions table.
**Detection formula:** `COUNTIF(Raw_Transactions!$A:$A, Raw_Alerts!B2) = 0`
**Findings:** **{n_orphan_alerts} orphaned alert records** identified. These alerts (TXN_ORPHAN_000 to TXN_ORPHAN_009) cannot be linked to a source transaction, making disposition investigation impossible.
**Remediation:** Escalate to source system owners; check if transactions were purged before alert log export.

### Check 2: Duplicate Transaction Records
**Issue:** Multiple Transaction_IDs with identical Customer_ID, Date, and Amount (near-duplicate pattern).
**Detection formula:** `COUNTIFS($B:$B,B2,$C:$C,C2,$D:$D,D2)>1`
**Findings:** **{n_dup_txn} near-duplicate records** found — same customer, same date, same amount. These may represent double-posted entries or batch processing errors.
**Remediation:** Deduplicate based on Customer_ID + Date + Amount combination; retain earliest Transaction_ID.

### Check 3: Missing Customer Risk Profiles
**Issue:** Transactions from customers with no corresponding entry in the Risk Profile table.
**Detection formula:** `XLOOKUP(B2, Raw_Risk_Profiles!$A:$A, Raw_Risk_Profiles!$B:$B, "Not Found")`
**Findings:** **{n_missing_risk} transactions** from customers with no risk profile (flag = "CRITICAL: No Risk Profile").
These transactions cannot be properly risk-scored for AML screening — a compliance gap.
**Remediation:** Enrich customer master for flagged Customer_IDs; apply default "High" risk rating pending review.

### Check 4: Expired / Pending KYC Status
**Issue:** Customers transacting with expired or pending KYC verification.
**Detection formula:** `IF(I2="Expired","WARN: KYC Expired",IF(I2="Pending","INFO: KYC Pending","OK"))`
**Findings:** **{n_kyc_issues} transactions** from customers with non-complete KYC status.
KYC-expired customers are flagged as "WARN" — transactions from these accounts should be reviewed before processing.
**Remediation:** Customer Operations to initiate KYC refresh; flag accounts for enhanced due diligence.

---

## Summary Table

| Check | Issue Found | Count | Severity | Action |
|---|---|---|---|---|
| Orphaned alerts | Alert with no matching transaction | {n_orphan_alerts} | High | Escalate to source system |
| Duplicate transactions | Same customer/date/amount | {n_dup_txn} | Medium | Dedup; retain earliest ID |
| Missing risk profiles | Customer not in risk master | {n_missing_risk} | High | Default to High risk; enrich master |
| Expired/Pending KYC | Customer KYC not current | {n_kyc_issues} | Medium | Enhanced due diligence |

---

## Structuring Pattern Identified
During weekly trend analysis, **Week 6 (Feb 9–15)** showed an anomalous spike in Structuring alerts.
Investigation revealed **20 customers** each made 4–6 cash deposits of **$9,000–$9,999** within the same 7-day window —
a classic structuring pattern designed to stay below the $10,000 CTR reporting threshold.

**All 20 customers** already carry a High Risk Rating in the customer master.
Recommended action: File Suspicious Activity Reports (SARs) and escalate to AML investigations team.
"""

with open(os.path.join(BASE, "report", "data_quality_notes.md"), "w", encoding="utf-8") as f:
    f.write(dq_md)
print("   -> report/data_quality_notes.md")
print("\nBuild complete.")
print(f"  Transactions : {N_TRANS}")
print(f"  Risk profiles: {len(risk_df)}")
print(f"  Alerts       : {N_ALERTS}")
