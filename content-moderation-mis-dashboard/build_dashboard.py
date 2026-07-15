#!/usr/bin/env python3
"""
Content Moderation Operations MIS Dashboard Generator
Generates a 90-day simulated dataset and builds an Excel workbook.
"""

import os
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

np.random.seed(42)

# ─────────────────────────────────────────────
# 1.  CONFIGURATION
# ─────────────────────────────────────────────
START_DATE = datetime(2026, 1, 1)
N_DAYS     = 90
MODERATORS = [f"MOD_{i:03d}" for i in range(1, 21)]
CATEGORIES = [
    "Hate Speech", "Violence", "Nudity", "Spam",
    "Misinformation", "Child Safety", "Bullying", "Self-Harm"
]

CAT_PARAMS = {
    "Hate Speech":    dict(base_rt=180, fp=.020, fn=.015, esc=.030),
    "Violence":       dict(base_rt=210, fp=.018, fn=.012, esc=.040),
    "Nudity":         dict(base_rt=120, fp=.025, fn=.020, esc=.015),
    "Spam":           dict(base_rt= 60, fp=.030, fn=.025, esc=.005),
    "Misinformation": dict(base_rt=240, fp=.022, fn=.018, esc=.035),
    "Child Safety":   dict(base_rt=300, fp=.008, fn=.005, esc=.060),
    "Bullying":       dict(base_rt=150, fp=.028, fn=.022, esc=.020),
    "Self-Harm":      dict(base_rt=270, fp=.010, fn=.008, esc=.050),
}

# Incident window: day indices 34-38 = Feb 4-8, 2026
INCIDENT_DAYS = set(range(34, 39))

# ─────────────────────────────────────────────
# 2.  GENERATE DATA
# ─────────────────────────────────────────────
rows = []
for day_idx in range(N_DAYS):
    date       = START_DATE + timedelta(days=day_idx)
    dow        = date.weekday()      # 0=Mon … 6=Sun
    is_monday  = (dow == 0)
    is_weekend = (dow >= 5)
    is_inc     = (day_idx in INCIDENT_DAYS)

    for mod in MODERATORS:
        eff = np.random.uniform(0.88, 1.12)  # moderator speed factor
        for cat in CATEGORIES:
            p = CAT_PARAMS[cat]

            # Volume with realistic day-of-week + incident multipliers
            vol = max(1, int(
                np.random.poisson(15)
                * (1.40 if is_monday  else 1.0)
                * (0.70 if is_weekend else 1.0)
                * (1.80 if is_inc     else 1.0)
            ))

            # Review time (seconds) — sensitive categories take longer
            rt = max(30, np.random.normal(
                p["base_rt"] * eff,
                p["base_rt"] * 0.15
            ) * (1.20 if is_inc else 1.0))

            # Error rates
            fp = int(np.random.binomial(vol, p["fp"]))
            fn = int(np.random.binomial(vol, p["fn"]))

            # Escalations (higher for sensitive categories + incident)
            esc_r = p["esc"] * (1.6 if is_inc else 1.0)
            esc   = int(np.random.binomial(vol, esc_r))

            # SLA breaches — Monday surge + incident spike are deliberate patterns
            br_r  = 0.12 if is_monday else (0.20 if is_inc else 0.04)
            brs   = int(np.random.binomial(vol, br_r))
            sla_c = round(100.0 * (1.0 - brs / vol), 2)

            rows.append((
                date.strftime("%Y-%m-%d"), mod, cat, vol,
                round(rt, 1), fp, fn, esc, 95.0, brs, sla_c
            ))

COLS = [
    "Date", "Moderator_ID", "Content_Category", "Videos_Reviewed",
    "Avg_Review_Time_Sec", "False_Positives", "False_Negatives",
    "Escalations", "SLA_Target_Pct", "SLA_Breaches", "SLA_Compliance_Pct"
]

df = pd.DataFrame(rows, columns=COLS)

# Status column — nested IF logic equivalent
df["Status"] = np.where(
    df.SLA_Compliance_Pct < 90, "Critical",
    np.where(df.SLA_Compliance_Pct < 95, "Warning", "Healthy")
)

os.makedirs("data",   exist_ok=True)
os.makedirs("excel",  exist_ok=True)
os.makedirs("report", exist_ok=True)

CSV_PATH  = "data/moderation_data_90day.csv"
XLSX_PATH = "excel/content_moderation_mis_dashboard.xlsx"

df.to_csv(CSV_PATH, index=False)
print(f"CSV saved  -> {CSV_PATH}  ({len(df):,} rows)")

# ─────────────────────────────────────────────
# 3.  STYLE HELPERS
# ─────────────────────────────────────────────
def mk_fill(hex_):   return PatternFill("solid", fgColor=hex_)
def mk_font(hex_, bold=False, sz=11, italic=False):
    return Font(color=hex_, bold=bold, size=sz, italic=italic)

F_NAVY   = mk_fill("1F3864"); F_BLUE   = mk_fill("2E75B6")
F_GREEN  = mk_fill("375623"); F_ORANGE = mk_fill("833C00")
F_RED    = mk_fill("FF0000"); F_YELLOW = mk_fill("FFFF00")
F_LGRN   = mk_fill("92D050"); F_LBLUE  = mk_fill("D9E1F2")
F_LGRAY  = mk_fill("F2F2F2"); F_CREAM  = mk_fill("FFF2CC")

WH     = mk_font("FFFFFF", bold=True)
BL     = mk_font("000000")
NV     = mk_font("1F3864", bold=True)
thin   = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP   = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row_num, fill=F_NAVY, fnt=WH):
    for cell in ws[row_num]:
        cell.fill = fill; cell.font = fnt
        cell.alignment = CENTER; cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────
# 4.  WORKBOOK
# ─────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════
# SHEET 1: Raw_Data
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Raw_Data"
ws1.sheet_properties.tabColor = "1F3864"

ws1.append(COLS + ["Status"])
style_header(ws1, 1)

for r in df.itertuples(index=False):
    ws1.append(list(r))

ws1.freeze_panes = "A2"
set_widths(ws1, [12, 14, 18, 18, 21, 17, 17, 13, 15, 13, 22, 10])

last = len(df) + 1
for formula, fill_, fnt_ in [
    ('$L2="Critical"', F_RED,    mk_font("FFFFFF", bold=True)),
    ('$L2="Warning"',  F_YELLOW, BL),
    ('$L2="Healthy"',  F_LGRN,   mk_font("375623")),
]:
    ws1.conditional_formatting.add(
        f"L2:L{last}",
        FormulaRule(formula=[formula], fill=fill_, font=fnt_)
    )

# ══════════════════════════════════════════════
# SHEET 2: Moderator_Summary
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Moderator_Summary")
ws2.sheet_properties.tabColor = "2E75B6"

ws2.append([
    "Moderator_ID", "Total_Videos", "Avg_Review_Time_Sec",
    "Total_FP", "Total_FN", "Error_Rate_Pct",
    "Escalations", "Avg_SLA_Pct", "Performance_Tier"
])
style_header(ws2, 1, F_BLUE)
ws2.freeze_panes = "A2"

for i, mod in enumerate(MODERATORS, 2):
    ws2[f"A{i}"] = mod
    ws2[f"B{i}"] = f"=SUMIF(Raw_Data!$B:$B,A{i},Raw_Data!$D:$D)"
    ws2[f"C{i}"] = f"=AVERAGEIF(Raw_Data!$B:$B,A{i},Raw_Data!$E:$E)"
    ws2[f"D{i}"] = f"=SUMIF(Raw_Data!$B:$B,A{i},Raw_Data!$F:$F)"
    ws2[f"E{i}"] = f"=SUMIF(Raw_Data!$B:$B,A{i},Raw_Data!$G:$G)"
    ws2[f"F{i}"] = f"=IFERROR((D{i}+E{i})/B{i}*100,0)"
    ws2[f"G{i}"] = f"=SUMIF(Raw_Data!$B:$B,A{i},Raw_Data!$H:$H)"
    ws2[f"H{i}"] = f"=AVERAGEIF(Raw_Data!$B:$B,A{i},Raw_Data!$K:$K)"
    # Four-tier nested IF for Performance_Tier
    ws2[f"I{i}"] = (
        f'=IF(H{i}>=97,"Top Performer",'
        f'IF(H{i}>=93,"Standard",'
        f'IF(H{i}>=88,"Needs Review","At Risk")))'
    )
    for c in range(1, 10):
        cell = ws2.cell(i, c)
        cell.border = BORDER; cell.alignment = CENTER
        cell.fill = F_LBLUE if i % 2 == 0 else F_LGRAY

set_widths(ws2, [15, 15, 22, 12, 12, 16, 14, 16, 16])

for formula, fill_, fnt_ in [
    ('$I2="Top Performer"', F_LGRN, mk_font("375623")),
    ('$I2="At Risk"',       F_RED,  mk_font("FFFFFF", bold=True)),
]:
    ws2.conditional_formatting.add(
        f"I2:I{len(MODERATORS)+1}",
        FormulaRule(formula=[formula], fill=fill_, font=fnt_)
    )

# ══════════════════════════════════════════════
# SHEET 3: Category_Breakdown
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("Category_Breakdown")
ws3.sheet_properties.tabColor = "375623"

ws3.append([
    "Content_Category", "Total_Videos", "Avg_Review_Time_Sec",
    "Total_FP", "Total_FN", "Error_Rate_Pct",
    "Escalations", "Avg_SLA_Pct", "Avg_Daily_Volume"
])
style_header(ws3, 1, F_GREEN)
ws3.freeze_panes = "A2"

for i, cat in enumerate(CATEGORIES, 2):
    ws3[f"A{i}"] = cat
    ws3[f"B{i}"] = f"=SUMIF(Raw_Data!$C:$C,A{i},Raw_Data!$D:$D)"
    ws3[f"C{i}"] = f"=AVERAGEIF(Raw_Data!$C:$C,A{i},Raw_Data!$E:$E)"
    ws3[f"D{i}"] = f"=SUMIF(Raw_Data!$C:$C,A{i},Raw_Data!$F:$F)"
    ws3[f"E{i}"] = f"=SUMIF(Raw_Data!$C:$C,A{i},Raw_Data!$G:$G)"
    ws3[f"F{i}"] = f"=IFERROR((D{i}+E{i})/B{i}*100,0)"
    ws3[f"G{i}"] = f"=SUMIF(Raw_Data!$C:$C,A{i},Raw_Data!$H:$H)"
    ws3[f"H{i}"] = f"=AVERAGEIF(Raw_Data!$C:$C,A{i},Raw_Data!$K:$K)"
    ws3[f"I{i}"] = f"=ROUND(B{i}/90,1)"   # avg daily volume
    for c in range(1, 10):
        cell = ws3.cell(i, c)
        cell.border = BORDER; cell.alignment = CENTER
        cell.fill = F_LBLUE if i % 2 == 0 else F_LGRAY

set_widths(ws3, [18, 15, 22, 12, 12, 14, 14, 14, 18])

# ══════════════════════════════════════════════
# SHEET 4: MIS_Report  (daily aggregates — pivot-table equivalent)
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("MIS_Report")
ws4.sheet_properties.tabColor = "833C00"

unique_dates = sorted(df["Date"].unique())

ws4.append([
    "Date", "Day_of_Week", "Week_Num", "Total_Videos",
    "SLA_Breaches", "Avg_SLA_Pct", "Escalations",
    "Total_Errors", "Status_Flag"
])
style_header(ws4, 1, F_ORANGE)
ws4.freeze_panes = "A2"

for i, d in enumerate(unique_dates, 2):
    ws4[f"A{i}"] = d
    ws4[f"B{i}"] = f'=TEXT(DATEVALUE(A{i}),"DDD")'
    ws4[f"C{i}"] = f"=WEEKNUM(DATEVALUE(A{i}),1)"
    ws4[f"D{i}"] = f"=SUMIF(Raw_Data!$A:$A,A{i},Raw_Data!$D:$D)"
    ws4[f"E{i}"] = f"=SUMIF(Raw_Data!$A:$A,A{i},Raw_Data!$J:$J)"
    ws4[f"F{i}"] = f"=AVERAGEIF(Raw_Data!$A:$A,A{i},Raw_Data!$K:$K)"
    ws4[f"G{i}"] = f"=SUMIF(Raw_Data!$A:$A,A{i},Raw_Data!$H:$H)"
    ws4[f"H{i}"] = (
        f"=SUMIF(Raw_Data!$A:$A,A{i},Raw_Data!$F:$F)"
        f"+SUMIF(Raw_Data!$A:$A,A{i},Raw_Data!$G:$G)"
    )
    # Nested IF status flag
    ws4[f"I{i}"] = f'=IF(F{i}<90,"Critical",IF(F{i}<95,"Warning","Healthy"))'
    for c in range(1, 10):
        ws4.cell(i, c).border = BORDER
        ws4.cell(i, c).alignment = CENTER

# Conditional formatting on SLA compliance (col F)
sla_r = f"F2:F{len(unique_dates)+1}"
ws4.conditional_formatting.add(sla_r,
    CellIsRule("lessThan",           ["90"],        fill=F_RED,    font=mk_font("FFFFFF", bold=True)))
ws4.conditional_formatting.add(sla_r,
    CellIsRule("between",            ["90","94.99"],fill=F_YELLOW, font=BL))
ws4.conditional_formatting.add(sla_r,
    CellIsRule("greaterThanOrEqual", ["95"],        fill=F_LGRN,   font=mk_font("375623")))

set_widths(ws4, [14, 13, 12, 14, 14, 16, 14, 14, 12])

# ══════════════════════════════════════════════
# SHEET 5: Dashboard
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("Dashboard")
ws5.sheet_properties.tabColor = "C00000"

# ── Title banner ──
ws5.merge_cells("A1:M1")
ws5["A1"].value = "CONTENT MODERATION OPERATIONS  —  MIS DASHBOARD"
ws5["A1"].font  = Font(bold=True, size=18, color="FFFFFF")
ws5["A1"].fill  = F_NAVY; ws5["A1"].alignment = CENTER
ws5.row_dimensions[1].height = 42

ws5.merge_cells("A2:M2")
ws5["A2"].value = "90-Day Simulated Operations Summary  |  Jan 1 – Mar 31, 2026  |  20 Moderators  |  8 Content Categories"
ws5["A2"].font  = mk_font("1F3864", italic=True, sz=10)
ws5["A2"].alignment = CENTER
ws5.row_dimensions[2].height = 20

# ── KPI section header ──
ws5.merge_cells("A4:M4")
ws5["A4"].value = "KEY PERFORMANCE INDICATORS"
ws5["A4"].font  = Font(bold=True, size=12, color="FFFFFF")
ws5["A4"].fill  = F_BLUE; ws5["A4"].alignment = CENTER

KPI_DATA = [
    ("Total Videos\nReviewed",       "=SUM(Raw_Data!D:D)"),
    ("Overall SLA\nCompliance %",    "=ROUND(AVERAGE(Raw_Data!K:K),1)"),
    ("Avg Error\nRate %",            "=ROUND(IFERROR((SUM(Raw_Data!F:F)+SUM(Raw_Data!G:G))/SUM(Raw_Data!D:D)*100,0),3)"),
    ("Escalation\nRate %",           "=ROUND(IFERROR(SUM(Raw_Data!H:H)/SUM(Raw_Data!D:D)*100,0),3)"),
    ("Avg Review\nTime (sec)",       "=ROUND(AVERAGE(Raw_Data!E:E),1)"),
    ("Active\nModerators",           f"={len(MODERATORS)}"),
]

ws5.row_dimensions[5].height = 26
ws5.row_dimensions[6].height = 36

for idx, (label, formula) in enumerate(KPI_DATA):
    c1 = idx * 2 + 1; c2 = c1 + 1
    l1 = get_column_letter(c1); l2 = get_column_letter(c2)
    ws5.merge_cells(f"{l1}5:{l2}5")
    lc = ws5[f"{l1}5"]; lc.value = label.replace("\\n", "\n")
    lc.font = Font(bold=True, size=9, color="FFFFFF"); lc.fill = F_BLUE
    lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws5.merge_cells(f"{l1}6:{l2}6")
    vc = ws5[f"{l1}6"]; vc.value = formula
    vc.font = Font(bold=True, size=15, color="1F3864")
    vc.fill = F_LBLUE; vc.alignment = CENTER

# ── Trend table header ──
ws5.merge_cells("A8:F8")
ws5["A8"].value = "DAILY SLA COMPLIANCE TREND (First 30 Days)"
ws5["A8"].font  = Font(bold=True, size=11, color="FFFFFF")
ws5["A8"].fill  = F_BLUE; ws5["A8"].alignment = CENTER

for c, h in enumerate(["Date","Day","Total Videos","SLA Breaches","SLA Compliance %","Status"], 1):
    cell = ws5.cell(9, c); cell.value = h
    cell.font = WH; cell.fill = F_NAVY; cell.alignment = CENTER; cell.border = BORDER

# ── 30-day detail rows using XLOOKUP to pull from MIS_Report ──
for r in range(10, 40):
    mis_row = r - 8   # Dashboard row 10 → MIS_Report row 2
    ws5.cell(r, 1).value = f"=MIS_Report!A{mis_row}"
    ws5.cell(r, 2).value = f'=TEXT(DATEVALUE(MIS_Report!A{mis_row}),"DDD")'
    # XLOOKUP: look up each date in MIS_Report and return corresponding metric
    ws5.cell(r, 3).value = f"=XLOOKUP(MIS_Report!A{mis_row},MIS_Report!$A:$A,MIS_Report!$D:$D,0)"
    ws5.cell(r, 4).value = f"=XLOOKUP(MIS_Report!A{mis_row},MIS_Report!$A:$A,MIS_Report!$E:$E,0)"
    ws5.cell(r, 5).value = f"=XLOOKUP(MIS_Report!A{mis_row},MIS_Report!$A:$A,MIS_Report!$F:$F,0)"
    # Nested IF for status
    ws5.cell(r, 6).value = f'=IF(E{r}<90,"Critical",IF(E{r}<95,"Warning","Healthy"))'
    for c in range(1, 7):
        ws5.cell(r, c).border = BORDER; ws5.cell(r, c).alignment = CENTER

# CF on SLA compliance (col E, rows 10-39) — traffic-light colours
ws5.conditional_formatting.add("E10:E39",
    CellIsRule("lessThan",           ["90"],        fill=F_RED,    font=mk_font("FFFFFF", bold=True)))
ws5.conditional_formatting.add("E10:E39",
    CellIsRule("between",            ["90","94.99"],fill=F_YELLOW, font=BL))
ws5.conditional_formatting.add("E10:E39",
    CellIsRule("greaterThanOrEqual", ["95"],        fill=F_LGRN,   font=mk_font("375623")))

for formula, fill_, fnt_ in [
    ('$F10="Critical"', F_RED,    mk_font("FFFFFF", bold=True)),
    ('$F10="Warning"',  F_YELLOW, BL),
    ('$F10="Healthy"',  F_LGRN,   mk_font("375623")),
]:
    ws5.conditional_formatting.add(
        "F10:F39",
        FormulaRule(formula=[formula], fill=fill_, font=fnt_)
    )

# ── Data validation dropdown ──
ws5["H8"].value = "Filter by Category ▼"
ws5["H8"].font  = mk_font("1F3864", bold=True, sz=10)
ws5["H8"].fill  = F_CREAM; ws5["H8"].border = BORDER; ws5["H8"].alignment = CENTER

dv = DataValidation(
    type="list",
    formula1='"All Categories,Hate Speech,Violence,Nudity,Spam,Misinformation,Child Safety,Bullying,Self-Harm"',
    allow_blank=True,
    showDropDown=False,
)
ws5.add_data_validation(dv)
ws5["H9"].value = "All Categories"
ws5["H9"].fill  = F_CREAM; ws5["H9"].border = BORDER; ws5["H9"].alignment = CENTER
dv.add("H9")

# ── Weekly summary data (rows 42-55) for charts ──
for c, h in enumerate(["Week","Weekly Volume","Avg SLA Compliance %"], 1):
    cell = ws5.cell(42, c); cell.value = h
    cell.font = WH; cell.fill = F_NAVY; cell.border = BORDER; cell.alignment = CENTER

for wk in range(1, 14):
    r = 42 + wk
    ws5[f"A{r}"] = f"Wk {wk}"
    ws5[f"B{r}"] = (
        f"=SUMPRODUCT((MIS_Report!$C$2:$C$91={wk})"
        f"*(MIS_Report!$D$2:$D$91))"
    )
    ws5[f"C{r}"] = (
        f"=IFERROR(SUMPRODUCT((MIS_Report!$C$2:$C$91={wk})"
        f"*(MIS_Report!$F$2:$F$91))"
        f"/COUNTIF(MIS_Report!$C$2:$C$91,{wk}),0)"
    )
    for c in range(1, 4):
        ws5.cell(r, c).border = BORDER; ws5.cell(r, c).alignment = CENTER

# ── Bar chart: weekly review volume ──
bar_chart = BarChart(); bar_chart.type = "col"; bar_chart.style = 10
bar_chart.title = "Weekly Video Review Volume"; bar_chart.grouping = "clustered"
bar_chart.y_axis.title = "Videos Reviewed"; bar_chart.x_axis.title = "Week"
vol_ref  = Reference(ws5, min_col=2, max_col=2, min_row=42, max_row=55)
cats_ref = Reference(ws5, min_col=1, min_row=43, max_row=55)
bar_chart.add_data(vol_ref, titles_from_data=True)
bar_chart.set_categories(cats_ref)
bar_chart.width = 20; bar_chart.height = 13
ws5.add_chart(bar_chart, "H11")

# ── Line chart: weekly SLA compliance ──
line_chart = LineChart(); line_chart.style = 10
line_chart.title = "Average Weekly SLA Compliance %"
line_chart.y_axis.title = "SLA Compliance %"
line_chart.y_axis.scaling.min = 80
sla_ref = Reference(ws5, min_col=3, max_col=3, min_row=42, max_row=55)
line_chart.add_data(sla_ref, titles_from_data=True)
line_chart.set_categories(cats_ref)
line_chart.width = 20; line_chart.height = 13
ws5.add_chart(line_chart, "H28")

set_widths(ws5, [14, 8, 16, 14, 20, 12, 12, 24, 16])

# ─────────────────────────────────────────────
# 5.  SAVE WORKBOOK
# ─────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"Excel saved -> {XLSX_PATH}")
print("\nDone. Open the .xlsx in Excel to see all formulas calculate live.")
