#!/usr/bin/env python3
"""
Advanced Excel Reporting Framework Builder
Generates 180-day support ticket dataset, Excel workbook, and chart images.
"""

import os, sys
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mtick
from matplotlib.gridspec import GridSpec
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

np.random.seed(42)

# ─────────────────────────────────────────────
# 1.  CONFIGURATION
# ─────────────────────────────────────────────
START   = datetime(2026, 1, 1)
N_DAYS  = 180
AGENTS  = [f"AGENT_{i:02d}" for i in range(1, 16)]
CATS    = ["Technical", "Billing", "Account", "Shipping",
           "Returns", "Product Quality", "General Inquiry"]
PRIS    = ["P1", "P2", "P3", "P4"]
DEPTS   = {"AGENT_01":"Technical","AGENT_02":"Technical","AGENT_03":"Technical",
           "AGENT_04":"Billing",  "AGENT_05":"Billing",  "AGENT_06":"Billing",
           "AGENT_07":"Account",  "AGENT_08":"Account",  "AGENT_09":"Account",
           "AGENT_10":"Shipping", "AGENT_11":"Shipping", "AGENT_12":"Returns",
           "AGENT_13":"Returns",  "AGENT_14":"General",  "AGENT_15":"General"}
TIERS   = {"AGENT_01":"L3","AGENT_02":"L2","AGENT_03":"L2",
           "AGENT_04":"L3","AGENT_05":"L2","AGENT_06":"L1",
           "AGENT_07":"L2","AGENT_08":"L1","AGENT_09":"L2",
           "AGENT_10":"L1","AGENT_11":"L2","AGENT_12":"L1",
           "AGENT_13":"L2","AGENT_14":"L1","AGENT_15":"L2"}

SLA_HRS = {"P1": 4, "P2": 8, "P3": 24, "P4": 48}
CAT_BASE = {"Technical": 18, "Billing": 6, "Account": 8, "Shipping": 12,
            "Returns": 10, "Product Quality": 14, "General Inquiry": 4}
PRI_MULT = {"P1": 0.8, "P2": 1.0, "P3": 1.2, "P4": 1.5}

# Incident surge: March 9-15, 2026  (day indices 67-73)
INCIDENT = set(range(67, 74))

# ─────────────────────────────────────────────
# 2.  GENERATE DATA
# ─────────────────────────────────────────────
rows, tid = [], 10000
for di in range(N_DAYS):
    d    = START + timedelta(days=di)
    dow  = d.weekday()
    wknd = dow >= 5
    inc  = di in INCIDENT

    n = max(5, np.random.poisson(18 if wknd else 32))
    for _ in range(n):
        cat = np.random.choice(CATS, p=[0.25,0.15,0.15,0.15,0.12,0.10,0.08])
        pri = np.random.choice(PRIS, p=[0.05,0.20,0.50,0.25])
        agt = np.random.choice(AGENTS)

        base = CAT_BASE[cat] * PRI_MULT[pri] * (1.65 if inc else 1.0)
        rt   = max(0.5, round(np.random.lognormal(np.log(base), 0.40), 2))

        sla_t = SLA_HRS[pri]
        sla_m = "Y" if rt <= sla_t else "N"

        if sla_m == "N" and pri in ("P1","P2"):
            status = "Escalated" if np.random.random() < 0.30 else "Resolved"
        else:
            status = "Open" if np.random.random() < 0.04 else "Resolved"

        csat = max(1, min(5, int(np.random.normal(4.1 if sla_m=="Y" else 2.8, 0.7))))

        rows.append((f"TKT-{tid}", d.strftime("%Y-%m-%d"), cat, pri, agt,
                     rt, status, csat, sla_t, sla_m))
        tid += 1

COLS = ["Ticket_ID","Date","Category","Priority","Assigned_Agent",
        "Resolution_Time_Hrs","Status","CSAT_Score","SLA_Target_Hrs","SLA_Met"]
df   = pd.DataFrame(rows, columns=COLS)

os.makedirs("data",   exist_ok=True)
os.makedirs("excel",  exist_ok=True)
os.makedirs("images", exist_ok=True)

df.to_csv("data/support_tickets_180day.csv", index=False)
print(f"CSV saved ({len(df):,} rows)")

# ─────────────────────────────────────────────
# 3.  PRE-COMPUTE AGGREGATES (for images + Excel helper rows)
# ─────────────────────────────────────────────
df["DateDT"] = pd.to_datetime(df["Date"])
df["Week"]   = df["DateDT"].dt.isocalendar().week.astype(int)
df["WeekDT"] = df["DateDT"] - pd.to_timedelta(df["DateDT"].dt.dayofweek, unit="d")

wk = (df.groupby("WeekDT")
        .agg(Tickets  =("Ticket_ID","count"),
             SLA_Rate =("SLA_Met",  lambda x: (x=="Y").mean()*100),
             Avg_RT   =("Resolution_Time_Hrs","mean"))
        .reset_index().reset_index(drop=False))
wk.columns = ["idx","WeekDT","Tickets","SLA_Rate","Avg_RT"]
wk["WkLabel"] = ["Wk "+str(i+1) for i in range(len(wk))]

cat_agg = (df.groupby("Category")
             .agg(Tickets=("Ticket_ID","count"),
                  Avg_RT =("Resolution_Time_Hrs","mean"),
                  SLA    =("SLA_Met",lambda x:(x=="Y").mean()*100),
                  CSAT   =("CSAT_Score","mean"))
             .reset_index().sort_values("Tickets", ascending=False))

agt_agg = (df.groupby("Assigned_Agent")
             .agg(Tickets=("Ticket_ID","count"),
                  Avg_RT =("Resolution_Time_Hrs","mean"),
                  SLA    =("SLA_Met",lambda x:(x=="Y").mean()*100))
             .reset_index().sort_values("SLA", ascending=False))

pri_agg = (df.groupby("Priority")
             .agg(Tickets=("Ticket_ID","count"),
                  SLA    =("SLA_Met",lambda x:(x=="Y").mean()*100))
             .reset_index())
pri_agg["Priority"] = pd.Categorical(pri_agg.Priority, categories=PRIS, ordered=True)
pri_agg = pri_agg.sort_values("Priority")

# Overall KPIs
total_tickets = len(df)
overall_sla   = (df.SLA_Met=="Y").mean()*100
avg_rt        = df.Resolution_Time_Hrs.mean()
avg_csat      = df.CSAT_Score.mean()

# ─────────────────────────────────────────────
# 4.  EXCEL WORKBOOK
# ─────────────────────────────────────────────
def mk_fill(h): return PatternFill("solid", fgColor=h)
def mk_font(h, bold=False, sz=11, italic=False):
    return Font(color=h, bold=bold, size=sz, italic=italic)

F_NAVY   = mk_fill("1F3864"); F_BLUE  = mk_fill("2E75B6")
F_GREEN  = mk_fill("375623"); F_ORG   = mk_fill("833C00")
F_RED    = mk_fill("FF0000"); F_YLW   = mk_fill("FFFF00")
F_LGRN   = mk_fill("92D050"); F_LBLUE = mk_fill("D9E1F2")
F_LGRAY  = mk_fill("F2F2F2"); F_CREAM = mk_fill("FFF2CC")
F_LRED   = mk_fill("FFCCCC")

WH     = mk_font("FFFFFF", bold=True)
BL     = mk_font("000000")
NV     = mk_font("1F3864", bold=True)
thin   = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP   = Alignment(horizontal="center", vertical="center", wrap_text=True)

def hdr(ws, row_num, fill=F_NAVY, fnt=WH):
    for cell in ws[row_num]:
        cell.fill=fill; cell.font=fnt; cell.alignment=CENTER; cell.border=BORDER

def widths(ws, ww):
    for i,w in enumerate(ww,1):
        ws.column_dimensions[get_column_letter(i)].width = w

N = len(df)  # total rows in Raw_Data (excluding header)

wb = Workbook()

# ══════════════════════════════════════════════
# SHEET 1: Raw_Data
# ══════════════════════════════════════════════
ws1 = wb.active; ws1.title="Raw_Data"
ws1.sheet_properties.tabColor="1F3864"
ws1.append(COLS); hdr(ws1,1)
ws1.freeze_panes="A2"

for r in df[COLS].itertuples(index=False):
    ws1.append(list(r))

widths(ws1,[12,12,17,9,14,21,12,12,16,9])

# CF on SLA_Met (col J)
rng_j = f"J2:J{N+1}"
ws1.conditional_formatting.add(rng_j, FormulaRule(['$J2="N"'], fill=F_LRED, font=mk_font("C00000")))
ws1.conditional_formatting.add(rng_j, FormulaRule(['$J2="Y"'], fill=F_LGRN, font=mk_font("375623")))

# ══════════════════════════════════════════════
# SHEET 2: Agent_Summary
# ══════════════════════════════════════════════

# Reference lookup table embedded in this sheet for INDEX/MATCH demo
ws2 = wb.create_sheet("Agent_Summary")
ws2.sheet_properties.tabColor="2E75B6"

# Agent reference table (cols K-M, hidden area)
ws2["K1"]="Agent_Ref"; ws2["K1"].font=WH; ws2["K1"].fill=F_NAVY
ws2["L1"]="Department"; ws2["L1"].font=WH; ws2["L1"].fill=F_NAVY
ws2["M1"]="Tier"; ws2["M1"].font=WH; ws2["M1"].fill=F_NAVY
for ir, ag in enumerate(AGENTS, 2):
    ws2.cell(ir,11).value = ag
    ws2.cell(ir,12).value = DEPTS[ag]
    ws2.cell(ir,13).value = TIERS[ag]

# Main summary table
ws2.append(["Agent","Ticket_Count","Avg_Resolution_Hrs","SLA_Met_Count",
            "SLA_Breach_Count","SLA_Compliance_Pct","Avg_CSAT",
            "Department","Tier","Health_Status"])
hdr(ws2, 1, F_BLUE); ws2.freeze_panes="A2"

for i,ag in enumerate(AGENTS,2):
    ws2[f"A{i}"] = ag
    ws2[f"B{i}"] = f"=COUNTIF(Raw_Data!$E$2:$E${N+1},A{i})"
    ws2[f"C{i}"] = f"=AVERAGEIF(Raw_Data!$E$2:$E${N+1},A{i},Raw_Data!$F$2:$F${N+1})"
    ws2[f"D{i}"] = f'=COUNTIFS(Raw_Data!$E$2:$E${N+1},A{i},Raw_Data!$J$2:$J${N+1},"Y")'
    ws2[f"E{i}"] = f'=COUNTIFS(Raw_Data!$E$2:$E${N+1},A{i},Raw_Data!$J$2:$J${N+1},"N")'
    ws2[f"F{i}"] = f"=IFERROR(D{i}/B{i}*100,0)"
    ws2[f"G{i}"] = f"=AVERAGEIF(Raw_Data!$E$2:$E${N+1},A{i},Raw_Data!$H$2:$H${N+1})"
    # INDEX/MATCH to look up Department from reference table
    ws2[f"H{i}"] = f"=IFERROR(INDEX($L$2:$L${len(AGENTS)+1},MATCH(A{i},$K$2:$K${len(AGENTS)+1},0)),\"Unknown\")"
    # INDEX/MATCH to look up Tier
    ws2[f"I{i}"] = f"=IFERROR(INDEX($M$2:$M${len(AGENTS)+1},MATCH(A{i},$K$2:$K${len(AGENTS)+1},0)),\"Unknown\")"
    # Nested IF: Health_Status based on SLA compliance % and avg resolution time
    ws2[f"J{i}"] = (
        f'=IF(AND(F{i}<75,C{i}>20),"Critical",'
        f'IF(OR(F{i}<85,C{i}>15),"Warning","Healthy"))'
    )
    for c in range(1,11):
        cell=ws2.cell(i,c); cell.border=BORDER; cell.alignment=CENTER
        cell.fill = F_LBLUE if i%2==0 else F_LGRAY

widths(ws2,[14,14,22,15,17,20,12,14,8,14, 14,14,8])

# CF on Health_Status col J
hs_rng = f"J2:J{len(AGENTS)+1}"
ws2.conditional_formatting.add(hs_rng, FormulaRule(['$J2="Critical"'], fill=F_RED,  font=mk_font("FFFFFF",bold=True)))
ws2.conditional_formatting.add(hs_rng, FormulaRule(['$J2="Warning"'],  fill=F_YLW,  font=BL))
ws2.conditional_formatting.add(hs_rng, FormulaRule(['$J2="Healthy"'],  fill=F_LGRN, font=mk_font("375623")))

# ══════════════════════════════════════════════
# SHEET 3: Category_Breakdown
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("Category_Breakdown")
ws3.sheet_properties.tabColor="375623"

ws3.append(["Category","Priority","Total_Tickets","Avg_Resolution_Hrs",
            "SLA_Met_Count","SLA_Breach_Count","SLA_Compliance_Pct","Avg_CSAT"])
hdr(ws3, 1, F_GREEN); ws3.freeze_panes="A2"

row3 = 2
for cat in CATS:
    for pri in PRIS:
        ws3.cell(row3,1).value = cat
        ws3.cell(row3,2).value = pri
        # COUNTIFS: count tickets matching both category AND priority
        ws3.cell(row3,3).value = (
            f'=COUNTIFS(Raw_Data!$C$2:$C${N+1},A{row3},'
            f'Raw_Data!$D$2:$D${N+1},B{row3})'
        )
        # AVERAGEIFS: avg resolution time for cat+priority
        ws3.cell(row3,4).value = (
            f'=IFERROR(AVERAGEIFS(Raw_Data!$F$2:$F${N+1},'
            f'Raw_Data!$C$2:$C${N+1},A{row3},'
            f'Raw_Data!$D$2:$D${N+1},B{row3}),0)'
        )
        # COUNTIFS: SLA Met count
        ws3.cell(row3,5).value = (
            f'=COUNTIFS(Raw_Data!$C$2:$C${N+1},A{row3},'
            f'Raw_Data!$D$2:$D${N+1},B{row3},'
            f'Raw_Data!$J$2:$J${N+1},"Y")'
        )
        # COUNTIFS: SLA Breach count
        ws3.cell(row3,6).value = (
            f'=COUNTIFS(Raw_Data!$C$2:$C${N+1},A{row3},'
            f'Raw_Data!$D$2:$D${N+1},B{row3},'
            f'Raw_Data!$J$2:$J${N+1},"N")'
        )
        # SLA Compliance %
        ws3.cell(row3,7).value = f"=IFERROR(E{row3}/C{row3}*100,0)"
        # AVERAGEIFS: avg CSAT
        ws3.cell(row3,8).value = (
            f'=IFERROR(AVERAGEIFS(Raw_Data!$H$2:$H${N+1},'
            f'Raw_Data!$C$2:$C${N+1},A{row3},'
            f'Raw_Data!$D$2:$D${N+1},B{row3}),0)'
        )
        for c in range(1,9):
            cell=ws3.cell(row3,c); cell.border=BORDER; cell.alignment=CENTER
            cell.fill = F_LBLUE if row3%2==0 else F_LGRAY
        row3 += 1

widths(ws3,[17,9,15,22,15,17,20,12])

# CF on SLA compliance col G
ws3.conditional_formatting.add(f"G2:G{row3-1}",
    CellIsRule("lessThan",["70"], fill=F_RED, font=mk_font("FFFFFF",bold=True)))
ws3.conditional_formatting.add(f"G2:G{row3-1}",
    CellIsRule("between",["70","84.99"], fill=F_YLW, font=BL))
ws3.conditional_formatting.add(f"G2:G{row3-1}",
    CellIsRule("greaterThanOrEqual",["85"], fill=F_LGRN, font=mk_font("375623")))

# ══════════════════════════════════════════════
# SHEET 4: Dashboard
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("Dashboard")
ws4.sheet_properties.tabColor="C00000"

# ── Title ──
ws4.merge_cells("A1:N1")
ws4["A1"].value = "SUPPORT OPERATIONS  —  ADVANCED MIS REPORTING FRAMEWORK"
ws4["A1"].font  = Font(bold=True, size=18, color="FFFFFF")
ws4["A1"].fill  = F_NAVY; ws4["A1"].alignment = CENTER
ws4.row_dimensions[1].height = 42

ws4.merge_cells("A2:N2")
ws4["A2"].value = "180-Day Simulated Operations Dataset  |  Jan 1 – Jun 29, 2026  |  15 Agents  |  7 Categories"
ws4["A2"].font  = mk_font("1F3864", italic=True, sz=10)
ws4["A2"].alignment = CENTER

# ── KPI Header ──
ws4.merge_cells("A4:N4")
ws4["A4"].value="KEY PERFORMANCE INDICATORS"
ws4["A4"].font=Font(bold=True,size=12,color="FFFFFF")
ws4["A4"].fill=F_BLUE; ws4["A4"].alignment=CENTER

# 6 KPI boxes
KPI = [
    ("Total Tickets",      f"=COUNTA(Raw_Data!$A$2:$A${N+1})"),
    ("Overall SLA %",      f"=ROUND(COUNTIF(Raw_Data!$J$2:$J${N+1},\"Y\")/COUNTA(Raw_Data!$J$2:$J${N+1})*100,1)"),
    ("Avg Resolution Hrs", f"=ROUND(AVERAGE(Raw_Data!$F$2:$F${N+1}),1)"),
    ("Avg CSAT Score",     f"=ROUND(AVERAGE(Raw_Data!$H$2:$H${N+1}),2)"),
    ("Total Escalated",    f'=COUNTIF(Raw_Data!$G$2:$G${N+1},"Escalated")'),
    ("Active Agents",      f"={len(AGENTS)}"),
]
ws4.row_dimensions[5].height=24; ws4.row_dimensions[6].height=36
for idx,(lbl,fml) in enumerate(KPI):
    c1=idx*2+1; c2=c1+1
    l1=get_column_letter(c1); l2=get_column_letter(c2)
    ws4.merge_cells(f"{l1}5:{l2}5")
    lc=ws4[f"{l1}5"]; lc.value=lbl
    lc.font=Font(bold=True,size=9,color="FFFFFF"); lc.fill=F_BLUE
    lc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws4.merge_cells(f"{l1}6:{l2}6")
    vc=ws4[f"{l1}6"]; vc.value=fml
    vc.font=Font(bold=True,size=15,color="1F3864"); vc.fill=F_LBLUE; vc.alignment=CENTER

# ── Agent XLOOKUP selector ──
ws4.merge_cells("A8:D8")
ws4["A8"].value="AGENT SPOTLIGHT  (select agent in B9)"
ws4["A8"].font=Font(bold=True,size=11,color="FFFFFF"); ws4["A8"].fill=F_BLUE; ws4["A8"].alignment=CENTER

ws4["A9"].value="Selected Agent:"; ws4["A9"].font=mk_font("1F3864",bold=True)
ws4["B9"].value="AGENT_01"
ws4["B9"].fill=F_CREAM; ws4["B9"].border=BORDER; ws4["B9"].alignment=CENTER

# Data validation dropdown on B9
dv_agt = DataValidation(
    type="list",
    formula1='"'+",".join(AGENTS)+'"',
    allow_blank=True, showDropDown=False
)
ws4.add_data_validation(dv_agt); dv_agt.add("B9")

ws4["C9"].value="Ticket Count:";  ws4["C9"].font=mk_font("1F3864",bold=True)
ws4["D9"].value='=XLOOKUP(B9,Agent_Summary!$A$2:$A$16,Agent_Summary!$B$2:$B$16,"Not Found")'
ws4["D9"].fill=F_LBLUE; ws4["D9"].border=BORDER; ws4["D9"].alignment=CENTER

ws4["E9"].value="Avg Res Hrs:"; ws4["E9"].font=mk_font("1F3864",bold=True)
ws4["F9"].value='=XLOOKUP(B9,Agent_Summary!$A$2:$A$16,Agent_Summary!$C$2:$C$16,"Not Found")'
ws4["F9"].fill=F_LBLUE; ws4["F9"].border=BORDER; ws4["F9"].alignment=CENTER

ws4["G9"].value="SLA %:"; ws4["G9"].font=mk_font("1F3864",bold=True)
ws4["H9"].value='=XLOOKUP(B9,Agent_Summary!$A$2:$A$16,Agent_Summary!$F$2:$F$16,"Not Found")'
ws4["H9"].fill=F_LBLUE; ws4["H9"].border=BORDER; ws4["H9"].alignment=CENTER

ws4["I9"].value="Health:"; ws4["I9"].font=mk_font("1F3864",bold=True)
ws4["J9"].value='=XLOOKUP(B9,Agent_Summary!$A$2:$A$16,Agent_Summary!$J$2:$J$16,"Not Found")'
ws4["J9"].fill=F_LBLUE; ws4["J9"].border=BORDER; ws4["J9"].alignment=CENTER

# ── Category filter dropdown ──
ws4["A11"].value="Filter by Category:"; ws4["A11"].font=mk_font("1F3864",bold=True)
ws4["B11"].value="All Categories"
ws4["B11"].fill=F_CREAM; ws4["B11"].border=BORDER; ws4["B11"].alignment=CENTER

dv_cat = DataValidation(
    type="list",
    formula1='"All Categories,'+",".join(CATS)+'"',
    allow_blank=True, showDropDown=False
)
ws4.add_data_validation(dv_cat); dv_cat.add("B11")

# ── SLA Summary table (rows 13-14 header, 15-21 data) ──
ws4.merge_cells("A13:G13")
ws4["A13"].value="SLA COMPLIANCE BY CATEGORY"
ws4["A13"].font=Font(bold=True,size=11,color="FFFFFF"); ws4["A13"].fill=F_BLUE; ws4["A13"].alignment=CENTER

for ci,h in enumerate(["Category","Total Tickets","SLA Met","SLA Breach","SLA Compliance %","Avg Res Hrs","Avg CSAT"],1):
    cell=ws4.cell(14,ci); cell.value=h
    cell.font=WH; cell.fill=F_NAVY; cell.alignment=CENTER; cell.border=BORDER

for ri,cat in enumerate(CATS,15):
    ws4.cell(ri,1).value=cat
    ws4.cell(ri,2).value=f'=COUNTIF(Raw_Data!$C$2:$C${N+1},A{ri})'
    ws4.cell(ri,3).value=f'=COUNTIFS(Raw_Data!$C$2:$C${N+1},A{ri},Raw_Data!$J$2:$J${N+1},"Y")'
    ws4.cell(ri,4).value=f'=COUNTIFS(Raw_Data!$C$2:$C${N+1},A{ri},Raw_Data!$J$2:$J${N+1},"N")'
    ws4.cell(ri,5).value=f'=IFERROR(C{ri}/B{ri}*100,0)'
    ws4.cell(ri,6).value=f'=AVERAGEIF(Raw_Data!$C$2:$C${N+1},A{ri},Raw_Data!$F$2:$F${N+1})'
    ws4.cell(ri,7).value=f'=AVERAGEIF(Raw_Data!$C$2:$C${N+1},A{ri},Raw_Data!$H$2:$H${N+1})'
    for c in range(1,8):
        cell=ws4.cell(ri,c); cell.border=BORDER; cell.alignment=CENTER
        cell.fill=F_LBLUE if ri%2==0 else F_LGRAY

# CF on SLA compliance col E (rows 15-21)
ws4.conditional_formatting.add("E15:E21",
    CellIsRule("lessThan",["70"],fill=F_RED,font=mk_font("FFFFFF",bold=True)))
ws4.conditional_formatting.add("E15:E21",
    CellIsRule("between",["70","84.99"],fill=F_YLW,font=BL))
ws4.conditional_formatting.add("E15:E21",
    CellIsRule("greaterThanOrEqual",["85"],fill=F_LGRN,font=mk_font("375623")))

# ── Weekly chart data helper (rows 24-50) ──
ws4["A23"]="Week Label"; ws4["B23"]="Ticket Count"; ws4["C23"]="SLA Compliance %"
for c in [1,2,3]:
    ws4.cell(23,c).font=WH; ws4.cell(23,c).fill=F_NAVY
    ws4.cell(23,c).border=BORDER; ws4.cell(23,c).alignment=CENTER

unique_weeks = sorted(df["WeekDT"].unique())
for ri, wdt in enumerate(unique_weeks, 24):
    wlbl = wdt.strftime("%b %d")
    wend = wdt + timedelta(days=6)
    wend_str = wend.strftime("%Y-%m-%d")
    wdt_str  = wdt.strftime("%Y-%m-%d")
    ws4.cell(ri,1).value = wlbl
    ws4.cell(ri,2).value = (
        f'=COUNTIFS(Raw_Data!$B$2:$B${N+1},">="&"{wdt_str}",'
        f'Raw_Data!$B$2:$B${N+1},"<="&"{wend_str}")'
    )
    ws4.cell(ri,3).value = (
        f'=IFERROR(COUNTIFS(Raw_Data!$B$2:$B${N+1},">="&"{wdt_str}",'
        f'Raw_Data!$B$2:$B${N+1},"<="&"{wend_str}",'
        f'Raw_Data!$J$2:$J${N+1},"Y")'
        f'/COUNTIFS(Raw_Data!$B$2:$B${N+1},">="&"{wdt_str}",'
        f'Raw_Data!$B$2:$B${N+1},"<="&"{wend_str}")*100,0)'
    )
    for c in range(1,4):
        ws4.cell(ri,c).border=BORDER; ws4.cell(ri,c).alignment=CENTER

n_wks = len(unique_weeks)
last_wk_row = 23 + n_wks

# ── Line chart: weekly ticket volume ──
line_c = LineChart(); line_c.style=10; line_c.title="Weekly Ticket Volume"
line_c.y_axis.title="Tickets"; line_c.x_axis.title="Week"
vol_ref = Reference(ws4, min_col=2, max_col=2, min_row=23, max_row=last_wk_row)
cat_ref = Reference(ws4, min_col=1, min_row=24, max_row=last_wk_row)
line_c.add_data(vol_ref, titles_from_data=True)
line_c.set_categories(cat_ref)
line_c.width=22; line_c.height=13
ws4.add_chart(line_c,"I1")

# ── Line chart: weekly SLA compliance ──
line_s = LineChart(); line_s.style=10; line_s.title="Weekly SLA Compliance %"
line_s.y_axis.title="SLA %"; line_s.y_axis.scaling.min=60
sla_ref = Reference(ws4, min_col=3, max_col=3, min_row=23, max_row=last_wk_row)
line_s.add_data(sla_ref, titles_from_data=True)
line_s.set_categories(cat_ref)
line_s.width=22; line_s.height=13
ws4.add_chart(line_s,"I18")

widths(ws4,[16,14,12,12,14,12,12,12,14,14])

# ── Save workbook ──
XLSX="excel/advanced_excel_reporting_framework.xlsx"
wb.save(XLSX)
print(f"Excel saved -> {XLSX}")

# ─────────────────────────────────────────────
# 5.  MATPLOTLIB CHART IMAGES
# ─────────────────────────────────────────────
STYLE = {
    "axes.facecolor":   "#F8FAFC",
    "axes.edgecolor":   "#CCCCCC",
    "axes.grid":        True,
    "grid.color":       "#E5E5E5",
    "grid.linewidth":   0.8,
    "axes.titlesize":   14,
    "axes.titlecolor":  "#1F3864",
    "axes.titleweight": "bold",
    "axes.labelcolor":  "#444444",
    "xtick.color":      "#555555",
    "ytick.color":      "#555555",
    "figure.facecolor": "#FFFFFF",
}
plt.rcParams.update(STYLE)
C_BLUE="#2E75B6"; C_NAVY="#1F3864"; C_GREEN="#375623"
C_RED="#C00000";  C_ORG="#ED7D31";  C_LBLUE="#D9E1F2"

# ── Chart 02: Weekly ticket volume + SLA trend ──
fig, (ax1,ax2) = plt.subplots(2,1, figsize=(14,10), sharex=True)
fig.suptitle("Weekly Operations Trend (Jan–Jun 2026)", fontsize=16, color=C_NAVY,
             fontweight="bold", y=0.98)

ax1.fill_between(wk.index, wk.Tickets, alpha=0.15, color=C_BLUE)
ax1.plot(wk.index, wk.Tickets, color=C_BLUE, linewidth=2.2, marker="o",
         markersize=5, label="Weekly Tickets")
# Shade incident week
inc_weeks = wk[wk.WeekDT.apply(lambda x: any(
    di in INCIDENT for di in range(int((x - START).days), int((x - START).days)+7)
))]
if not inc_weeks.empty:
    ax1.axvspan(inc_weeks.index[0]-0.5, inc_weeks.index[-1]+0.5,
                alpha=0.15, color=C_RED, label="Incident Window (Mar 9-15)")
ax1.set_ylabel("Tickets per Week", fontsize=11); ax1.legend(fontsize=9)
ax1.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x,_: f"{int(x):,}"))

ax2.plot(wk.index, wk.SLA_Rate, color=C_GREEN, linewidth=2.2, marker="s",
         markersize=5, label="SLA Compliance %")
ax2.axhline(85, color=C_RED, linestyle="--", linewidth=1.2, label="Warning Threshold (85%)")
ax2.axhline(95, color=C_GREEN, linestyle="--", linewidth=1.0, alpha=0.5, label="Target (95%)")
if not inc_weeks.empty:
    ax2.axvspan(inc_weeks.index[0]-0.5, inc_weeks.index[-1]+0.5, alpha=0.15, color=C_RED)
ax2.set_ylim(60,105); ax2.set_ylabel("SLA Compliance %", fontsize=11)
ax2.set_xlabel("Week Number", fontsize=11); ax2.legend(fontsize=9)

step = max(1, len(wk)//12)
ax2.set_xticks(wk.index[::step])
ax2.set_xticklabels(wk.WkLabel[::step], rotation=45, ha="right", fontsize=8)

plt.tight_layout()
plt.savefig("images/02-trend-chart.png", dpi=150, bbox_inches="tight")
plt.close()
print("Image 02 saved")

# ── Chart 03: Category breakdown ──
fig, axes = plt.subplots(1,2, figsize=(16,7))
fig.suptitle("Category Performance Breakdown", fontsize=16, color=C_NAVY, fontweight="bold")

cats_sorted = cat_agg.sort_values("Tickets")
bars = axes[0].barh(cats_sorted.Category, cats_sorted.Tickets,
                     color=C_BLUE, edgecolor="white", linewidth=0.5)
axes[0].set_title("Total Tickets by Category"); axes[0].set_xlabel("Ticket Count")
for bar in bars:
    w=bar.get_width()
    axes[0].text(w+5, bar.get_y()+bar.get_height()/2,
                 f"{int(w):,}", va="center", fontsize=9)
axes[0].set_xlim(0, cats_sorted.Tickets.max()*1.15)

colors = [C_RED if v<70 else C_ORG if v<85 else C_GREEN for v in cat_agg.SLA]
bars2 = axes[1].barh(cat_agg.Category, cat_agg.SLA,
                      color=colors, edgecolor="white", linewidth=0.5)
axes[1].set_title("SLA Compliance % by Category"); axes[1].set_xlabel("SLA Compliance %")
axes[1].axvline(85, color=C_RED, linestyle="--", linewidth=1.2, label="Warning (85%)")
axes[1].axvline(95, color=C_GREEN, linestyle="--", linewidth=1.0, alpha=0.7, label="Target (95%)")
axes[1].set_xlim(0,110); axes[1].legend(fontsize=9)
for bar in bars2:
    w=bar.get_width()
    axes[1].text(w+0.5, bar.get_y()+bar.get_height()/2,
                 f"{w:.1f}%", va="center", fontsize=9)
# Annotation for Technical outlier
tech_sla = cat_agg[cat_agg.Category=="Technical"]["SLA"].values[0]
tech_y   = list(cat_agg.Category).index("Technical")
axes[1].annotate("Technical: lowest\nresolution SLA",
                  xy=(tech_sla, tech_y), xytext=(tech_sla-20, tech_y+1),
                  arrowprops=dict(arrowstyle="->",color=C_RED),
                  color=C_RED, fontsize=8, fontweight="bold")

plt.tight_layout()
plt.savefig("images/03-category-breakdown.png", dpi=150, bbox_inches="tight")
plt.close()
print("Image 03 saved")

# ── Chart 04: Agent performance ──
fig, axes = plt.subplots(1,2, figsize=(16,7))
fig.suptitle("Agent Performance Summary", fontsize=16, color=C_NAVY, fontweight="bold")

agt_sorted = agt_agg.sort_values("SLA")
colors_agt = [C_RED if v<75 else C_ORG if v<85 else C_GREEN for v in agt_sorted.SLA]
axes[0].barh(agt_sorted.Assigned_Agent, agt_sorted.SLA,
              color=colors_agt, edgecolor="white", linewidth=0.5)
axes[0].set_title("SLA Compliance % by Agent"); axes[0].set_xlabel("SLA Compliance %")
axes[0].axvline(85, color=C_RED, linestyle="--", linewidth=1.2, label="Warning (85%)")
axes[0].set_xlim(0,110); axes[0].legend(fontsize=9)
for i,(v,ag) in enumerate(zip(agt_sorted.SLA, agt_sorted.Assigned_Agent)):
    axes[0].text(v+0.5, i, f"{v:.1f}%", va="center", fontsize=8)

agt_rt = agt_agg.sort_values("Avg_RT", ascending=False)
colors_rt = [C_RED if v>20 else C_ORG if v>12 else C_GREEN for v in agt_rt.Avg_RT]
axes[1].barh(agt_rt.Assigned_Agent, agt_rt.Avg_RT,
              color=colors_rt, edgecolor="white", linewidth=0.5)
axes[1].set_title("Avg Resolution Time (hrs) by Agent"); axes[1].set_xlabel("Hours")
axes[1].axvline(15, color=C_RED, linestyle="--", linewidth=1.2, label="Warning (15h)")
axes[1].legend(fontsize=9)
for i,(v,ag) in enumerate(zip(agt_rt.Avg_RT, agt_rt.Assigned_Agent)):
    axes[1].text(v+0.1, i, f"{v:.1f}h", va="center", fontsize=8)

plt.tight_layout()
plt.savefig("images/04-agent-performance.png", dpi=150, bbox_inches="tight")
plt.close()
print("Image 04 saved")

# ── Chart 05: SLA compliance by priority ──
fig, ax = plt.subplots(figsize=(10,6))
fig.suptitle("SLA Compliance by Priority Tier", fontsize=16, color=C_NAVY, fontweight="bold")

colors_p = [C_RED if v<70 else C_ORG if v<85 else C_GREEN for v in pri_agg.SLA]
bars = ax.bar(pri_agg.Priority, pri_agg.SLA, color=colors_p,
               edgecolor="white", linewidth=0.5, width=0.5)
ax.axhline(95, color=C_GREEN, linestyle="--", linewidth=1.5, label="SLA Target (95%)")
ax.axhline(85, color=C_ORG,   linestyle="--", linewidth=1.2, label="Warning (85%)")
ax.set_ylim(0,110); ax.set_ylabel("SLA Compliance %", fontsize=12)
ax.set_xlabel("Priority Tier", fontsize=12); ax.legend(fontsize=10)

for bar in bars:
    h=bar.get_height()
    ax.text(bar.get_x()+bar.get_width()/2, h+1, f"{h:.1f}%",
            ha="center", va="bottom", fontsize=11, fontweight="bold")

# Ticket count annotations
for i,(p,tc) in enumerate(zip(pri_agg.Priority, pri_agg.Tickets)):
    ax.text(i, 5, f"n={tc:,}", ha="center", fontsize=9, color="white", fontweight="bold")

ax.set_title("P1 (Critical) tickets have the lowest SLA compliance\ndue to tight 4-hour resolution windows",
             fontsize=11, color="#555555", pad=10)

plt.tight_layout()
plt.savefig("images/05-sla-compliance.png", dpi=150, bbox_inches="tight")
plt.close()
print("Image 05 saved")

# ── Chart 06: Formula bar simulation ──
fig = plt.figure(figsize=(14, 4))
fig.patch.set_facecolor("#F0F0F0")
ax = fig.add_axes([0, 0, 1, 1])
ax.set_xlim(0, 14); ax.set_ylim(0, 4); ax.axis("off")

# Ribbon bar (top)
ribbon = mpatches.FancyBboxPatch((0,3.1), 14, 0.9, boxstyle="square,pad=0",
                                  facecolor="#D4D4D4", edgecolor="none")
ax.add_patch(ribbon)
ax.text(0.3, 3.55, "File    Home    Insert    Page Layout    Formulas    Data    Review    View",
        fontsize=8.5, va="center", color="#111111")

# Formula bar area
fbar_bg = mpatches.FancyBboxPatch((0,2.2), 14, 0.8, boxstyle="square,pad=0",
                                   facecolor="#FFFFFF", edgecolor="#ABABAB", linewidth=0.8)
ax.add_patch(fbar_bg)

# Name box (cell address)
nb = mpatches.FancyBboxPatch((0.05,2.28), 1.3, 0.64, boxstyle="square,pad=0",
                               facecolor="#FFFFFF", edgecolor="#ABABAB", linewidth=1.2)
ax.add_patch(nb)
ax.text(0.7, 2.60, "D9", fontsize=11, ha="center", va="center")

# Divider
ax.axvline(1.45, ymin=0.55, ymax=0.73, color="#ABABAB", linewidth=1.2)

# fx button
fx_bg = mpatches.FancyBboxPatch((1.5,2.28), 0.5, 0.64, boxstyle="square,pad=0",
                                  facecolor="#EBEBEB", edgecolor="#ABABAB", linewidth=1.2)
ax.add_patch(fx_bg)
ax.text(1.75, 2.60, "fx", fontsize=12, ha="center", va="center",
        fontstyle="italic", fontfamily="serif")

# Formula text
formula_txt = '=XLOOKUP(B9, Agent_Summary!$A$2:$A$16, Agent_Summary!$B$2:$B$16, "Not Found")'
ax.text(2.1, 2.60, formula_txt, fontsize=11.5, va="center",
        fontfamily="monospace", color="#1F1F1F")

# Spreadsheet area (mock grid)
grid_bg = mpatches.FancyBboxPatch((0, 0), 14, 2.1, boxstyle="square,pad=0",
                                    facecolor="#FFFFFF", edgecolor="none")
ax.add_patch(grid_bg)

# Column headers
col_headers = ["","A","B","C","D","E","F","G","H","I","J"]
col_xs = [0, 0.9, 2.2, 3.8, 5.2, 6.6, 8.0, 9.2, 10.4, 11.6, 12.8]
for ci, (ch, cx) in enumerate(zip(col_headers, col_xs)):
    hdr_c = mpatches.FancyBboxPatch((col_xs[ci],1.75), col_xs[ci+1]-col_xs[ci] if ci<len(col_xs)-1 else 1.2,
                                     0.28, boxstyle="square,pad=0",
                                     facecolor="#2E75B6" if ch=="D" else "#E4E4E4",
                                     edgecolor="#CCCCCC", linewidth=0.5)
    if ci < len(col_headers)-1:
        ax.add_patch(hdr_c)
    ax.text(cx + (col_xs[ci+1]-cx)/2 if ci < len(col_xs)-1 else cx+0.6, 1.89, ch,
            fontsize=9, ha="center", va="center",
            color="white" if ch=="D" else "#555555",
            fontweight="bold" if ch=="D" else "normal")

# Row headers + cell contents
row_data = [
    ("8","Selected Agent:", "AGENT_01", "Ticket Count:", "", ""),
    ("9","", "AGENT_01", "", "15 ←XLOOKUP result", ""),
]
row_ys = [1.2, 0.6]
row_fill = ["#FFFFFF","#EEF4FF"]

for ri,(rlbl,*cells) in enumerate(row_data):
    ry = row_ys[ri]
    row_rect = mpatches.FancyBboxPatch((0, ry-0.25), 14, 0.55, boxstyle="square,pad=0",
                                        facecolor=row_fill[ri], edgecolor="none")
    ax.add_patch(row_rect)
    # Row number
    rn_bg = mpatches.FancyBboxPatch((0, ry-0.25), 0.88, 0.55, boxstyle="square,pad=0",
                                      facecolor="#1F3864" if ri==1 else "#E4E4E4",
                                      edgecolor="#CCCCCC", linewidth=0.5)
    ax.add_patch(rn_bg)
    ax.text(0.44, ry, rlbl, fontsize=9, ha="center", va="center",
            color="white" if ri==1 else "#555", fontweight="bold" if ri==1 else "normal")
    # Cell grid lines
    for ci,cx in enumerate(col_xs[:-1]):
        ax.plot([cx, cx], [ry-0.25, ry+0.30], color="#CCCCCC", linewidth=0.5)
    # Cell values
    cell_vals = [rlbl] + list(cells)
    cell_texts = ["Selected Agent:", "AGENT_01", "Ticket Count:", "15", "", ""]
    if ri==1:
        cell_texts = ["", "AGENT_01", "", "15  ← XLOOKUP result", "", ""]
    for ci,(ctext,cx) in enumerate(zip(cell_texts, col_xs[1:])):
        highlight = (ri==1 and ci==2)  # D9 cell
        if highlight:
            hl = mpatches.FancyBboxPatch((col_xs[ci+1], ry-0.25),
                                          col_xs[ci+2]-col_xs[ci+1] if ci+2<len(col_xs) else 1.2,
                                          0.55, boxstyle="square,pad=0",
                                          facecolor="#D9E1F2", edgecolor="#2E75B6", linewidth=2)
            ax.add_patch(hl)
        ax.text(cx+0.1, ry, ctext, fontsize=9, va="center",
                color="#1F3864" if highlight else "#333333",
                fontweight="bold" if highlight else "normal")

ax.text(7, -0.15, "Cell D9 in the Dashboard sheet contains the XLOOKUP formula shown in the formula bar above",
        fontsize=9, ha="center", color="#555555", style="italic")

plt.savefig("images/06-formula-xlookup.png", dpi=150, bbox_inches="tight",
            facecolor="#F0F0F0")
plt.close()
print("Image 06 saved")

# ── Chart 07: Conditional formatting table simulation ──
fig, ax = plt.subplots(figsize=(14, 6))
fig.suptitle("Dashboard — SLA Compliance Table with Conditional Formatting",
             fontsize=14, color=C_NAVY, fontweight="bold")
ax.axis("off")

# Build table data
table_cats = CATS
table_data = []
for cat in table_cats:
    sub = df[df.Category==cat]
    n   = len(sub)
    sla = (sub.SLA_Met=="Y").mean()*100
    rt  = sub.Resolution_Time_Hrs.mean()
    cs  = sub.CSAT_Score.mean()
    # Health status
    health = "Critical" if sla<70 else ("Warning" if sla<85 else "Healthy")
    table_data.append([cat, f"{n:,}", f"{sla:.1f}%", f"{rt:.1f}h", f"{cs:.2f}", health])

col_lbls = ["Category","Tickets","SLA Compliance %","Avg Res Time","Avg CSAT","Health"]
n_rows   = len(table_data)
n_cols   = len(col_lbls)

col_widths_tbl = [0.22, 0.10, 0.18, 0.14, 0.12, 0.14]
row_height = 0.12
start_x, start_y = 0.02, 0.85

# Draw header
for ci,(lbl,cw) in enumerate(zip(col_lbls, col_widths_tbl)):
    x = start_x + sum(col_widths_tbl[:ci])
    rect = mpatches.FancyBboxPatch((x, start_y), cw-0.005, row_height,
                                    boxstyle="square,pad=0",
                                    facecolor="#1F3864", edgecolor="white", linewidth=1.5)
    ax.add_patch(rect)
    ax.text(x+cw/2-0.002, start_y+row_height/2, lbl,
            ha="center", va="center", fontsize=9, color="white", fontweight="bold",
            transform=ax.transAxes)

# Draw data rows
for ri, row in enumerate(table_data):
    y = start_y - (ri+1) * (row_height + 0.01)
    for ci,(val,cw) in enumerate(zip(row, col_widths_tbl)):
        x = start_x + sum(col_widths_tbl[:ci])
        # Color logic
        if ci == 2:  # SLA column
            sla_val = float(val.replace("%",""))
            fc = "#FFCCCC" if sla_val<70 else ("#FFFF99" if sla_val<85 else "#92D050")
            tc = "#C00000" if sla_val<70 else ("#555500" if sla_val<85 else "#375623")
        elif ci == 5:  # Health column
            fc = "#FFCCCC" if val=="Critical" else ("#FFFF99" if val=="Warning" else "#92D050")
            tc = "#C00000" if val=="Critical" else ("#555500" if val=="Warning" else "#375623")
        else:
            fc = "#D9E1F2" if ri%2==0 else "#FFFFFF"
            tc = "#1F1F1F"
        rect = mpatches.FancyBboxPatch((x, y), cw-0.005, row_height,
                                        boxstyle="square,pad=0",
                                        facecolor=fc, edgecolor="#DDDDDD", linewidth=0.8)
        ax.add_patch(rect)
        ax.text(x+cw/2-0.002, y+row_height/2, val,
                ha="center", va="center", fontsize=9, color=tc,
                fontweight="bold" if ci in [2,5] else "normal",
                transform=ax.transAxes)

# Legend
legend_items = [
    mpatches.Patch(facecolor="#92D050", edgecolor="#375623", label="Healthy (SLA ≥ 85%)"),
    mpatches.Patch(facecolor="#FFFF99", edgecolor="#555500", label="Warning (70% ≤ SLA < 85%)"),
    mpatches.Patch(facecolor="#FFCCCC", edgecolor="#C00000", label="Critical (SLA < 70%)"),
]
ax.legend(handles=legend_items, loc="lower right", fontsize=9,
          framealpha=0.9, edgecolor="#CCCCCC")

plt.tight_layout()
plt.savefig("images/07-conditional-formatting.png", dpi=150, bbox_inches="tight")
plt.close()
print("Image 07 saved")

# ── Chart 01: Dashboard overview (composite) ──
fig = plt.figure(figsize=(18, 12))
fig.patch.set_facecolor("#F8FAFC")
gs  = GridSpec(3, 3, figure=fig, hspace=0.4, wspace=0.35)
fig.suptitle("Support Operations MIS Dashboard  —  Jan–Jun 2026",
             fontsize=18, color=C_NAVY, fontweight="bold", y=0.98)

# KPI boxes (top row)
kpi_ax = fig.add_subplot(gs[0,:])
kpi_ax.axis("off")
kpis = [
    ("Total Tickets",   f"{total_tickets:,}",  C_BLUE),
    ("Overall SLA",     f"{overall_sla:.1f}%",  C_GREEN if overall_sla>=85 else C_ORG),
    ("Avg Resolution",  f"{avg_rt:.1f} hrs",    C_NAVY),
    ("Avg CSAT",        f"{avg_csat:.2f}/5",    C_GREEN),
    ("Agents Active",   f"{len(AGENTS)}",        C_NAVY),
    ("Categories",      f"{len(CATS)}",          C_BLUE),
]
for ki,(lbl,val,clr) in enumerate(kpis):
    x0 = ki/len(kpis); x1 = (ki+1)/len(kpis)
    card = mpatches.FancyBboxPatch((x0+0.005, 0.05), x1-x0-0.01, 0.9,
                                    boxstyle="round,pad=0.01",
                                    facecolor=C_LBLUE, edgecolor=clr, linewidth=2,
                                    transform=kpi_ax.transAxes)
    kpi_ax.add_patch(card)
    kpi_ax.text((x0+x1)/2, 0.68, lbl, ha="center", fontsize=9, color="#555555",
                transform=kpi_ax.transAxes)
    kpi_ax.text((x0+x1)/2, 0.32, val, ha="center", fontsize=18, color=clr,
                fontweight="bold", transform=kpi_ax.transAxes)

# Trend chart (middle left)
ax_trend = fig.add_subplot(gs[1,:2])
ax_trend.fill_between(wk.index, wk.Tickets, alpha=0.12, color=C_BLUE)
ax_trend.plot(wk.index, wk.Tickets, color=C_BLUE, linewidth=2, marker="o", markersize=4)
inc_idx = wk[wk.WeekDT.apply(lambda x: any(
    di in INCIDENT for di in range(int((x-START).days), int((x-START).days)+7)
))].index
if len(inc_idx)>0:
    ax_trend.axvspan(inc_idx[0]-0.5, inc_idx[-1]+0.5, alpha=0.15, color=C_RED, label="Incident")
ax_trend.set_title("Weekly Ticket Volume"); ax_trend.set_ylabel("Tickets")
ax_trend.set_xticks(wk.index[::4]); ax_trend.set_xticklabels(wk.WkLabel[::4], rotation=45, fontsize=7)
ax_trend.legend(fontsize=8)

# Category SLA (middle right)
ax_cat = fig.add_subplot(gs[1,2])
colors_c = [C_RED if v<70 else C_ORG if v<85 else C_GREEN for v in cat_agg.SLA]
ax_cat.barh(cat_agg.Category, cat_agg.SLA, color=colors_c, edgecolor="white")
ax_cat.axvline(85, color=C_RED, linestyle="--", linewidth=1)
ax_cat.set_title("Category SLA %"); ax_cat.set_xlim(0,110)
ax_cat.set_xlabel("SLA Compliance %")

# Agent performance (bottom left)
ax_agt = fig.add_subplot(gs[2,0])
colors_a = [C_RED if v<75 else C_ORG if v<85 else C_GREEN for v in agt_sorted.SLA]
ax_agt.barh(agt_sorted.Assigned_Agent, agt_sorted.SLA, color=colors_a, edgecolor="white")
ax_agt.axvline(85, color=C_RED, linestyle="--", linewidth=1)
ax_agt.set_title("Agent SLA %"); ax_agt.set_xlim(0,110)
ax_agt.tick_params(axis="y", labelsize=7)

# Priority SLA (bottom middle)
ax_pri = fig.add_subplot(gs[2,1])
colors_pr = [C_RED if v<70 else C_ORG if v<85 else C_GREEN for v in pri_agg.SLA]
ax_pri.bar(pri_agg.Priority, pri_agg.SLA, color=colors_pr, edgecolor="white", width=0.5)
ax_pri.axhline(95, color=C_GREEN, linestyle="--", linewidth=1, label="Target")
ax_pri.axhline(85, color=C_RED,   linestyle="--", linewidth=1, label="Warning")
ax_pri.set_title("Priority SLA %"); ax_pri.set_ylabel("%"); ax_pri.legend(fontsize=7)
ax_pri.set_ylim(0,110)

# SLA trend (bottom right)
ax_sla = fig.add_subplot(gs[2,2])
ax_sla.plot(wk.index, wk.SLA_Rate, color=C_GREEN, linewidth=2, marker="s", markersize=3)
ax_sla.axhline(85, color=C_RED,   linestyle="--", linewidth=1, label="Warning")
ax_sla.axhline(95, color=C_GREEN, linestyle="--", linewidth=1, alpha=0.5, label="Target")
ax_sla.set_ylim(60,105); ax_sla.set_title("Weekly SLA % Trend")
ax_sla.set_ylabel("%"); ax_sla.legend(fontsize=7)
ax_sla.set_xticks(wk.index[::4]); ax_sla.set_xticklabels(wk.WkLabel[::4], rotation=45, fontsize=7)

plt.savefig("images/01-dashboard-overview.png", dpi=150, bbox_inches="tight")
plt.close()
print("Image 01 saved")

# ── Print summary stats for README ──
print("\n=== KEY STATS ===")
print(f"Total tickets: {total_tickets:,}")
print(f"Overall SLA: {overall_sla:.1f}%")
print(f"Avg resolution time: {avg_rt:.1f} hrs")
print(f"Avg CSAT: {avg_csat:.2f}")
print(f"Technical SLA: {cat_agg[cat_agg.Category=='Technical']['SLA'].values[0]:.1f}%")
print(f"Incident week lowest SLA: {wk.SLA_Rate.min():.1f}%  (Week {wk.SLA_Rate.idxmin()+1})")
print(f"Normal week avg SLA: {wk[~wk.WeekDT.apply(lambda x: any(di in INCIDENT for di in range(int((x-START).days), int((x-START).days)+7)))].SLA_Rate.mean():.1f}%")

print("\nAll done.")
