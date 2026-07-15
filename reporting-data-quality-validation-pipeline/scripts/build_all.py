import matplotlib
matplotlib.use("Agg")

import os, random
import numpy as np
import pandas as pd
from datetime import date, timedelta, datetime
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

rng = np.random.default_rng(42)
random.seed(42)

BASE   = r"D:\reporting-data-quality-validation-pipeline"
D_DIR  = os.path.join(BASE, "data")
E_DIR  = os.path.join(BASE, "excel")
I_DIR  = os.path.join(BASE, "images")
S_DIR  = os.path.join(BASE, "scripts")

# ─── generate clean base data ──────────────────────────────────────────────────
print("[1/5] Generating base dataset ...")

N_BASE = 9_660

# master reference data
CUST_IDS_MASTER = [f"CUST_{i:05d}" for i in range(1, 401)]   # 400 valid customers
CHANNELS   = ["UPI", "NEFT", "IMPS", "Card", "Cash", "Net Banking"]
CATEGORIES = ["Retail", "Food & Beverage", "Travel", "Utilities", "Healthcare",
               "Entertainment", "Education", "Transfers"]
REGIONS    = ["North", "South", "East", "West", "Central"]

start_dt = date(2025, 7, 1)
all_dates = [start_dt + timedelta(days=d) for d in range(365)]

base_rows = []
for i in range(N_BASE):
    d = rng.choice(all_dates)
    amt = round(float(rng.lognormal(6.5, 1.0)), 2)
    base_rows.append({
        "Transaction_ID": f"TX_{i+1:07d}",
        "Customer_ID":    rng.choice(CUST_IDS_MASTER),
        "Date":           str(d),
        "Amount":         amt,
        "Channel":        rng.choice(CHANNELS),
        "Category":       rng.choice(CATEGORIES),
        "Region":         rng.choice(REGIONS),
        "Transaction_Status": rng.choice(["Success","Success","Success","Failed","Pending"]),
        "Reference_ID":   f"REF_{rng.integers(100000, 999999)}",
    })

base_df = pd.DataFrame(base_rows)

# ─── seed issues ──────────────────────────────────────────────────────────────
print("[2/5] Seeding data quality issues ...")

messy_df = base_df.copy()

# Issue 1: Duplicate records — rows 0-339 appended again (340 dupes)
N_DUPES = 340
dup_rows = base_df.iloc[:N_DUPES].copy()
# slight variation: keep same TX ID on some to make them exact dupes
messy_df = pd.concat([messy_df, dup_rows], ignore_index=True)
TOTAL_WITH_DUPES = len(messy_df)  # 10,000

# Issue 2: Missing values — rows 340-789: Amount = NaN
IDX_MISSING_AMT = list(range(340, 790))
messy_df.loc[IDX_MISSING_AMT, "Amount"] = np.nan
N_MISSING_AMT = len(IDX_MISSING_AMT)

# Issue 3: Missing Customer_ID — rows 790-989
IDX_MISSING_CID = list(range(790, 990))
messy_df.loc[IDX_MISSING_CID, "Customer_ID"] = np.nan
N_MISSING_CID = len(IDX_MISSING_CID)

# Issue 4: Inconsistent date formats — rows 990-1489
date_formats = [
    lambda d: datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y"),   # DD/MM/YYYY
    lambda d: datetime.strptime(d, "%Y-%m-%d").strftime("%m-%d-%Y"),   # MM-DD-YYYY
    lambda d: datetime.strptime(d, "%Y-%m-%d").strftime("%d-%b-%Y"),   # DD-Mon-YYYY
    lambda d: datetime.strptime(d, "%Y-%m-%d").strftime("%B %d, %Y"),  # Month DD, YYYY
]
IDX_DATE_FMT = list(range(990, 1490))
for i, idx in enumerate(IDX_DATE_FMT):
    orig_date = messy_df.loc[idx, "Date"]
    if pd.notna(orig_date):
        fmt_fn = date_formats[i % len(date_formats)]
        try:
            messy_df.loc[idx, "Date"] = fmt_fn(str(orig_date))
        except Exception:
            pass
N_DATE_FMT = len(IDX_DATE_FMT)

# Issue 5: Negative amounts (outlier values) — rows 1490-1539
IDX_NEG_AMT = list(range(1490, 1540))
messy_df.loc[IDX_NEG_AMT, "Amount"] = messy_df.loc[IDX_NEG_AMT, "Amount"].apply(
    lambda x: -abs(x) if pd.notna(x) else x)
N_NEG_AMT = len(IDX_NEG_AMT)

# Issue 6: Category label variants — rows 1540-1939
category_variants = {
    "UPI":         ["upi", "U.P.I.", "Upi", "UPI ", " UPI", "u.p.i"],
    "NEFT":        ["neft", "N.E.F.T", "Neft"],
    "Card":        ["card", "CARD ", "Card "],
    "Cash":        ["cash", "CASH", "Cash "],
}
IDX_CAT_VAR = list(range(1540, 1940))
for i, idx in enumerate(IDX_CAT_VAR):
    orig_chan = messy_df.loc[idx, "Channel"]
    if orig_chan in category_variants:
        variants = category_variants[orig_chan]
        messy_df.loc[idx, "Channel"] = rng.choice(variants)
    else:
        messy_df.loc[idx, "Channel"] = orig_chan.lower()
N_CAT_VAR = len(IDX_CAT_VAR)

# Issue 7: Orphaned foreign keys — rows 1940-2019 (Customer_IDs not in master)
orphan_cids = [f"CUST_ORPH_{i:03d}" for i in range(1, 81)]
IDX_ORPHAN = list(range(1940, 2020))
for i, idx in enumerate(IDX_ORPHAN):
    messy_df.loc[idx, "Customer_ID"] = orphan_cids[i % len(orphan_cids)]
N_ORPHAN = len(IDX_ORPHAN)

# shuffle so issues are not obviously in sequence
messy_df = messy_df.sample(frac=1, random_state=42).reset_index(drop=True)

messy_df.to_csv(os.path.join(D_DIR, "raw_mis_source_data.csv"), index=False)
print(f"   -> raw_mis_source_data.csv  ({len(messy_df)} rows)")
print(f"      Issues seeded: {N_DUPES} dupes | {N_MISSING_AMT} missing amt | "
      f"{N_MISSING_CID} missing CID | {N_DATE_FMT} date variants | "
      f"{N_NEG_AMT} negatives | {N_CAT_VAR} channel variants | {N_ORPHAN} orphan FKs")

# ─── validation pipeline ──────────────────────────────────────────────────────
print("[3/5] Running validation pipeline ...")

audit_log = []

def log(rule, before, after, note=""):
    audit_log.append({"Rule": rule, "Before_Count": before,
                       "After_Count": after, "Records_Affected": before - after, "Note": note})

def profile(df, stage=""):
    n = len(df)
    dup_count         = df.duplicated(subset=["Transaction_ID","Customer_ID","Date","Amount","Channel"]).sum()
    missing_amt       = df["Amount"].isna().sum()
    missing_cid       = df["Customer_ID"].isna().sum()
    negative_amt      = (df["Amount"].dropna() < 0).sum()
    # channel variants: any non-standard value
    std_channels = set(CHANNELS)
    bad_channel   = (~df["Channel"].isin(std_channels)).sum()
    # orphaned FKs: customer not in master
    valid_mask    = df["Customer_ID"].isin(CUST_IDS_MASTER) | df["Customer_ID"].isna()
    orphan_fk     = (~valid_mask).sum()
    # non-standard dates: anything that fails strict YYYY-MM-DD
    def is_std_date(v):
        try:
            datetime.strptime(str(v), "%Y-%m-%d")
            return True
        except Exception:
            return False
    bad_date = (~df["Date"].apply(is_std_date)).sum()

    total_issues = dup_count + missing_amt + missing_cid + negative_amt + bad_channel + orphan_fk + bad_date
    dq_score = max(0, round((1 - total_issues / max(n, 1)) * 100, 1))
    return {
        "Stage": stage, "Total_Rows": n,
        "Duplicate_Records": int(dup_count),
        "Missing_Amount": int(missing_amt),
        "Missing_Customer_ID": int(missing_cid),
        "Negative_Amount": int(negative_amt),
        "Channel_Label_Variants": int(bad_channel),
        "Orphaned_FK": int(orphan_fk),
        "Non_Standard_Dates": int(bad_date),
        "Total_Issues": int(total_issues),
        "DQ_Score_%": dq_score,
    }

before_profile = profile(messy_df, "Before Cleaning")
print(f"   BEFORE: {before_profile['Total_Rows']} rows | DQ Score: {before_profile['DQ_Score_%']}%")

clean_df = messy_df.copy()

# Rule 1: Remove exact duplicate records
before_n = len(clean_df)
clean_df = clean_df.drop_duplicates(
    subset=["Transaction_ID","Customer_ID","Date","Amount","Channel"], keep="first")
log("Remove duplicate records", before_n, len(clean_df), "Kept first occurrence of each duplicate set")
print(f"   Rule 1 (dedup): {before_n} -> {len(clean_df)} rows")

# Rule 2: Standardize date formats to YYYY-MM-DD
def standardize_date(v):
    if pd.isna(v):
        return v
    v = str(v).strip()
    formats = ["%Y-%m-%d","%d/%m/%Y","%m-%d-%Y","%d-%b-%Y","%B %d, %Y","%d-%B-%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return v   # return as-is if unparseable

bad_dates_before = int((~clean_df["Date"].apply(lambda v: bool(
    pd.notna(v) and len(str(v)) == 10 and str(v)[4] == "-" and str(v)[7] == "-"
))).sum())
clean_df["Date"] = clean_df["Date"].apply(standardize_date)
bad_dates_after = int((~clean_df["Date"].apply(lambda v: bool(
    pd.notna(v) and len(str(v)) == 10 and str(v)[4] == "-" and str(v)[7] == "-"
))).sum())
log("Standardize date formats", bad_dates_before, bad_dates_after,
    "Parsed 4 non-standard formats; converted to YYYY-MM-DD")
print(f"   Rule 2 (dates): {bad_dates_before} non-standard -> {bad_dates_after}")

# Rule 3: Standardize channel/category label variants
channel_map = {}
std_channels = set(CHANNELS)
for chan in clean_df["Channel"].dropna().unique():
    if chan in std_channels:
        continue
    chan_clean = chan.strip().upper().replace(".", "").replace(" ", "")
    for std in std_channels:
        if std.upper().replace(".", "").replace(" ", "") == chan_clean:
            channel_map[chan] = std
            break
    else:
        channel_map[chan] = chan  # no match; leave as-is

bad_chan_before = int((~clean_df["Channel"].isin(std_channels)).sum())
clean_df["Channel"] = clean_df["Channel"].map(lambda x: channel_map.get(x, x) if pd.notna(x) else x)
bad_chan_after = int((~clean_df["Channel"].isin(std_channels)).sum())
log("Standardize channel labels", bad_chan_before, bad_chan_after,
    "Normalized case/punctuation variants to canonical labels (UPI/NEFT/Card/Cash/IMPS/Net Banking)")
print(f"   Rule 3 (channels): {bad_chan_before} variants -> {bad_chan_after}")

# Rule 4: Flag missing Amount (do not remove — flag for investigation)
clean_df["Missing_Amount_Flag"] = clean_df["Amount"].isna().astype(int)
n_missing_flag = int(clean_df["Missing_Amount_Flag"].sum())
log("Flag missing Amount", n_missing_flag, n_missing_flag,
    "Flagged (not removed) — requires source-system reconciliation")
print(f"   Rule 4 (missing amt): {n_missing_flag} records flagged")

# Rule 5: Flag missing Customer_ID
clean_df["Missing_CID_Flag"] = clean_df["Customer_ID"].isna().astype(int)
n_missing_cid = int(clean_df["Missing_CID_Flag"].sum())
log("Flag missing Customer_ID", n_missing_cid, n_missing_cid,
    "Flagged — anonymous transactions require manual attribution")
print(f"   Rule 5 (missing CID): {n_missing_cid} records flagged")

# Rule 6: Flag negative amounts
clean_df["Negative_Amount_Flag"] = (clean_df["Amount"].dropna() < 0).astype(int)
clean_df["Negative_Amount_Flag"] = clean_df["Negative_Amount_Flag"].fillna(0).astype(int)
n_neg = int(clean_df["Negative_Amount_Flag"].sum())
log("Flag negative amounts", n_neg, n_neg,
    "Flagged as outliers — may be reversals; requires accounting review")
print(f"   Rule 6 (negatives): {n_neg} records flagged")

# Rule 7: Flag orphaned foreign keys
clean_df["Orphaned_FK_Flag"] = (~clean_df["Customer_ID"].isin(CUST_IDS_MASTER) &
                                  clean_df["Customer_ID"].notna()).astype(int)
n_orphan = int(clean_df["Orphaned_FK_Flag"].sum())
log("Flag orphaned Customer_ID foreign keys", n_orphan, n_orphan,
    "Customer_IDs not found in customer master — requires customer onboarding verification")
print(f"   Rule 7 (orphan FKs): {n_orphan} records flagged")

# save cleaned file
clean_df.to_csv(os.path.join(D_DIR, "cleaned_mis_data.csv"), index=False)
after_profile = profile(clean_df, "After Cleaning")
print(f"   AFTER:  {after_profile['Total_Rows']} rows | DQ Score: {after_profile['DQ_Score_%']}%")

# audit log
audit_df = pd.DataFrame(audit_log)
audit_df.to_csv(os.path.join(D_DIR, "audit_log.csv"), index=False)
print(f"   -> cleaned_mis_data.csv saved  |  audit_log.csv saved")

# ─── Excel workbook ───────────────────────────────────────────────────────────
print("[4/5] Building Excel before/after report ...")

wb = Workbook()
wb.remove(wb.active)

HDR_FILL  = PatternFill("solid", fgColor="1B3A6B")
HDR2_FILL = PatternFill("solid", fgColor="2E5E9E")
RED_FILL  = PatternFill("solid", fgColor="FFC7CE")
YEL_FILL  = PatternFill("solid", fgColor="FFEB9C")
GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")
ALT_FILL  = PatternFill("solid", fgColor="EEF2FA")
WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
SM_FONT   = Font(size=9)
thin      = Side(style="thin", color="B0BAC9")
med       = Side(style="medium", color="1B3A6B")
def B(): return Border(top=thin,left=thin,right=thin,bottom=thin)
CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left",   vertical="center")

# ── Sheet 1: DQ_Summary ────────────────────────────────────────────────────────
ws1 = wb.create_sheet("DQ_Summary")
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:H1")
t = ws1.cell(1, 1, "Data Quality Improvement Scorecard — Before vs After Cleaning")
t.fill = HDR_FILL; t.font = Font(bold=True, color="FFFFFF", size=13)
t.alignment = CTR
ws1.row_dimensions[1].height = 30

ws1.cell(3, 1, f"Source file: raw_mis_source_data.csv").font = SM_FONT
ws1.cell(4, 1, f"Cleaned file: cleaned_mis_data.csv").font = SM_FONT
ws1.cell(3, 5, f"Report generated: {date.today()}").font = SM_FONT

cols = ["Issue Type","Before Count","Before %","After Count","After %","Change","Status","Severity"]
for c, v in enumerate(cols, 1):
    cell = ws1.cell(6, c, v)
    cell.fill = HDR2_FILL; cell.font = HDR_FONT
    cell.alignment = CTR; cell.border = B()

rows_data = [
    ("Duplicate Records",
     before_profile["Duplicate_Records"], before_profile["Total_Rows"],
     after_profile["Duplicate_Records"],  after_profile["Total_Rows"],
     "Resolved", "High"),
    ("Missing Amount",
     before_profile["Missing_Amount"], before_profile["Total_Rows"],
     after_profile["Missing_Amount"],  after_profile["Total_Rows"],
     "Flagged", "Medium"),
    ("Missing Customer ID",
     before_profile["Missing_Customer_ID"], before_profile["Total_Rows"],
     after_profile["Missing_Customer_ID"],  after_profile["Total_Rows"],
     "Flagged", "High"),
    ("Non-Standard Date Formats",
     before_profile["Non_Standard_Dates"], before_profile["Total_Rows"],
     after_profile["Non_Standard_Dates"],  after_profile["Total_Rows"],
     "Resolved", "Medium"),
    ("Negative Amounts",
     before_profile["Negative_Amount"], before_profile["Total_Rows"],
     after_profile["Negative_Amount"],  after_profile["Total_Rows"],
     "Flagged", "Low"),
    ("Channel Label Variants",
     before_profile["Channel_Label_Variants"], before_profile["Total_Rows"],
     after_profile["Channel_Label_Variants"],  after_profile["Total_Rows"],
     "Resolved", "Medium"),
    ("Orphaned Foreign Keys",
     before_profile["Orphaned_FK"], before_profile["Total_Rows"],
     after_profile["Orphaned_FK"],  after_profile["Total_Rows"],
     "Flagged", "High"),
]

for r_off, (issue, b_cnt, b_tot, a_cnt, a_tot, status, severity) in enumerate(rows_data, 7):
    b_pct = b_cnt / b_tot if b_tot > 0 else 0
    a_pct = a_cnt / a_tot if a_tot > 0 else 0
    change = a_cnt - b_cnt

    ws1.cell(r_off, 1, issue).font      = SM_FONT
    ws1.cell(r_off, 2, b_cnt).font      = SM_FONT
    ws1.cell(r_off, 3, b_pct).font      = SM_FONT
    ws1.cell(r_off, 4, a_cnt).font      = SM_FONT
    ws1.cell(r_off, 5, a_pct).font      = SM_FONT
    ws1.cell(r_off, 6, change).font     = BOLD_FONT
    ws1.cell(r_off, 7, status).font     = BOLD_FONT
    ws1.cell(r_off, 8, severity).font   = BOLD_FONT

    ws1.cell(r_off, 3).number_format = "0.00%"
    ws1.cell(r_off, 5).number_format = "0.00%"

    fill = ALT_FILL if r_off % 2 == 0 else WHT_FILL
    for c in range(1, 9):
        ws1.cell(r_off, c).border = B()
        ws1.cell(r_off, c).fill   = fill
        ws1.cell(r_off, c).alignment = CTR if c > 1 else LFT

# DQ Score row
score_row = 7 + len(rows_data)
ws1.merge_cells(f"A{score_row}:A{score_row}")
ws1.cell(score_row, 1, "Overall DQ Score").font = Font(bold=True, size=10, color="1B3A6B")
ws1.cell(score_row, 2, f"{before_profile['DQ_Score_%']}%").font = Font(bold=True, size=14, color="C0392B")
ws1.cell(score_row, 4, f"{after_profile['DQ_Score_%']}%").font  = Font(bold=True, size=14, color="27AE60")
ws1.cell(score_row, 6, f"+{after_profile['DQ_Score_%'] - before_profile['DQ_Score_%']:.1f}pp").font = Font(bold=True, size=12, color="1B3A6B")
for c in range(1, 9):
    ws1.cell(score_row, c).border = Border(top=med,left=thin,right=thin,bottom=med)
    ws1.cell(score_row, c).fill   = PatternFill("solid", fgColor="F0F4FB")
    ws1.cell(score_row, c).alignment = CTR

# conditional formatting on Before Count (col B) and After Count (col D)
ws1.conditional_formatting.add(f"B7:B{score_row-1}",
    CellIsRule(operator="greaterThan", formula=["0"], fill=RED_FILL))
ws1.conditional_formatting.add(f"D7:D{score_row-1}",
    CellIsRule(operator="equal", formula=["0"], fill=GRN_FILL))
ws1.conditional_formatting.add(f"D7:D{score_row-1}",
    CellIsRule(operator="greaterThan", formula=["0"], fill=YEL_FILL))
ws1.conditional_formatting.add(f"G7:G{score_row-1}",
    FormulaRule(formula=['G7="Resolved"'], fill=GRN_FILL, font=Font(bold=True, color="276221")))
ws1.conditional_formatting.add(f"G7:G{score_row-1}",
    FormulaRule(formula=['G7="Flagged"'],  fill=YEL_FILL, font=Font(bold=True, color="7D6608")))

col_widths = [30, 14, 12, 12, 12, 12, 12, 12]
for w, col in zip(col_widths, [get_column_letter(c) for c in range(1, 9)]):
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[6].height = 22
print("   -> Sheet: DQ_Summary")

# ── Sheet 2: Audit_Log ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Audit_Log")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:F1")
t2 = ws2.cell(1, 1, "Validation Pipeline Audit Log")
t2.fill = HDR_FILL; t2.font = Font(bold=True, color="FFFFFF", size=12)
t2.alignment = CTR

audit_cols = list(audit_df.columns)
for c, v in enumerate(audit_cols, 1):
    cell = ws2.cell(3, c, v)
    cell.fill = HDR2_FILL; cell.font = HDR_FONT
    cell.alignment = CTR; cell.border = B()

for r_off, row_data in enumerate(audit_df.itertuples(index=False), 4):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(r_off, c, val)
        cell.fill   = ALT_FILL if r_off % 2 == 0 else WHT_FILL
        cell.font   = SM_FONT
        cell.border = B()
        cell.alignment = LFT if c in (1, 5) else CTR

for w, col in zip([40, 14, 12, 18, 50], "ABCDE"):
    ws2.column_dimensions[col].width = w
print("   -> Sheet: Audit_Log")

# ── Sheet 3: Validation_Rules ─────────────────────────────────────────────────
ws3 = wb.create_sheet("Validation_Rules")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
t3 = ws3.cell(1, 1, "Validation Rule Definitions")
t3.fill = HDR_FILL; t3.font = Font(bold=True, color="FFFFFF", size=12)
t3.alignment = CTR

rule_cols = ["Rule ID","Rule Name","Field(s)","Logic","Action","Priority"]
for c, v in enumerate(rule_cols, 1):
    cell = ws3.cell(3, c, v)
    cell.fill = HDR2_FILL; cell.font = HDR_FONT
    cell.alignment = CTR; cell.border = B()

rules = [
    ("VR-01","Remove Duplicates",
     "Transaction_ID, Customer_ID, Date, Amount, Channel",
     "If all 5 key fields match an earlier record, mark as duplicate",
     "DROP — keep first occurrence","High"),
    ("VR-02","Standardize Date Format",
     "Date",
     "Parse DD/MM/YYYY, MM-DD-YYYY, DD-Mon-YYYY, 'Month DD, YYYY'; convert to YYYY-MM-DD",
     "TRANSFORM — overwrite with ISO format","Medium"),
    ("VR-03","Standardize Channel Labels",
     "Channel",
     "Strip whitespace; normalize case/punctuation; map variants to canonical label",
     "TRANSFORM — map to: UPI/NEFT/IMPS/Card/Cash/Net Banking","Medium"),
    ("VR-04","Flag Missing Amount",
     "Amount",
     "Amount IS NULL or blank",
     "FLAG — add Missing_Amount_Flag=1; retain record for reconciliation","High"),
    ("VR-05","Flag Missing Customer ID",
     "Customer_ID",
     "Customer_ID IS NULL or blank",
     "FLAG — add Missing_CID_Flag=1; retain for manual attribution","High"),
    ("VR-06","Flag Negative Amounts",
     "Amount",
     "Amount < 0 (excluding null)",
     "FLAG — add Negative_Amount_Flag=1; may be valid reversal","Low"),
    ("VR-07","Flag Orphaned Foreign Keys",
     "Customer_ID",
     "Customer_ID not found in customer master (CUST_00001-CUST_00400)",
     "FLAG — add Orphaned_FK_Flag=1; customer must be onboarded before reporting","High"),
]

for r_off, rule_row in enumerate(rules, 4):
    for c, val in enumerate(rule_row, 1):
        cell = ws3.cell(r_off, c, val)
        cell.fill   = ALT_FILL if r_off % 2 == 0 else WHT_FILL
        cell.font   = SM_FONT
        cell.border = B()
        cell.alignment = LFT
        if c == 6:
            if val == "High":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(bold=True, size=9, color="9C0006")
            elif val == "Medium":
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
                cell.font = Font(bold=True, size=9, color="7D6608")

ws3.row_dimensions[1].height = 28
for w, col in zip([10, 28, 40, 55, 50, 12], "ABCDEF"):
    ws3.column_dimensions[col].width = w
for r in range(4, 4+len(rules)):
    ws3.row_dimensions[r].height = 30

wb.save(os.path.join(E_DIR, "dq_validation_report.xlsx"))
print("   -> dq_validation_report.xlsx saved")
print("   -> Sheet: Validation_Rules")

# ─── matplotlib images ────────────────────────────────────────────────────────
print("[5/5] Generating comparison images ...")

# ── image 1: before vs after scorecard ────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(14, 6))
fig.patch.set_facecolor("#F8FAFF")
fig.suptitle("Data Quality Improvement — Before vs After Validation Pipeline",
             fontsize=13, fontweight="bold", color="#1B3A6B", y=1.01)

issue_names = ["Duplicates","Missing\nAmount","Missing\nCustomer ID",
               "Non-Standard\nDates","Negative\nAmounts","Channel\nVariants","Orphaned\nFKs"]
before_vals = [before_profile["Duplicate_Records"],
               before_profile["Missing_Amount"],
               before_profile["Missing_Customer_ID"],
               before_profile["Non_Standard_Dates"],
               before_profile["Negative_Amount"],
               before_profile["Channel_Label_Variants"],
               before_profile["Orphaned_FK"]]
after_vals  = [after_profile["Duplicate_Records"],
               after_profile["Missing_Amount"],
               after_profile["Missing_Customer_ID"],
               after_profile["Non_Standard_Dates"],
               after_profile["Negative_Amount"],
               after_profile["Channel_Label_Variants"],
               after_profile["Orphaned_FK"]]

ax1 = axes[0]
ax1.set_facecolor("#F8FAFF")
x = range(len(issue_names))
w = 0.38
bars_b = ax1.bar([i - w/2 for i in x], before_vals, width=w, label="Before",
                  color="#D9534F", alpha=0.85, zorder=3)
bars_a = ax1.bar([i + w/2 for i in x], after_vals,  width=w, label="After",
                  color="#5CB85C", alpha=0.85, zorder=3)
ax1.set_xticks(list(x))
ax1.set_xticklabels(issue_names, fontsize=8)
ax1.set_title("Issue Count: Before vs After", fontsize=11, color="#1B3A6B", fontweight="bold")
ax1.set_ylabel("Record Count", fontsize=9)
ax1.legend(fontsize=9)
ax1.yaxis.grid(True, alpha=0.3, linestyle="--")
for bar, val in zip(bars_b, before_vals):
    ax1.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
             str(val), ha="center", fontsize=7.5, color="#9C0006", fontweight="bold")
for bar, val in zip(bars_a, after_vals):
    ax1.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
             str(val), ha="center", fontsize=7.5, color="#276221", fontweight="bold")

# DQ score gauge (right panel)
ax2 = axes[1]
ax2.set_facecolor("#F8FAFF")
ax2.set_xlim(0, 1); ax2.set_ylim(0, 1)
ax2.axis("off")

scores = [before_profile["DQ_Score_%"], after_profile["DQ_Score_%"]]
score_colors = ["#D9534F", "#5CB85C"]
labels = [f"Before\n{before_profile['DQ_Score_%']}%", f"After\n{after_profile['DQ_Score_%']}%"]
for i, (score, color, label) in enumerate(zip(scores, score_colors, labels)):
    x_pos = 0.20 + i * 0.50
    circle = plt.Circle((x_pos, 0.52), 0.22, color=color, zorder=3, alpha=0.9)
    ax2.add_patch(circle)
    ax2.text(x_pos, 0.52, f"{score}%", ha="center", va="center",
             fontsize=22, fontweight="bold", color="white")
    ax2.text(x_pos, 0.24, label, ha="center", va="top", fontsize=11,
             color="#1B3A6B", fontweight="bold")

ax2.annotate("", xy=(0.58, 0.52), xytext=(0.42, 0.52),
             arrowprops=dict(arrowstyle="->", color="#1B3A6B", lw=2.5))
delta = after_profile["DQ_Score_%"] - before_profile["DQ_Score_%"]
ax2.text(0.50, 0.60, f"+{delta:.1f}pp", ha="center", va="bottom",
         fontsize=13, color="#27AE60", fontweight="bold")
ax2.set_title("Overall DQ Score Improvement", fontsize=11, color="#1B3A6B",
              fontweight="bold", pad=10)

plt.tight_layout()
plt.savefig(os.path.join(I_DIR, "01-dq-scorecard-comparison.png"), dpi=150, bbox_inches="tight")
plt.close()
print("   -> 01-dq-scorecard-comparison.png")

# ── image 2: issue breakdown + resolution breakdown ────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5))
fig.patch.set_facecolor("#F8FAFF")
fig.suptitle("Issue Type Analysis — Reporting Data Quality Pipeline",
             fontsize=12, fontweight="bold", color="#1B3A6B")

ax1 = axes[0]
ax1.set_facecolor("#F8FAFF")
y_labels = issue_names
severity_colors = ["#D9534F","#E8A838","#D9534F","#E8A838","#5BC0DE","#E8A838","#D9534F"]
hb = ax1.barh(y_labels, before_vals, color=severity_colors, height=0.6, zorder=3)
ax1.xaxis.grid(True, alpha=0.3, linestyle="--")
ax1.set_title("Issue Count by Type (Before Cleaning)", fontsize=10,
              color="#1B3A6B", fontweight="bold")
ax1.set_xlabel("Records Affected", fontsize=9)
for bar, val in zip(hb, before_vals):
    ax1.text(bar.get_width()+1, bar.get_y()+bar.get_height()/2,
             str(val), va="center", fontsize=9, fontweight="bold")

# resolution status (right)
ax2 = axes[1]
statuses = ["Resolved\n(Removed)", "Standardized\n(Transformed)", "Flagged\n(Retained)"]
# resolved = dupes removed
# standardized = dates + channels fixed
# flagged = missing amt, missing CID, negatives, orphan FKs
resolved_count    = N_DUPES
standardized_count = N_DATE_FMT + N_CAT_VAR
flagged_count      = N_MISSING_AMT + N_MISSING_CID + N_NEG_AMT + N_ORPHAN
status_vals = [resolved_count, standardized_count, flagged_count]
status_colors = ["#5CB85C", "#2E5E9E", "#E8A838"]
wedges, texts, autotexts = ax2.pie(
    status_vals, labels=statuses, autopct="%1.1f%%",
    colors=status_colors, startangle=90,
    wedgeprops={"edgecolor":"white","linewidth":2},
    textprops={"fontsize":9})
for at in autotexts:
    at.set_fontsize(9)
    at.set_fontweight("bold")
    at.set_color("white")
ax2.set_title("Issue Resolution Method", fontsize=10, color="#1B3A6B", fontweight="bold")
plt.tight_layout()
plt.savefig(os.path.join(I_DIR, "02-issue-breakdown.png"), dpi=150, bbox_inches="tight")
plt.close()
print("   -> 02-issue-breakdown.png")

# ── image 3: formula bar simulation ───────────────────────────────────────────
fig, axes = plt.subplots(3, 1, figsize=(14, 7))
fig.patch.set_facecolor("#252526")
fig.suptitle("Python Validation Pipeline — Core Logic", fontsize=11,
             color="#CCCCCC", y=1.01)

snippets = [
    ("Rule VR-01: Deduplication",
     "clean_df = messy_df.drop_duplicates(\n"
     "    subset=['Transaction_ID','Customer_ID','Date','Amount','Channel'],\n"
     "    keep='first'  # retain earliest occurrence\n"
     ")"),
    ("Rule VR-02: Date Standardization (multi-format parser)",
     "def standardize_date(v):\n"
     "    formats = ['%Y-%m-%d','%d/%m/%Y','%m-%d-%Y','%d-%b-%Y','%B %d, %Y']\n"
     "    for fmt in formats:\n"
     "        try: return datetime.strptime(str(v), fmt).strftime('%Y-%m-%d')\n"
     "        except: pass\n"
     "clean_df['Date'] = clean_df['Date'].apply(standardize_date)"),
    ("Rule VR-07: Orphaned FK Detection",
     "CUST_MASTER = set(pd.read_csv('customer_master.csv')['Customer_ID'])\n"
     "clean_df['Orphaned_FK_Flag'] = (\n"
     "    ~clean_df['Customer_ID'].isin(CUST_MASTER) &\n"
     "    clean_df['Customer_ID'].notna()\n"
     ").astype(int)"),
]

for ax, (title, code) in zip(axes, snippets):
    ax.set_facecolor("#1E1E1E")
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.axis("off")
    ax.add_patch(mpatches.FancyBboxPatch((0, 0.75), 1.0, 0.24,
        boxstyle="square,pad=0", facecolor="#2D2D30", edgecolor="none"))
    ax.text(0.01, 0.87, title, color="#DCDCAA", fontsize=8.5,
            fontfamily="monospace", fontweight="bold", va="center")
    ax.text(0.02, 0.66, code, color="#9CDCFE", fontsize=8.5,
            fontfamily="monospace", va="top", multialignment="left")

plt.tight_layout(pad=0.3)
plt.savefig(os.path.join(I_DIR, "03-pipeline-code-view.png"), dpi=150, bbox_inches="tight",
            facecolor="#252526")
plt.close()
print("   -> 03-pipeline-code-view.png")

print("\nBuild complete.")
print(f"  Raw rows     : {len(messy_df)}")
print(f"  Cleaned rows : {len(clean_df)}")
print(f"  DQ Before    : {before_profile['DQ_Score_%']}%")
print(f"  DQ After     : {after_profile['DQ_Score_%']}%")
print(f"  Improvement  : +{after_profile['DQ_Score_%'] - before_profile['DQ_Score_%']:.1f}pp")
