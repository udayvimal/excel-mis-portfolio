#!/usr/bin/env python3
"""
Advisor Incentive Performance Dashboard Builder
Generates 6-month simulated performance data, Excel data model, DAX measures file, and chart images.
"""
import os, sys, warnings
warnings.filterwarnings("ignore")
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

np.random.seed(42)

# ─────────────────────────────────────────────
# 1.  CONFIGURATION & HIERARCHY
# ─────────────────────────────────────────────
MONTHS = ["Jan-26","Feb-26","Mar-26","Apr-26","May-26","Jun-26"]
MONTH_DATES = ["2026-01-01","2026-02-01","2026-03-01","2026-04-01","2026-05-01","2026-06-01"]

AM_NAMES  = {1:"Rajesh Kumar",   2:"Sonia Mehta",   3:"Arjun Sharma"}
TL_NAMES  = {
    1:"Kavita Singh",  2:"Deepak Rao",   3:"Priya Nair",
    4:"Suresh Pillai", 5:"Meera Joshi",  6:"Rohit Gupta",  # TL5=Meera Joshi (underperforming team)
    7:"Anita Das",     8:"Vikram Bhat",  9:"Pooja Tiwari", 10:"Anil Verma"
}
TL_AM = {1:1,2:1,3:1, 4:2,5:2,6:2, 7:3,8:3,9:3,10:3}

ADV_FIRST = ["Amit","Priyanka","Rahul","Sneha","Vijay","Ritu","Kiran","Manish",
             "Divya","Saurabh","Ananya","Nikhil","Pooja","Ravi","Kavitha",
             "Arjun","Swati","Tarun","Nidhi","Harish","Sunita","Akash","Lakshmi",
             "Deepa","Sanjay","Meghna","Pradeep","Asha","Girish","Vinita",
             "Suresh","Padma","Ramesh","Geetha","Dinesh","Sarita","Mahesh",
             "Usha","Ganesh","Pushpa","Srinivas","Nalini","Venkat","Rekha",
             "Bharat","Savita","Naresh","Kamla","Yogesh","Sheela","Vivek",
             "Saroj","Umesh","Heena","Mukesh","Vandana","Rajiv","Neeta",
             "Lalit","Sunanda"]
ADV_LAST = ["Sharma","Gupta","Singh","Kumar","Patel","Joshi","Nair","Rao","Das","Mehta",
            "Reddy","Pillai","Bhat","Tiwari","Verma","Agarwal","Mishra","Chaudhary",
            "Dubey","Iyer","Murthy","Naidu","Pandey","Shukla","Srivastava",
            "Thakur","Trivedi","Malhotra","Kapoor","Saxena",
            "Rastogi","Bhatt","Chopra","Gill","Bansal","Goyal","Shah","Doshi",
            "Dalal","Jain","Kothari","Modi","Parikh","Sanghvi","Vakil",
            "Bhavsar","Desai","Parekh","Phadke","Sathe","Kadam","Patil",
            "Kulkarni","Jadhav","More","Shinde","Salvi","Gaikwad","Mane","Sawant"]

# Build advisor dimension (60 advisors, 6 per TL, 10 TLs)
adv_rows = []
for tid in range(1,11):
    for j in range(6):
        idx = (tid-1)*6 + j
        aid = f"ADV_{idx+1:03d}"
        name = f"{ADV_FIRST[idx]} {ADV_LAST[idx]}"
        adv_rows.append(dict(
            Advisor_ID=aid, Advisor_Name=name,
            TL_ID=f"TL_{tid:02d}", TL_Name=TL_NAMES[tid],
            AM_ID=f"AM_{TL_AM[tid]:02d}", AM_Name=AM_NAMES[TL_AM[tid]]
        ))
advisor_dim = pd.DataFrame(adv_rows)

tl_dim = pd.DataFrame([
    dict(TL_ID=f"TL_{t:02d}", TL_Name=TL_NAMES[t],
         AM_ID=f"AM_{TL_AM[t]:02d}", AM_Name=AM_NAMES[TL_AM[t]])
    for t in range(1,11)
])
am_dim = pd.DataFrame([
    dict(AM_ID=f"AM_{a:02d}", AM_Name=AM_NAMES[a], Region=["North","West","South"][a-1])
    for a in range(1,4)
])
date_dim = pd.DataFrame([
    dict(Month_Year=m, Date=d, Month_Num=i+1,
         Quarter=f"Q{(i//3)+1}-26", FY="FY2026")
    for i,(m,d) in enumerate(zip(MONTHS,MONTH_DATES))
])

# ─────────────────────────────────────────────
# 2.  PERFORMANCE DATA GENERATION
# ─────────────────────────────────────────────
# Program targets (per advisor per month)
TARGETS = dict(Retention=50, Migration=20, Rider=30, CI=25, Referral=15)
# Normal achievement % distributions (mean, std)
NORMAL = dict(Retention=(76,10), Migration=(68,12), Rider=(72,11), CI=(67,12), Referral=(79,9))
# TL_05 team = advisors ADV_025 to ADV_030, underperforming on Rider & CI in months Feb-Apr
UNDERPERFORM_TL = "TL_05"
UNDERPERFORM_MONTHS = ["Feb-26","Mar-26","Apr-26"]
UNDERPERFORM_PROGS = dict(Rider=(38,9), CI=(32,8))

perf_rows = []
for _, adv in advisor_dim.iterrows():
    aid = adv["Advisor_ID"]
    tl  = adv["TL_ID"]
    for mi, month in enumerate(MONTHS):
        row = dict(Advisor_ID=aid, Month_Year=month)
        scores = {}
        for prog, (tgt,) in [(p,( TARGETS[p],)) for p in ["Retention","Migration","Rider","CI","Referral"]]:
            # Achievement %
            if tl == UNDERPERFORM_TL and month in UNDERPERFORM_MONTHS and prog in ("Rider","CI"):
                mu, sd = UNDERPERFORM_PROGS[prog]
            else:
                mu, sd = NORMAL[prog]
            ach_pct = np.clip(np.random.normal(mu, sd), 20, 130) / 100
            achieved = max(0, round(TARGETS[prog] * ach_pct))
            target   = TARGETS[prog]
            row[f"{prog}_Target"]   = target
            row[f"{prog}_Achieved"] = achieved
            row[f"{prog}_Pct"]      = round(min(achieved/target, 1.30) * 100, 1)
            scores[prog] = min(achieved/target, 1.30)

        # Consolidated Score (weighted)
        cs = (scores["Retention"] * 0.30 + scores["Migration"] * 0.15 +
              scores["Rider"] * 0.20 + scores["CI"] * 0.20 + scores["Referral"] * 0.15)
        cs_pct = round(cs * 100, 1)
        row["Consolidated_Score"] = cs_pct

        # Incentive Slab
        if cs_pct < 60:   slab, amt = "No Incentive", 0
        elif cs_pct < 70: slab, amt = "Bronze",        2000
        elif cs_pct < 80: slab, amt = "Silver",        5000
        elif cs_pct < 90: slab, amt = "Gold",          8000
        else:             slab, amt = "Platinum",      12000
        row["Incentive_Slab"]   = slab
        row["Incentive_Amount"] = amt

        perf_rows.append(row)

perf_df = pd.DataFrame(perf_rows)

# Merge hierarchy for convenience
perf_full = perf_df.merge(advisor_dim[["Advisor_ID","TL_ID","TL_Name","AM_ID","AM_Name"]], on="Advisor_ID")

# ─────────────────────────────────────────────
# 3.  SAVE CSVs
# ─────────────────────────────────────────────
os.makedirs("data",  exist_ok=True)
os.makedirs("excel", exist_ok=True)
os.makedirs("dax",   exist_ok=True)
os.makedirs("images",exist_ok=True)
os.makedirs("report",exist_ok=True)

advisor_dim.to_csv("data/advisor_dim.csv",     index=False)
tl_dim.to_csv     ("data/tl_dim.csv",          index=False)
am_dim.to_csv     ("data/am_dim.csv",          index=False)
date_dim.to_csv   ("data/date_dim.csv",        index=False)
perf_df.to_csv    ("data/performance_fact.csv",index=False)
print(f"CSVs saved (performance rows: {len(perf_df)})")

# ─────────────────────────────────────────────
# 4.  DAX MEASURES FILE
# ─────────────────────────────────────────────
dax_text = r"""
// ════════════════════════════════════════════════════════════════════════
// ADVISOR INCENTIVE PERFORMANCE — DAX MEASURES
// Data Model: Star schema — performance_fact linked to advisor_dim,
//             advisor_dim linked to tl_dim, tl_dim linked to am_dim,
//             performance_fact linked to date_dim via Month_Year
// ════════════════════════════════════════════════════════════════════════

// ─────────────────────────────────────────────────────────────────────
// MEASURE 1: Consolidated Incentive Score (%)
// Weighted achievement across all 5 programs
// Retention 30% | Migration 15% | Rider 20% | CI 20% | Referrals 15%
// ─────────────────────────────────────────────────────────────────────
Consolidated Incentive Score (%) =
VAR RetPct =
    DIVIDE(
        SUM(performance_fact[Retention_Achieved]),
        SUM(performance_fact[Retention_Target]),
        0
    )
VAR MigPct =
    DIVIDE(
        SUM(performance_fact[Migration_Achieved]),
        SUM(performance_fact[Migration_Target]),
        0
    )
VAR RidPct =
    DIVIDE(
        SUM(performance_fact[Rider_Achieved]),
        SUM(performance_fact[Rider_Target]),
        0
    )
VAR CIPct =
    DIVIDE(
        SUM(performance_fact[CI_Achieved]),
        SUM(performance_fact[CI_Target]),
        0
    )
VAR RefPct =
    DIVIDE(
        SUM(performance_fact[Referral_Achieved]),
        SUM(performance_fact[Referral_Target]),
        0
    )
RETURN
    ROUND(
        (RetPct * 0.30) +
        (MigPct * 0.15) +
        (RidPct * 0.20) +
        (CIPct  * 0.20) +
        (RefPct * 0.15),
        4
    ) * 100


// ─────────────────────────────────────────────────────────────────────
// MEASURE 2: Retention Achievement %
// ─────────────────────────────────────────────────────────────────────
Retention Achievement % =
DIVIDE(
    SUM(performance_fact[Retention_Achieved]),
    SUM(performance_fact[Retention_Target]),
    0
) * 100


// ─────────────────────────────────────────────────────────────────────
// MEASURE 3: Rider Attachment Achievement %
// ─────────────────────────────────────────────────────────────────────
Rider Attachment % =
DIVIDE(
    SUM(performance_fact[Rider_Achieved]),
    SUM(performance_fact[Rider_Target]),
    0
) * 100


// ─────────────────────────────────────────────────────────────────────
// MEASURE 4: CI Attachment Achievement %
// ─────────────────────────────────────────────────────────────────────
CI Attachment % =
DIVIDE(
    SUM(performance_fact[CI_Achieved]),
    SUM(performance_fact[CI_Target]),
    0
) * 100


// ─────────────────────────────────────────────────────────────────────
// MEASURE 5: Month-over-Month Change in Consolidated Score
// Uses DATEADD time intelligence — requires date_dim date table
// ─────────────────────────────────────────────────────────────────────
Consolidated Score MoM Change =
VAR CurrentScore =
    [Consolidated Incentive Score (%)]
VAR PreviousMonthScore =
    CALCULATE(
        [Consolidated Incentive Score (%)],
        DATEADD(date_dim[Date], -1, MONTH)
    )
RETURN
    IF(
        NOT ISBLANK(PreviousMonthScore) && PreviousMonthScore <> 0,
        DIVIDE(CurrentScore - PreviousMonthScore, PreviousMonthScore, BLANK()),
        BLANK()
    )


// ─────────────────────────────────────────────────────────────────────
// MEASURE 6: TL-Level Average Consolidated Score
// Rolls up to Team Lead level via advisor_dim relationship
// ─────────────────────────────────────────────────────────────────────
TL Avg Consolidated Score =
AVERAGEX(
    SUMMARIZE(
        advisor_dim,
        advisor_dim[TL_ID],
        advisor_dim[TL_Name]
    ),
    CALCULATE( [Consolidated Incentive Score (%)] )
)


// ─────────────────────────────────────────────────────────────────────
// MEASURE 7: AM-Level Average Consolidated Score
// Rolls up through tl_dim to area manager
// ─────────────────────────────────────────────────────────────────────
AM Avg Consolidated Score =
AVERAGEX(
    SUMMARIZE(
        am_dim,
        am_dim[AM_ID],
        am_dim[AM_Name]
    ),
    CALCULATE( [Consolidated Incentive Score (%)] )
)


// ─────────────────────────────────────────────────────────────────────
// MEASURE 8: Total Incentive Amount Paid
// ─────────────────────────────────────────────────────────────────────
Total Incentive Paid =
SUM(performance_fact[Incentive_Amount])


// ─────────────────────────────────────────────────────────────────────
// MEASURE 9: % Advisors at Platinum
// ─────────────────────────────────────────────────────────────────────
% Advisors at Platinum =
DIVIDE(
    CALCULATE(
        COUNTROWS(performance_fact),
        performance_fact[Incentive_Slab] = "Platinum"
    ),
    COUNTROWS(performance_fact),
    0
) * 100


// ─────────────────────────────────────────────────────────────────────
// MEASURE 10: Incentive Variance vs Prior Month (INR)
// ─────────────────────────────────────────────────────────────────────
Incentive Amount MoM Variance =
VAR CurrAmt =
    SUM(performance_fact[Incentive_Amount])
VAR PrevAmt =
    CALCULATE(
        SUM(performance_fact[Incentive_Amount]),
        DATEADD(date_dim[Date], -1, MONTH)
    )
RETURN
    IF(NOT ISBLANK(PrevAmt), CurrAmt - PrevAmt, BLANK())


// ─────────────────────────────────────────────────────────────────────
// RELATIONSHIPS (define in Power BI Model view):
//   performance_fact[Advisor_ID]  → advisor_dim[Advisor_ID]   (Many-to-One)
//   advisor_dim[TL_ID]            → tl_dim[TL_ID]             (Many-to-One)
//   tl_dim[AM_ID]                 → am_dim[AM_ID]             (Many-to-One)
//   performance_fact[Month_Year]  → date_dim[Month_Year]      (Many-to-One)
//   Mark date_dim as "Date Table" using the [Date] column
// ─────────────────────────────────────────────────────────────────────
"""
with open("dax/measures.dax","w", encoding="utf-8") as f:
    f.write(dax_text)
print("DAX file saved")

# ─────────────────────────────────────────────
# 5.  EXCEL DATA MODEL WORKBOOK
# ─────────────────────────────────────────────
def mk_fill(h): return PatternFill("solid", fgColor=h)
def mk_font(h, bold=False, sz=11): return Font(color=h, bold=bold, size=sz)
thin  = Side(style="thin")
BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR   = Alignment(horizontal="center", vertical="center")
WRAP  = Alignment(horizontal="center", vertical="center", wrap_text=True)

F_NAVY  = mk_fill("0F2D5A"); F_BLUE = mk_fill("1565C0"); F_TEAL = mk_fill("00695C")
F_ORG   = mk_fill("E65100"); F_LBLUE= mk_fill("BBDEFB"); F_LGRAY= mk_fill("F5F5F5")
F_LGRN  = mk_fill("C8E6C9"); F_RED  = mk_fill("FFCDD2")
WH = mk_font("FFFFFF", bold=True); BL = mk_font("000000")

def hdr(ws, row, fill=F_NAVY, fnt=WH):
    for c in ws[row]:
        c.fill=fill; c.font=fnt; c.alignment=CTR; c.border=BDR

def widths(ws, ww):
    for i,w in enumerate(ww,1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# Sheet: Advisor_Dim
ws_adv = wb.active; ws_adv.title = "Advisor_Dim"
ws_adv.sheet_properties.tabColor = "1565C0"
ws_adv.append(list(advisor_dim.columns)); hdr(ws_adv,1,F_BLUE)
for _, r in advisor_dim.iterrows():
    ws_adv.append(list(r))
    for c in range(1,7):
        ws_adv.cell(ws_adv.max_row,c).border=BDR
        ws_adv.cell(ws_adv.max_row,c).alignment=CTR
        ws_adv.cell(ws_adv.max_row,c).fill = F_LBLUE if ws_adv.max_row%2==0 else F_LGRAY
widths(ws_adv,[12,22,8,16,8,16])
ws_adv.freeze_panes="A2"

# Sheet: TL_Dim
ws_tl = wb.create_sheet("TL_Dim"); ws_tl.sheet_properties.tabColor="00695C"
ws_tl.append(list(tl_dim.columns)); hdr(ws_tl,1,F_TEAL)
for _, r in tl_dim.iterrows():
    ws_tl.append(list(r))
    for c in range(1,5):
        ws_tl.cell(ws_tl.max_row,c).border=BDR; ws_tl.cell(ws_tl.max_row,c).alignment=CTR
        ws_tl.cell(ws_tl.max_row,c).fill = F_LGRN if ws_tl.max_row%2==0 else F_LGRAY
widths(ws_tl,[10,18,8,16]); ws_tl.freeze_panes="A2"

# Sheet: AM_Dim
ws_am = wb.create_sheet("AM_Dim"); ws_am.sheet_properties.tabColor="E65100"
ws_am.append(list(am_dim.columns)); hdr(ws_am,1,F_ORG)
for _, r in am_dim.iterrows():
    ws_am.append(list(r))
    for c in range(1,4):
        ws_am.cell(ws_am.max_row,c).border=BDR; ws_am.cell(ws_am.max_row,c).alignment=CTR
widths(ws_am,[8,18,10]); ws_am.freeze_panes="A2"

# Sheet: Date_Dim
ws_dt = wb.create_sheet("Date_Dim"); ws_dt.sheet_properties.tabColor="4A148C"
ws_dt.append(list(date_dim.columns))
hdr(ws_dt,1,mk_fill("4A148C"))
for _, r in date_dim.iterrows():
    ws_dt.append(list(r))
    for c in range(1,6):
        ws_dt.cell(ws_dt.max_row,c).border=BDR; ws_dt.cell(ws_dt.max_row,c).alignment=CTR
widths(ws_dt,[10,12,10,10,8])

# Sheet: Performance_Fact
ws_pf = wb.create_sheet("Performance_Fact"); ws_pf.sheet_properties.tabColor="C62828"
ws_pf.append(list(perf_df.columns)); hdr(ws_pf,1,mk_fill("C62828"))
for _, r in perf_df.iterrows():
    ws_pf.append(list(r))
    row_n = ws_pf.max_row
    for c in range(1, len(perf_df.columns)+1):
        ws_pf.cell(row_n,c).border=BDR; ws_pf.cell(row_n,c).alignment=CTR
        ws_pf.cell(row_n,c).fill = F_LBLUE if row_n%2==0 else F_LGRAY
ws_pf.freeze_panes="A2"
widths(ws_pf,[12,8,16,17,13,14,15,10,13,9,10,11,9,18,17,17])

# Sheet: Data_Model_Schema
ws_schema = wb.create_sheet("Data_Model_Schema")
ws_schema.sheet_properties.tabColor="37474F"
schema_info = [
    ["TABLE", "KEY COLUMN", "RELATIONSHIP", "RELATED TABLE", "TYPE"],
    ["performance_fact","Advisor_ID","→ (Many-to-One)","Advisor_Dim[Advisor_ID]","Fact-to-Dim"],
    ["performance_fact","Month_Year","→ (Many-to-One)","Date_Dim[Month_Year]","Fact-to-Dim"],
    ["Advisor_Dim","TL_ID","→ (Many-to-One)","TL_Dim[TL_ID]","Dim-to-Dim"],
    ["TL_Dim","AM_ID","→ (Many-to-One)","AM_Dim[AM_ID]","Dim-to-Dim"],
    ["","","","",""],
    ["STAR SCHEMA","","","",""],
    ["Fact Table: performance_fact (360 rows)","","","",""],
    ["Dimensions: Advisor_Dim, TL_Dim, AM_Dim, Date_Dim","","","",""],
    ["Hierarchy: AM → TL → Advisor (3 levels)","","","",""],
]
ws_schema.merge_cells("A1:E1")
ws_schema["A1"].value="DATA MODEL SCHEMA — Advisor Incentive Performance"
ws_schema["A1"].font=Font(bold=True,size=14,color="FFFFFF")
ws_schema["A1"].fill=mk_fill("37474F"); ws_schema["A1"].alignment=CTR
for i,row in enumerate(schema_info,2):
    ws_schema.append(row)
    if i==3:
        hdr(ws_schema,i,mk_fill("546E7A"))
widths(ws_schema,[22,18,20,28,16])

wb.save("excel/advisor_incentive_data_model.xlsx")
print("Excel data model saved")

# ─────────────────────────────────────────────
# 6.  PRE-COMPUTE AGGREGATES FOR IMAGES
# ─────────────────────────────────────────────
# TL-level monthly summary
tl_monthly = (perf_full.groupby(["TL_ID","TL_Name","Month_Year"])
              .agg(Avg_CS=("Consolidated_Score","mean"),
                   Avg_Rider=("Rider_Pct","mean"),
                   Avg_CI=("CI_Pct","mean"),
                   Total_Incentive=("Incentive_Amount","sum"))
              .reset_index())

# Overall program monthly trend
prog_monthly = {}
for p in ["Retention","Migration","Rider","CI","Referral"]:
    prog_monthly[p] = (perf_full.groupby("Month_Year")[f"{p}_Pct"].mean().values)

# TL consolidated score by month
tl_cs = tl_monthly.pivot(index="Month_Year", columns="TL_Name", values="Avg_CS")
month_order = MONTHS; tl_cs = tl_cs.reindex(month_order)

# Incentive distribution
incentive_dist = (perf_full.groupby(["Incentive_Slab"])["Advisor_ID"]
                  .count().reset_index()
                  .rename(columns={"Advisor_ID":"Count"}))
slab_order = ["No Incentive","Bronze","Silver","Gold","Platinum"]
incentive_dist["Incentive_Slab"] = pd.Categorical(incentive_dist.Incentive_Slab, categories=slab_order, ordered=True)
incentive_dist = incentive_dist.sort_values("Incentive_Slab")

# AM-level summary
am_summary = (perf_full.groupby(["AM_ID","AM_Name"])
              .agg(Avg_CS=("Consolidated_Score","mean"),
                   Total_Incentive=("Incentive_Amount","sum"),
                   Advisor_Count=("Advisor_ID","nunique"))
              .reset_index())

# TL_05 drill-down
tl5_data = perf_full[perf_full.TL_ID=="TL_05"].copy()
tl5_monthly = tl5_data.groupby("Month_Year").agg(
    Rider=("Rider_Pct","mean"), CI=("CI_Pct","mean"),
    CS=("Consolidated_Score","mean")).reindex(MONTHS)
all_others = perf_full[perf_full.TL_ID!="TL_05"].groupby("Month_Year").agg(
    Rider=("Rider_Pct","mean"), CI=("CI_Pct","mean"),
    CS=("Consolidated_Score","mean")).reindex(MONTHS)

overall_stats = {
    "total_advisors": 60,
    "total_incentive": int(perf_full.Incentive_Amount.sum()),
    "avg_cs": round(perf_full.Consolidated_Score.mean(), 1),
    "pct_platinum": round((perf_full.Incentive_Slab=="Platinum").mean()*100, 1),
    "pct_no_incentive": round((perf_full.Incentive_Slab=="No Incentive").mean()*100, 1),
    "tl5_avg_cs": round(tl5_data.Consolidated_Score.mean(), 1),
    "others_avg_cs": round(perf_full[perf_full.TL_ID!="TL_05"].Consolidated_Score.mean(), 1),
    "tl5_rider_feb_apr": round(tl5_data[tl5_data.Month_Year.isin(UNDERPERFORM_MONTHS)].Rider_Pct.mean(), 1),
    "all_rider_feb_apr": round(perf_full[perf_full.Month_Year.isin(UNDERPERFORM_MONTHS)].Rider_Pct.mean(), 1),
    "tl5_ci_feb_apr": round(tl5_data[tl5_data.Month_Year.isin(UNDERPERFORM_MONTHS)].CI_Pct.mean(), 1),
    "all_ci_feb_apr": round(perf_full[perf_full.Month_Year.isin(UNDERPERFORM_MONTHS)].CI_Pct.mean(), 1),
}
print("Key stats:", overall_stats)

# ─────────────────────────────────────────────
# 7.  MATPLOTLIB IMAGES
# ─────────────────────────────────────────────
PBI_BG  = "#F3F2F1"
C_NAVY  = "#0F2D5A"; C_BLUE = "#1565C0"; C_TEAL = "#00695C"
C_ORG   = "#E65100"; C_RED  = "#B71C1C"; C_GRN  = "#1B5E20"
C_LBLUE = "#BBDEFB"; C_YLW  = "#FFF9C4"
PROG_COLORS = {"Retention":C_BLUE,"Migration":C_TEAL,"Rider":C_ORG,
               "CI":C_RED,"Referral":C_GRN}

plt.rcParams.update({
    "font.family":"DejaVu Sans", "axes.facecolor":PBI_BG,
    "figure.facecolor":"#FFFFFF", "axes.grid":True,
    "grid.color":"#E0E0E0","grid.linewidth":0.7,
    "axes.titlesize":13,"axes.titlecolor":C_NAVY,"axes.titleweight":"bold",
    "xtick.labelsize":9,"ytick.labelsize":9,
})

# ── Image 01: Dashboard Overview ──────────────────────────────────────
fig = plt.figure(figsize=(20, 14))
fig.patch.set_facecolor(PBI_BG)
gs = gridspec.GridSpec(4, 4, figure=fig, hspace=0.5, wspace=0.4,
                       top=0.93, bottom=0.05, left=0.05, right=0.97)

# Header
hdr_ax = fig.add_subplot(gs[0,:])
hdr_ax.set_facecolor(C_NAVY); hdr_ax.axis("off")
hdr_ax.text(0.5, 0.65, "Advisor Incentive Performance Dashboard",
            ha="center", va="center", transform=hdr_ax.transAxes,
            fontsize=22, fontweight="bold", color="white")
hdr_ax.text(0.5, 0.22, "6-Month Performance Overview  |  Jan–Jun 2026  |  60 Advisors  |  10 Teams",
            ha="center", va="center", transform=hdr_ax.transAxes,
            fontsize=11, color="#BBDEFB", style="italic")

# KPI cards (row 1)
kpi_data = [
    ("Total Advisors",    "60",                 C_BLUE, "#FFFFFF"),
    ("Total Incentive",   f'Rs.{overall_stats["total_incentive"]/1e6:.2f}M', C_TEAL, "#FFFFFF"),
    ("Avg Consolidated",  f'{overall_stats["avg_cs"]}%',                C_ORG, "#FFFFFF"),
    ("% Platinum Slab",   f'{overall_stats["pct_platinum"]}%',          C_GRN, "#FFFFFF"),
]
for ki,(lbl,val,clr,tc) in enumerate(kpi_data):
    kax = fig.add_subplot(gs[1,ki])
    kax.set_facecolor(clr); kax.axis("off")
    kax.text(0.5, 0.7, val, ha="center", va="center", transform=kax.transAxes,
             fontsize=24, fontweight="bold", color=tc)
    kax.text(0.5, 0.22, lbl, ha="center", va="center", transform=kax.transAxes,
             fontsize=10, color="#EEEEEE")
    kax.set_title("", pad=0)

# Program achievement trend (row 2, cols 0-2)
ax_trend = fig.add_subplot(gs[2,:3])
ax_trend.set_facecolor("#FFFFFF")
for p in ["Retention","Migration","Rider","CI","Referral"]:
    ax_trend.plot(MONTHS, prog_monthly[p], marker="o", linewidth=2.2, markersize=5,
                  label=p, color=PROG_COLORS[p])
ax_trend.axhline(70, color="#999999", linestyle="--", linewidth=0.8, alpha=0.6)
ax_trend.set_title("Program Achievement % — Monthly Trend"); ax_trend.set_ylabel("Achievement %")
ax_trend.set_ylim(20,110); ax_trend.legend(fontsize=8, ncol=5, loc="upper right")
# Shade underperform months
for mi,m in enumerate(MONTHS):
    if m in UNDERPERFORM_MONTHS:
        ax_trend.axvspan(mi-0.4, mi+0.4, alpha=0.08, color=C_RED)

# Incentive distribution (row 2, col 3)
ax_inc = fig.add_subplot(gs[2,3])
ax_inc.set_facecolor("#FFFFFF")
slab_colors = {"No Incentive":C_RED,"Bronze":"#CD7F32","Silver":"#9E9E9E",
               "Gold":"#FFC107","Platinum":"#5C6BC0"}
colors_list = [slab_colors[s] for s in incentive_dist.Incentive_Slab]
wedges,texts,autotexts = ax_inc.pie(
    incentive_dist.Count, labels=None,
    colors=colors_list, autopct="%1.0f%%",
    pctdistance=0.75, startangle=90,
    wedgeprops=dict(edgecolor="white",linewidth=1.5)
)
for at in autotexts: at.set_fontsize(8)
ax_inc.set_title("Incentive Slab Distribution")
ax_inc.legend(incentive_dist.Incentive_Slab.tolist(), fontsize=7,
              loc="lower center", ncol=2, bbox_to_anchor=(0.5,-0.18))

# TL Heatmap (row 3)
ax_tl = fig.add_subplot(gs[3,:])
ax_tl.set_facecolor("#FFFFFF")
tl_cs_vals = tl_cs.values.T
im = ax_tl.imshow(tl_cs_vals, aspect="auto", cmap="RdYlGn",
                   vmin=55, vmax=90, interpolation="nearest")
ax_tl.set_xticks(range(len(MONTHS))); ax_tl.set_xticklabels(MONTHS, fontsize=9)
ax_tl.set_yticks(range(len(tl_cs.columns))); ax_tl.set_yticklabels(tl_cs.columns, fontsize=8)
ax_tl.set_title("Team Lead Consolidated Score Heatmap (Jan–Jun 2026)")
for yi in range(len(tl_cs.columns)):
    for xi in range(len(MONTHS)):
        v = tl_cs_vals[yi,xi]
        ax_tl.text(xi, yi, f"{v:.0f}%", ha="center", va="center",
                   fontsize=8, color="black" if v>65 else "white", fontweight="bold")
plt.colorbar(im, ax=ax_tl, fraction=0.02, pad=0.01, label="Consolidated Score %")

fig.suptitle("", y=0.97)
plt.savefig("images/01-dashboard-overview.png", dpi=150, bbox_inches="tight",
            facecolor=PBI_BG)
plt.close(); print("Image 01 saved")

# ── Image 02: Program Achievement Trend ───────────────────────────────
fig, axes = plt.subplots(2,3, figsize=(18,10))
fig.suptitle("Program-wise Achievement % — Monthly Trend (Jan–Jun 2026)",
             fontsize=16, color=C_NAVY, fontweight="bold")
programs_info = [("Retention","Retention_Pct",0.30),
                 ("Migration","Migration_Pct",0.15),
                 ("Rider","Rider_Pct",0.20),
                 ("CI","CI_Pct",0.20),
                 ("Referral","Referral_Pct",0.15)]
for ai,((pname,pcol,wt),ax) in enumerate(zip(programs_info, axes.flat[:5])):
    # All TLs
    tl_prog = perf_full.groupby(["TL_ID","TL_Name","Month_Year"])[pcol].mean().reset_index()
    for tl_name, grp in tl_prog.groupby("TL_Name"):
        is_bad = "Meera" in tl_name
        ax.plot(MONTHS, grp.set_index("Month_Year").reindex(MONTHS)[pcol],
                linewidth=3 if is_bad else 1.2,
                color=C_RED if is_bad else "#AAAAAA",
                alpha=1 if is_bad else 0.5,
                label="Team Meera Joshi (TL_05)" if is_bad else "_",
                zorder=5 if is_bad else 2)
    # Overall average
    overall = perf_full.groupby("Month_Year")[pcol].mean().reindex(MONTHS)
    ax.plot(MONTHS, overall.values, color=PROG_COLORS[pname],
            linewidth=2, linestyle="--", marker="s", markersize=5, label="Overall Avg")
    ax.set_title(f"{pname}  (weight: {int(wt*100)}%)")
    ax.set_ylabel("Achievement %"); ax.set_ylim(0,120)
    ax.tick_params(axis="x", rotation=30)
    if ai==0: ax.legend(fontsize=8)
    if pname in ("Rider","CI"):
        for mi,m in enumerate(MONTHS):
            if m in UNDERPERFORM_MONTHS:
                ax.axvspan(mi-0.45,mi+0.45, alpha=0.12, color=C_RED)
        ax.text(1.5, 95, "TL_05 drag\nperiod", ha="center", fontsize=8,
                color=C_RED, fontweight="bold")

axes.flat[5].axis("off")
plt.tight_layout()
plt.savefig("images/02-program-achievement-trend.png", dpi=150, bbox_inches="tight")
plt.close(); print("Image 02 saved")

# ── Image 03: TL Performance Matrix ──────────────────────────────────
fig, axes = plt.subplots(1,2, figsize=(18,7))
fig.suptitle("Team Lead Performance Matrix", fontsize=16, color=C_NAVY, fontweight="bold")

tl_summary = (perf_full.groupby(["TL_ID","TL_Name","AM_Name"])
              .agg(Avg_CS=("Consolidated_Score","mean"),
                   Total_Incentive=("Incentive_Amount","sum"),
                   Avg_Rider=("Rider_Pct","mean"),
                   Avg_CI=("CI_Pct","mean"))
              .reset_index().sort_values("Avg_CS"))

colors_tl = [C_RED if v<65 else (C_ORG if v<72 else C_GRN) for v in tl_summary.Avg_CS]
bars = axes[0].barh(tl_summary.TL_Name, tl_summary.Avg_CS, color=colors_tl, edgecolor="white")
axes[0].axvline(perf_full.Consolidated_Score.mean(), color=C_NAVY, linestyle="--",
                linewidth=1.5, label=f"Avg {perf_full.Consolidated_Score.mean():.1f}%")
axes[0].set_title("Avg Consolidated Score by TL"); axes[0].set_xlabel("Consolidated Score %")
axes[0].legend(fontsize=9)
for bar in bars:
    w=bar.get_width()
    axes[0].text(w+0.3,bar.get_y()+bar.get_height()/2, f"{w:.1f}%", va="center", fontsize=9)

# Scatter: Rider vs CI by TL
scatter_colors = [C_RED if "Meera" in nm else C_BLUE for nm in tl_summary.TL_Name]
sc = axes[1].scatter(tl_summary.Avg_Rider, tl_summary.Avg_CI,
                      c=scatter_colors, s=200, zorder=5, edgecolors="white", linewidth=1.5)
for _, row in tl_summary.iterrows():
    is_bad = "Meera" in row.TL_Name
    axes[1].annotate(row.TL_Name.split()[0],
                     xy=(row.Avg_Rider, row.Avg_CI),
                     xytext=(5,5), textcoords="offset points",
                     fontsize=8, color=C_RED if is_bad else "#555555",
                     fontweight="bold" if is_bad else "normal")
axes[1].axhline(perf_full.CI_Pct.mean(), color="#AAAAAA", linestyle=":", linewidth=1)
axes[1].axvline(perf_full.Rider_Pct.mean(), color="#AAAAAA", linestyle=":", linewidth=1)
axes[1].set_xlabel("Avg Rider Attachment %"); axes[1].set_ylabel("Avg CI Attachment %")
axes[1].set_title("Rider vs CI Attachment by TL\n(TL_05 visible as outlier — bottom-left)")
axes[1].text(perf_full.Rider_Pct.mean()+0.5, 22, "TL_05 outlier", color=C_RED,
             fontsize=9, fontweight="bold",
             bbox=dict(boxstyle="round,pad=0.3", facecolor=C_YLW, edgecolor=C_RED))

plt.tight_layout()
plt.savefig("images/03-tl-performance-matrix.png", dpi=150, bbox_inches="tight")
plt.close(); print("Image 03 saved")

# ── Image 04: Underperforming Team Drill-down ─────────────────────────
fig, axes = plt.subplots(1,3, figsize=(18,7))
fig.suptitle("Root Cause Drill-Down: Team Meera Joshi (TL_05) vs All Others",
             fontsize=15, color=C_RED, fontweight="bold")

for ax,prog,title in zip(axes, ["Rider","CI","CS"],
                          ["Rider Attachment %","CI Attachment %","Consolidated Score %"]):
    if prog == "CS":
        t5 = tl5_monthly.CS.values
        ot = all_others.CS.values
    else:
        t5 = getattr(tl5_monthly, prog).values
        ot = getattr(all_others,  prog).values
    x = np.arange(len(MONTHS))
    w = 0.35
    ax.bar(x-w/2, t5, w, label="TL_05 (Meera Joshi)", color=C_RED,   alpha=0.85, edgecolor="white")
    ax.bar(x+w/2, ot, w, label="All Other Teams",       color=C_TEAL,  alpha=0.85, edgecolor="white")
    ax.set_xticks(x); ax.set_xticklabels(MONTHS, rotation=30, fontsize=9)
    ax.set_ylabel(title); ax.set_title(title); ax.legend(fontsize=9)
    # Shade underperform months
    for mi,m in enumerate(MONTHS):
        if m in UNDERPERFORM_MONTHS:
            ax.axvspan(mi-0.5, mi+0.5, alpha=0.07, color=C_RED)
    ax.text(1, max(max(t5),max(ot))*0.95, "Underperform\nwindow", ha="center",
            fontsize=8, color=C_RED, fontweight="bold")

plt.tight_layout()
plt.savefig("images/04-underperforming-team-drilldown.png", dpi=150, bbox_inches="tight")
plt.close(); print("Image 04 saved")

# ── Image 05: Incentive Distribution & AM Rollup ──────────────────────
fig, axes = plt.subplots(1,2, figsize=(16,7))
fig.suptitle("Incentive Distribution & Area Manager Performance",
             fontsize=15, color=C_NAVY, fontweight="bold")

slab_c2 = {"No Incentive":C_RED,"Bronze":"#8D6E63","Silver":"#78909C",
           "Gold":"#F9A825","Platinum":"#5C6BC0"}
bars_s = axes[0].bar(incentive_dist.Incentive_Slab.astype(str),
                      incentive_dist.Count,
                      color=[slab_c2[s] for s in incentive_dist.Incentive_Slab],
                      edgecolor="white", linewidth=1.2)
axes[0].set_title("Incentive Slab Distribution (All Advisors × All Months)")
axes[0].set_ylabel("Count"); axes[0].set_xlabel("Incentive Slab")
for bar in bars_s:
    h=bar.get_height()
    axes[0].text(bar.get_x()+bar.get_width()/2, h+2, str(int(h)),
                 ha="center", fontsize=10, fontweight="bold")

# AM performance
am_colors = [C_BLUE,C_TEAL,C_GRN]
bars_am = axes[1].bar(am_summary.AM_Name, am_summary.Avg_CS,
                       color=am_colors, edgecolor="white", linewidth=1.2, width=0.5)
axes[1].axhline(perf_full.Consolidated_Score.mean(), color=C_RED, linestyle="--",
                linewidth=1.5, label=f"Overall avg {perf_full.Consolidated_Score.mean():.1f}%")
axes[1].set_title("Avg Consolidated Score by Area Manager")
axes[1].set_ylabel("Consolidated Score %"); axes[1].legend(fontsize=9)
axes[1].set_ylim(60,90)
for bar in bars_am:
    h=bar.get_height()
    axes[1].text(bar.get_x()+bar.get_width()/2, h+0.3, f"{h:.1f}%",
                 ha="center", fontsize=11, fontweight="bold")

plt.tight_layout()
plt.savefig("images/05-incentive-am-rollup.png", dpi=150, bbox_inches="tight")
plt.close(); print("Image 05 saved")

# ── Image 06: DAX Formula Bar — Consolidated Score ────────────────────
def make_dax_screenshot(filename, measure_name, dax_code, title):
    fig = plt.figure(figsize=(15, 6))
    fig.patch.set_facecolor("#1E1E1E")
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0,15); ax.set_ylim(0,6); ax.axis("off")

    # Title bar
    title_bar = mpatches.FancyBboxPatch((0,5.3),15,0.7,boxstyle="square,pad=0",
                                         facecolor="#252526",edgecolor="none")
    ax.add_patch(title_bar)
    ax.text(0.3,5.65,f"Power BI Desktop — Data view — Measure: {measure_name}",
            fontsize=10, color="#CCCCCC", va="center")
    ax.text(14.7,5.65,"X", fontsize=12, color="#999999", va="center", ha="right")

    # Formula bar header
    fbar_hdr = mpatches.FancyBboxPatch((0,4.7),15,0.55,boxstyle="square,pad=0",
                                        facecolor="#2D2D30",edgecolor="none")
    ax.add_patch(fbar_hdr)
    ax.text(0.3,4.97,"Measure:", fontsize=9, color="#9CDCFE", va="center")
    ax.text(1.8,4.97,measure_name, fontsize=10, color="#CE9178", va="center", fontweight="bold")
    ax.text(8.0,4.97,"Table: performance_fact", fontsize=9, color="#6A9955", va="center")

    # DAX editor area
    editor = mpatches.FancyBboxPatch((0,0.1),15,4.55,boxstyle="square,pad=0",
                                      facecolor="#1E1E1E",edgecolor="#444444",linewidth=0.5)
    ax.add_patch(editor)

    # Line numbers + code
    lines = dax_code.strip().split("\n")
    for li, line in enumerate(lines[:20]):
        y = 4.45 - li * 0.20
        # Line number
        ax.text(0.25, y, str(li+1), fontsize=8.5, color="#6E6E6E", va="center",
                fontfamily="monospace", ha="right")
        ax.axvline(0.35, ymin=(y-0.09)/6, ymax=(y+0.09)/6, color="#444444", linewidth=0.5)
        # Syntax coloring
        stripped = line.rstrip()
        if stripped.startswith("//"):
            color = "#6A9955"
        elif any(kw in stripped for kw in ["VAR ","RETURN","CALCULATE","DIVIDE","SUM","DATEADD","AVERAGEX","SUMMARIZE","IF","ROUND","NOT","ISBLANK"]):
            color = "#569CD6"
        elif stripped.startswith("    ") and "=" in stripped and not stripped.startswith("    //"):
            color = "#9CDCFE"
        else:
            color = "#D4D4D4"
        ax.text(0.45, y, stripped[:90], fontsize=8.5, color=color, va="center",
                fontfamily="monospace")

    ax.text(7.5, 0.0, title, ha="center", fontsize=9, color="#858585", style="italic")
    plt.savefig(filename, dpi=150, bbox_inches="tight", facecolor="#1E1E1E")
    plt.close()

dax_cs = """// Consolidated Incentive Score — Weighted measure combining all 5 programs
// Weights: Retention 30% | Migration 15% | Rider 20% | CI 20% | Referrals 15%
Consolidated Incentive Score (%) =
VAR RetPct =
    DIVIDE( SUM(performance_fact[Retention_Achieved]),
            SUM(performance_fact[Retention_Target]), 0 )
VAR MigPct =
    DIVIDE( SUM(performance_fact[Migration_Achieved]),
            SUM(performance_fact[Migration_Target]), 0 )
VAR RidPct =
    DIVIDE( SUM(performance_fact[Rider_Achieved]),
            SUM(performance_fact[Rider_Target]), 0 )
VAR CIPct =
    DIVIDE( SUM(performance_fact[CI_Achieved]),
            SUM(performance_fact[CI_Target]), 0 )
VAR RefPct =
    DIVIDE( SUM(performance_fact[Referral_Achieved]),
            SUM(performance_fact[Referral_Target]), 0 )
RETURN
    ROUND(
        (RetPct * 0.30) + (MigPct * 0.15) +
        (RidPct * 0.20) + (CIPct  * 0.20) + (RefPct * 0.15),
        4
    ) * 100"""

make_dax_screenshot("images/06-dax-consolidated-score.png",
                    "Consolidated Incentive Score (%)", dax_cs,
                    "Power BI DAX Measure — Consolidated Incentive Score (weighted average across 5 programs)")
print("Image 06 saved")

dax_mom = """// Month-over-Month Change — requires date_dim marked as Date Table in Power BI
// Uses DATEADD time-intelligence function to shift context by -1 MONTH
Consolidated Score MoM Change =
VAR CurrentScore =
    [Consolidated Incentive Score (%)]
VAR PreviousMonthScore =
    CALCULATE(
        [Consolidated Incentive Score (%)],
        DATEADD(date_dim[Date], -1, MONTH)
    )
RETURN
    IF(
        NOT ISBLANK(PreviousMonthScore) && PreviousMonthScore <> 0,
        DIVIDE(CurrentScore - PreviousMonthScore, PreviousMonthScore, BLANK()),
        BLANK()
    )"""

make_dax_screenshot("images/07-dax-mom-trend.png",
                    "Consolidated Score MoM Change", dax_mom,
                    "Power BI DAX Measure — Month-over-Month Change using DATEADD time intelligence")
print("Image 07 saved")

# ── Image 08: Advisor Drill-down under TL_05 ──────────────────────────
fig, axes = plt.subplots(1,2, figsize=(18,7))
fig.suptitle("Individual Advisor Performance — TL_05 (Meera Joshi) vs Peers",
             fontsize=14, color=C_NAVY, fontweight="bold")

adv_summary = (perf_full.groupby(["Advisor_ID","TL_Name"])
               .agg(Avg_CS=("Consolidated_Score","mean"),
                    Avg_Rider=("Rider_Pct","mean"),
                    Avg_CI=("CI_Pct","mean"),
                    Incentive=("Incentive_Amount","sum"))
               .reset_index())
tl5_advisors = adv_summary[adv_summary.TL_Name=="Meera Joshi"].sort_values("Avg_CS")
other_advisors = adv_summary[adv_summary.TL_Name!="Meera Joshi"]

# TL_05 advisors: all 3 metrics
x  = np.arange(len(tl5_advisors)); w=0.25
ax = axes[0]
ax.bar(x-w, tl5_advisors.Avg_CS,    w, label="Consolidated %", color=C_NAVY,  alpha=0.85)
ax.bar(x,   tl5_advisors.Avg_Rider, w, label="Rider %",        color=C_ORG,   alpha=0.85)
ax.bar(x+w, tl5_advisors.Avg_CI,    w, label="CI %",           color=C_RED,   alpha=0.85)
ax.axhline(other_advisors.Avg_CS.mean(),    color=C_NAVY,  linestyle="--", linewidth=1, alpha=0.7)
ax.axhline(other_advisors.Avg_Rider.mean(), color=C_ORG,   linestyle="--", linewidth=1, alpha=0.7)
ax.axhline(other_advisors.Avg_CI.mean(),    color=C_RED,   linestyle="--", linewidth=1, alpha=0.7)
ax.set_xticks(x); ax.set_xticklabels(tl5_advisors.Advisor_ID, rotation=30, fontsize=9)
ax.set_title("TL_05 Advisors — 6-Month Average Performance\n(dashed = peer benchmark)")
ax.legend(fontsize=9); ax.set_ylabel("Achievement %"); ax.set_ylim(0,100)

# Incentive amount comparison: TL_05 vs others
ax2 = axes[1]
cat_labels = ["TL_05 Total 6M", "Others Per-Advisor Avg 6M", "TL_05 Per-Advisor Avg 6M"]
total_tl5 = tl5_data.Incentive_Amount.sum()
avg_other_per_adv = (perf_full[perf_full.TL_ID!="TL_05"].groupby("Advisor_ID").Incentive_Amount.sum().mean())
avg_tl5_per_adv   = (tl5_data.groupby("Advisor_ID").Incentive_Amount.sum().mean())
vals = [total_tl5, avg_other_per_adv, avg_tl5_per_adv]
colors_b = [C_RED, C_TEAL, C_ORG]
bars_i = ax2.bar(cat_labels, vals, color=colors_b, edgecolor="white", width=0.5)
ax2.set_title("Incentive Earned: TL_05 vs Peer Benchmark")
ax2.set_ylabel("INR (Rs.)"); ax2.tick_params(axis="x", rotation=15)
for bar in bars_i:
    h=bar.get_height()
    ax2.text(bar.get_x()+bar.get_width()/2, h+300, f"Rs.{int(h):,}",
             ha="center", fontsize=10, fontweight="bold")
gap = avg_other_per_adv - avg_tl5_per_adv
ax2.annotate(f"Gap: Rs.{int(gap):,}/advisor\n(lost incentive)",
             xy=(2, avg_tl5_per_adv), xytext=(1.5, avg_other_per_adv*0.8),
             arrowprops=dict(arrowstyle="->", color=C_RED),
             color=C_RED, fontsize=10, fontweight="bold")

plt.tight_layout()
plt.savefig("images/08-advisor-drill-down.png", dpi=150, bbox_inches="tight")
plt.close(); print("Image 08 saved")

print("\n=== PROJECT 1 STATS ===")
for k,v in overall_stats.items():
    print(f"  {k}: {v}")
print("Project 1 complete.")
